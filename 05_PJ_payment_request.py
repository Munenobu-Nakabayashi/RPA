# 【経費】支払依頼書
import logging
# --- 2021.11.29 ADD Start
import openpyxl

import sys
# import pyautogui    # --- 2021.11.30 Add New（※証憑ファイル提出が発生する為）
import pyautogui as pg  # --- UPDATE 2021.12.23
pg.FAILSAFE = False     # --- UPDATE 2021.12.23
from datetime import datetime     # --- 2021.11.30 Add New（本日日付取得の為）
import pyperclip    # 2021.12.01 クリップボードコピー追加 (Nakabayashi)
import gc           # 2021.12.01 ガーベジコレクター追加 (Nakabayashi)
import concurrent.futures   # --- ADD 2021.12.02
import re                   # --- ADD 2021.12.02

from selenium import webdriver
driver = webdriver.Chrome(
    executable_path="C:/Users/digiworker_biz_02/PycharmProjects/pythonProject/Chrome/chromedriver.exe"
)
# --- 2021.11.29 ADD End
import paymentdestination as pmd     # --- ADD NEW 2021.12.24 PM
import item as itm                   # --- ADD NEW 2021.12.25 AM
import department as dpt             # --- ADD NEW 2021.12.25 PM
# from chrome_driver_dl import get_latest_driver    # 2021.11.27 DELETE
from common import *

# Gets or creates a logger
logger = logging.getLogger("02")
# set log level
logger.setLevel(logging.INFO)
# define file handler and set formatter
file_handler = logging.FileHandler("../Logs/05_payment_request.log")
formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")
file_handler.setFormatter(formatter)
# add file handler to logger
logger.addHandler(file_handler)

# driver = webdriver.Chrome(get_latest_driver())    # --- DELETE 2021.11.30
driver.get(MFZC_URL)
# driver.fullscreen_window()

file_handler.setFormatter(formatter)
# add file handler to logger
logger.addHandler(file_handler)

def reloadBrowser():        # --- ADD 2021.12.02 Chrome原因不明不調につき急遽追記 - Start
    try:
        time.sleep(2)
        driver.refresh()
        time.sleep(2)
    except Exceptio as e:
        logger.info("Chromeブラウザーをリロードできず。ドライバーを確認せよ")
        print("Chromeブラウザーをリロードできず。ドライバーを確認せよ")
        # driver.quit()
    finally:
        pass                # --- ADD 2021.12.02 Chrome原因不明不調につき急遽追記 - End

def click_element(web_el):
    my_sleep_click(driver.find_element_by_xpath(web_el))

class PaymentRequest:
    def __init__(self,
                 wkUserId,
                 wkPassWd,
                 wkAim,
                 wkWishDate,
                 wkInvoiceNo,
                 wkPaymentDestCd,
                 wkRemarksA,
                 wkRemarksB,
                 wkItem,
                 wkBearDeptId,
                 wkContent,
                 wkCost,
                 wkTaxIncludedAmount,
                 wkConsumptionTax,
                 wkWithholdingTaxCalclationTargetClassification,
                 wkWithholdingTaxAmount,
                 wkBurdenDepartmentCode,
                 wkAnalysisCode,
                 wkVoucherPDF0,
                 wkVoucherPDF1,
                 wkVoucherPDF2,
                 wkVoucherPDF3,
                 wkVoucherPDF4,
                 wkVoucherPDF5,
                 wkVoucherPDF6,
                 wkVoucherPDF7,
                 wkVoucherPDF8,
                 wkVoucherPDF9,
                 wkIid01,  # --- ADD NEW 2022.01.11 スコープ変更に伴う追加 x 16 （配列厳しい為、直打ち）
                 wkDocno01,
                 wkFormname01,
                 wkCreated01,
                 wkAname01,
                 wkUid01,
                 wkEmpcd01,
                 wkEname01,
                 wkDeptid01,
                 wkDfullname01,
                 wkPost01,
                 wkTs01,
                 wkIid02,
                 wkDocno02,
                 wkFormname02,
                 wkCreated02,
                 wkAname02,
                 wkUid02,
                 wkEmpcd02,
                 wkEname02,
                 wkDeptid02,
                 wkDfullname02,
                 wkPost02,
                 wkTs02,
                 wkIid03,
                 wkDocno03,
                 wkFormname03,
                 wkCreated03,
                 wkAname03,
                 wkUid03,
                 wkEmpcd03,
                 wkEname03,
                 wkDeptid03,
                 wkDfullname03,
                 wkPost03,
                 wkTs03,
                 wkIid04,
                 wkDocno04,
                 wkFormname04,
                 wkCreated04,
                 wkAname04,
                 wkUid04,
                 wkEmpcd04,
                 wkEname04,
                 wkDeptid04,
                 wkDfullname04,
                 wkPost04,
                 wkTs04,
                 wkIid05,
                 wkDocno05,
                 wkFormname05,
                 wkCreated05,
                 wkAname05,
                 wkUid05,
                 wkEmpcd05,
                 wkEname05,
                 wkDeptid05,
                 wkDfullname05,
                 wkPost05,
                 wkTs05,
                 wkIid06,
                 wkDocno06,
                 wkFormname06,
                 wkCreated06,
                 wkAname06,
                 wkUid06,
                 wkEmpcd06,
                 wkEname06,
                 wkDeptid06,
                 wkDfullname06,
                 wkPost06,
                 wkTs06,
                 wkIid07,
                 wkDocno07,
                 wkFormname07,
                 wkCreated07,
                 wkAname07,
                 wkUid07,
                 wkEmpcd07,
                 wkEname07,
                 wkDeptid07,
                 wkDfullname07,
                 wkPost07,
                 wkTs07,
                 wkIid08,
                 wkDocno08,
                 wkFormname08,
                 wkCreated08,
                 wkAname08,
                 wkUid08,
                 wkEmpcd08,
                 wkEname08,
                 wkDeptid08,
                 wkDfullname08,
                 wkPost08,
                 wkTs08,
                 wkIid09,
                 wkDocno09,
                 wkFormname09,
                 wkCreated09,
                 wkAname09,
                 wkUid09,
                 wkEmpcd09,
                 wkEname09,
                 wkDeptid09,
                 wkDfullname09,
                 wkPost09,
                 wkTs09,
                 wkIid10,
                 wkDocno10,
                 wkFormname10,
                 wkCreated10,
                 wkAname10,
                 wkUid10,
                 wkEmpcd10,
                 wkEname10,
                 wkDeptid10,
                 wkDfullname10,
                 wkPost10,
                 wkTs10,
                 wkIid11,
                 wkDocno11,
                 wkFormname11,
                 wkCreated11,
                 wkAname11,
                 wkUid11,
                 wkEmpcd11,
                 wkEname11,
                 wkDeptid11,
                 wkDfullname11,
                 wkPost11,
                 wkTs11,
                 wkIid12,
                 wkDocno12,
                 wkFormname12,
                 wkCreated12,
                 wkAname12,
                 wkUid12,
                 wkEmpcd12,
                 wkEname12,
                 wkDeptid12,
                 wkDfullname12,
                 wkPost12,
                 wkTs12,
                 wkIid13,
                 wkDocno13,
                 wkFormname13,
                 wkCreated13,
                 wkAname13,
                 wkUid13,
                 wkEmpcd13,
                 wkEname13,
                 wkDeptid13,
                 wkDfullname13,
                 wkPost13,
                 wkTs13,
                 wkIid14,
                 wkDocno14,
                 wkFormname14,
                 wkCreated14,
                 wkAname14,
                 wkUid14,
                 wkEmpcd14,
                 wkEname14,
                 wkDeptid14,
                 wkDfullname14,
                 wkPost14,
                 wkTs14,
                 wkIid15,
                 wkDocno15,
                 wkFormname15,
                 wkCreated15,
                 wkAname15,
                 wkUid15,
                 wkEmpcd15,
                 wkEname15,
                 wkDeptid15,
                 wkDfullname15,
                 wkPost15,
                 wkTs15,
                 wkIid16,
                 wkDocno16,
                 wkFormname16,
                 wkCreated16,
                 wkAname16,
                 wkUid16,
                 wkEmpcd16,
                 wkEname16,
                 wkDeptid16,
                 wkDfullname16,
                 wkPost16,
                 wkTs16
                 ):
            self.userId = wkUserId                    # ユーザーID            →I列[9]:EMPCD
            self.passWd = wkPassWd                    # パスワード            →ナシ
            self.aim = wkAim                          # 目的                 →CA列[79]:SS13
            self.wishDate = wkWishDate                # 支払希望日            →AV列[48]:WISH_DATE
            self.invoiceNo = wkInvoiceNo              # 請求書番号            →E列[5]:DOCNO
            self.paymentDestCd = wkPaymentDestCd      # 支払先コード          →BK列[63]:CUSTCD
            self.remarksA = wkRemarksA                  # 備考A              →CA列[79]:SS13
            self.remarksB = wkRemarksB                  # 備考B              →CH列[86]:SS20
            self.item = wkItem                        # 品目                 →CB列[80]:SS14
            self.bearDeptId = wkBearDeptId            # 支払部門コード         →R列[18]:BEARDEPTID
            self.content = wkContent                  # 内容                 →CE列[83]:SS17
            self.cost = wkCost                        # 費用                 →AO列[41]:PAY_PRICE
            self.taxIncludedAmount = wkTaxIncludedAmount    # 税込金額        →AB列[28]:SUBTOTAL_1
            self.consumptionTax = wkConsumptionTax          # 消費税          →？
            self.withholdingTaxCalclationTargetClassification\
                = wkWithholdingTaxCalclationTargetClassification    # 源泉税計算対象区分 →AX列[50]:PAY_GEN_FLG
            self.withholdingTaxAmount = wkWithholdingTaxAmount  # 源泉税額              →？
            self.burdenDepartmentCode = wkBurdenDepartmentCode  # 負担部門コード         →P列[16]:BEARDEPTIC_C
            self.analysisCode = wkAnalysisCode         # 分析コード              →？
            self.voucherPDF0 = wkVoucherPDF0           # PDF0                  →DE列[109]
            self.voucherPDF1 = wkVoucherPDF1           # PDF1                  →DF列[110]
            self.voucherPDF2 = wkVoucherPDF2           # PDF2                  →DG列[111]
            self.voucherPDF3 = wkVoucherPDF3           # PDF3                  →DH列[112]
            self.voucherPDF4 = wkVoucherPDF4           # PDF4                  →DI列[113]
            self.voucherPDF5 = wkVoucherPDF5           # PDF5                  →DJ列[114]
            self.voucherPDF6 = wkVoucherPDF6           # PDF6                  →DK列[115]
            self.voucherPDF7 = wkVoucherPDF7           # PDF7                  →DL列[116]
            self.voucherPDF8 = wkVoucherPDF8           # PDF8                  →DM列[117]
            self.voucherPDF9 = wkVoucherPDF9           # PDF9                  →DN列[118]
            self.iid01 = wkIid01
            self.docno01 = wkDocno01
            self.formname01 = wkFormname01
            self.createdd01 = wkCreated01
            self.aname01 = wkAname01
            self.uid01 = wkUid01
            self.empcd01 = wkEmpcd01
            self.ename01 = wkEname01
            self.deptid01 = wkDeptid01
            self.dfullname01 = wkDfullname01
            self.post01 = wkPost01
            self.ts01 = wkTs01
            self.iid02 = wkIid02
            self.docno02 = wkDocno02
            self.forname02 = wkFormname02
            self.created02 = wkCreated02
            self.aname02 = wkAname02
            self.uid02 = wkUid02
            self.empcd02 = wkEmpcd02
            self.ename02 = wkEname02
            self.deptid02 = wkDeptid02
            self.dfullname02 = wkDfullname02
            self.post02 = wkPost02
            self.ts02 = wkTs02
            self.iid03 = wkIid03
            self.docno03 = wkDocno03
            self.fornmame03 = wkFormname03
            self.created03 = wkCreated03
            self.aname03 = wkAname03
            self.uid03 = wkUid03
            self.empcd03 = wkEmpcd03
            self.ename03 = wkEname03
            self.deptid03 = wkDeptid03
            self.dfullname03 = wkDfullname03
            self.post03 = wkPost03
            self.ts03 = wkTs03
            self.iid04 = wkIid04
            self.docno04 = wkDocno04
            self.formname04 = wkFormname04
            self.created04 = wkCreated04
            self.aname04 = wkAname04
            self.uid04 = wkUid04
            self.empcd04 = wkEmpcd04
            self.ename04 = wkEname04
            self.deptid04 = wkDeptid04
            self.dfullname04 = wkDfullname04
            self.post04 = wkPost04
            self.ts04 = wkTs04
            self.iid05 = wkIid05
            self.docno05 = wkDocno05
            self.formname05 = wkFormname05
            self.created05 = wkCreated05
            self.aname05 = wkAname05
            self.uid05 = wkUid05
            self.empcd05 = wkEmpcd05
            self.ename05 = wkEname05
            self.deptid05 = wkDeptid05
            self.dfullname05 = wkDfullname05
            self.post05 = wkPost05
            self.ts05 = wkTs05
            self.iid06 = wkIid06
            self.docno06 = wkDocno06
            self.formname06 = wkFormname06
            self.created06 = wkCreated06
            self.aname06 = wkAname06
            self.uid06 = wkUid06
            self.empcd06 = wkEmpcd06
            self.ename06 = wkEname06
            self.deptid06 = wkDeptid06
            self.dfullname06 = wkDfullname06
            self.post06 = wkPost06
            self.ts06 = wkTs06
            self.iid07 = wkIid07
            self.docno07 = wkDocno07
            self.forname07 = wkFormname07
            self.created07 = wkCreated07
            self.aname07 = wkAname07
            self.uid07 = wkUid07
            self.empcd07 = wkEmpcd07
            self.ename07 = wkEname07
            self.deptid07 = wkDeptid07
            self.dfullname07 = wkDfullname07
            self.post07 = wkPost07
            self.ts07 = wkTs07
            self.iid08 = wkIid08
            self.docno08 = wkDocno08
            self.forname08 = wkFormname08
            self.created08 = wkCreated08
            self.aname08 = wkAname08
            self.uid08 = wkUid08
            self.empcd08 = wkEmpcd08
            self.ename08 = wkEname08
            self.deptid08 = wkDeptid08
            self.dfullname08 = wkDfullname08
            self.post08 = wkPost08
            self.ts08 = wkTs08
            self.iid09 = wkIid09
            self.docno09 = wkDocno09
            self.formname09 = wkFormname09
            self.created09 = wkCreated09
            self.aname09 = wkAname09
            self.uid09 = wkUid09
            self.empcd09 = wkEmpcd09
            self.ename09 = wkEname09
            self.deptid09 = wkDeptid09
            self.dfullname09 = wkDfullname09
            self.post09 = wkPost09
            self.ts09 = wkTs09
            self.iid10 = wkIid10
            self.docno10 = wkDocno10
            self.formname10 = wkFormname10
            self.created10 = wkCreated10
            self.aname10 = wkAname10
            self.uid10 = wkUid10
            self.empcd10 = wkEmpcd10
            self.ename10 = wkEname10
            self.deptid10 = wkDeptid10
            self.dfullname10 = wkDfullname10
            self.post10 = wkPost10
            self.ts = wkTs10
            self.iid11 = wkIid11
            self.docno11 = wkDocno11
            self.fromname11 = wkFormname11
            self.created11 = wkCreated11
            self.aname11 = wkAname11
            self.uid11 = wkUid11
            self.empcd11 = wkEmpcd11
            self.ename11 = wkEname11
            self.deptid11 = wkDeptid11
            self.dfullname11 = wkDfullname11
            self.post11 = wkPost11
            self.ts11 = wkTs11
            self.iid12 = wkIid12
            self.docno12 = wkDocno12
            self.fromname12 = wkFormname12
            self.created12 = wkCreated12
            self.aname12 = wkAname12
            self.uid12 = wkUid12
            self.empcd12 = wkEmpcd12
            self.ename12 = wkEname12
            self.deptid12 = wkDeptid12
            self.dfullname12 = wkDfullname12
            self.post12 = wkPost12
            self.ts12 = wkTs12
            self.iid13 = wkIid13
            self.docno13 = wkDocno13
            self.formname13 = wkFormname13
            self.created13 = wkCreated13
            self.aname13 = wkAname13
            self.uid13 = wkUid13
            self.empcd13 = wkEmpcd13
            self.ename13 = wkEname13
            self.deptid13 = wkDeptid13
            self.dfullname13 = wkDfullname13
            self.post13 = wkPost13
            self.ts13 = wkTs13
            self.iid14 = wkIid14
            self.docno14 = wkDocno14
            self.formname14 = wkFormname14
            self.created14 = wkCreated14
            self.aname14 = wkAname14
            self.uid14 = wkUid14
            self.empcd14 = wkEmpcd14
            self.ename14 = wkEname14
            self.deptid14 = wkDeptid14
            self.dfullname14 = wkDfullname14
            self.post14 = wkPost14
            self.ts14 = wkTs14
            self.iid15 = wkIid15
            self.docno15 = wkDocno15
            self.formname15 = wkFormname15
            self.created15 = wkCreated15
            self.aname15 = wkAname15
            self.uid15 = wkUid15
            self.empcd15 = wkEmpcd15
            self.ename15 = wkEname15
            self.deptid15 = wkDeptid15
            self.dfullname15 = wkDfullname15
            self.post15 = wkPost15
            self.ts15 = wkTs15
            self.iid16 = wkIid16
            self.docno16 = wkDocno16
            self.fromname16 = wkFormname16
            self.created16 = wkCreated16
            self.aname16 = wkAname16
            self.uid16 = wkUid16
            self.empcd16 = wkEmpcd16
            self.ename16 = wkEname16
            self.deptid16 = wkDeptid16
            self.dfullname16 = wkDfullname16
            self.post16 = wkPost16
            self.ts16 = wkTs16

    def getSignInInfo(self):
        yield 100
        yield self.userID
        yield self.passWd
    
    def getAim(self):
        if self.aim is not None:
            return self.aim
        else:
            return '移行元データにおいてブランク。'   # --- エントリー必須項目である為、回避措置を講じた
    
    def getWishDate(self):
        # if self.wishDate != 0 and self.wishDate is not None:  # --- 2021.11.30 PEND Start
        #     return self.wishDate
        # else:
        #     return datetime.today()                           # --- 2021.11.30 PEND End
        today = datetime.now()              # --- 支払希望日がカラである為、本日日付で代替。
        return today.strftime('%Y-%m-%d')

    def getInvoiceNo(self):
        if self.invoiceNo is not None and self.invoiceNo != "":      # --- UPDATE 2022.01.06
            return self.invoiceNo
        else:
            return "N/A"

    def getPaymentDestCd(self):
        return self.paymentDestCd         # --- 2021.12.24 VALID

    def getRemarks(self):
        # if self.remarks is not None:                  # --- 2021.12.21 DELETE Start
        #    return self.remarks
        # else:
        #    # return ''
        #    return '元データにおいて備考は空白であった。'    # --- 2021.12.21 DELETE End
        if self.remarksA is None:
            self.remarksA = ""
        if self.remarksB is None:
            self.remarksB = ""
        yield self.remarksA     # --- 2021.012.21 UPDATE
        yield self.remarksB     # --- 2021.012.21 UPDATE
        approvalInfo = ""
        if self.iid01 is not None and self.iid01 != "":
            approvalInfo += f'決裁:{self.aname01} 従業員ID:{self.uid01} 従業員番号:{self.empcd01} 従業員指名:{self.ename01} ;'
        if self.iid02 is not None and self.iid02 != "":
            approvalInfo += f'決裁:{self.aname02} 従業員ID:{self.uid02} 従業員番号:{self.empcd02} 従業員指名:{self.ename02} ;'
        if self.iid03 is not None and self.iid03 != "":
            approvalInfo += f'決裁:{self.aname03} 従業員ID:{self.uid03} 従業員番号:{self.empcd03} 従業員指名:{self.ename03} ;'
        if self.iid04 is not None and self.iid04 != "":
            approvalInfo += f'決裁:{self.aname04} 従業員ID:{self.uid04} 従業員番号:{self.empcd04} 従業員指名:{self.ename04} ;'
        if self.iid05 is not None and self.iid05 != "":
            approvalInfo += f'決裁:{self.aname05} 従業員ID:{self.uid05} 従業員番号:{self.empcd05} 従業員指名:{self.ename05} ;'
        if self.iid06 is not None and self.iid06 != "":
            approvalInfo += f'決裁:{self.aname06} 従業員ID:{self.uid06} 従業員番号:{self.empcd06} 従業員指名:{self.ename06} ;'
        if self.iid07 is not None and self.iid07 != "":
            approvalInfo += f'決裁:{self.aname07} 従業員ID:{self.uid07} 従業員番号:{self.empcd07} 従業員指名:{self.ename07} ;'
        if self.iid08 is not None and self.iid08 != "":
            approvalInfo += f'決裁:{self.aname08} 従業員ID:{self.uid08} 従業員番号:{self.empcd08} 従業員指名:{self.ename08} ;'
        if self.iid09 is not None and self.iid09 != "":
            approvalInfo += f'決裁:{self.aname09} 従業員ID:{self.uid09} 従業員番号:{self.empcd09} 従業員指名:{self.ename09} ;'
        if self.iid10 is not None and self.iid10 != "":
            approvalInfo += f'決裁:{self.aname10} 従業員ID:{self.uid10} 従業員番号:{self.empcd10} 従業員指名:{self.ename10} ;'
        if self.iid11 is not None and self.iid11 != "":
            approvalInfo += f'決裁:{self.aname11} 従業員ID:{self.uid11} 従業員番号:{self.empcd11} 従業員指名:{self.ename11} ;'
        if self.iid12 is not None and self.iid12 != "":
            approvalInfo += f'決裁:{self.aname12} 従業員ID:{self.uid12} 従業員番号:{self.empcd12} 従業員指名:{self.ename12} ;'
        if self.iid13 is not None and self.iid13 != "":
            approvalInfo += f'決裁:{self.aname13} 従業員ID:{self.uid13} 従業員番号:{self.empcd13} 従業員指名:{self.ename13} ;'
        if self.iid14 is not None and self.iid14 != "":
            approvalInfo += f'決裁:{self.aname14} 従業員ID:{self.uid14} 従業員番号:{self.empcd14} 従業員指名:{self.ename14} ;'
        if self.iid15 is not None and self.iid15 != "":
            approvalInfo += f'決裁:{self.aname15} 従業員ID:{self.uid15} 従業員番号:{self.empcd15} 従業員指名:{self.ename15} ;'
        if self.iid16 is not None and self.iid16 != "":
            approvalInfo += f'決裁:{self.aname16} 従業員ID:{self.uid16} 従業員番号:{self.empcd16} 従業員指名:{self.ename16} ;'
        yield approvalInfo

    def getItem(self):  # 品目
        return self.item           # --- 2021.12.25 REMAKE --- 外部処理に任せる

    def getContent(self):
        if self.content != None and self.content != "":
            return self.content   # --- UPDATE 2022.01.06
        else:
            return "N/A"
    
    def getCost(self):
        if self.cost != None and self.cost != "":
            return self.cost
        else:
            return 0

    def getTaxIncludedAmount(self):
        return self.taxIncludedAmount
    
    def getConsumptionTax(self):
        if self.consumptionTax is None:
            return 0
        else:
            return self.consumptionTax

    def getWithholdingTaxCalclationTargetClassification(self):
        if self.withholdingTaxCalclationTargetClassification != 0:
            return "税込金額"
        else:
            return "税抜金額"

    def getWithholdingTaxAmount(self):
        if self.withholdingTaxAmount is None:
            return 0
        else:
            return self.withholdingTaxAmount

    def getBurdenDepartmentCode(self):
        return self.burdenDepartmentCode
    
    def getAnalysisCode(self):
        if self.analysisCode is not None:
            return self.analysisCode
        else:
            return '006'    # 2021.11.30 分析コード不明につき、「その他」オールマイティーコード006を返す
    
    def getVoucherPDFs(self):   #PDFその他証憑ファイルの絶対パス情報のまとまりをyieldで返す→next()を使って配列へ格納する
        yield self.voucherPDF0
        yield self.voucherPDF1
        yield self.voucherPDF2
        yield self.voucherPDF3
        yield self.voucherPDF4
        yield self.voucherPDF5
        yield self.voucherPDF6
        yield self.voucherPDF7
        yield self.voucherPDF8
        yield self.voucherPDF9

def signin_procedure(tenantId, empId, passWd):
    logger.info(            # --- ADD NEW 共通部品 20212.01.04
        "テナントID: " + str(tenantId) + " 職員番号: " + str(empId) + " 共通パスワード: " + str(passWd)
    )
    # テナントID = 100
    print(str(tenantId))
    tenant_id_textbox = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[1]/input"
    )
    print(str(empId))
    tenant_id_textbox.send_keys(
        str(tenantId)
    )
    # 職員コード
    employee_cd_textbox = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[2]/input"
    )
    employee_cd_textbox.send_keys(
        str(empId)
    )
    # パスワード（※2021.11.17時点、方針未決）
    password_textbox = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[3]/input"
    )
    password_textbox.send_keys(
        str(passWd)
    )
    # ログインボタンをクリック
    logger.info(            # --- ADD NEW 共通部品 20212.01.04
        "ログインボタンクリック"
    )
    click_element(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[4]/button"
    )
    time.sleep(2)

def sign_out_procedure():    # 2021.12.15
    # ①申請完了後の「閉じる」ボタンをクリック
    logger.info(  # --- ADD NEW 共通部品 20212.01.04
        "閉じるボタンクリック"
    )
    close_button = driver.find_element_by_xpath(
        # "/html/body/div[1]/div/div/form/div/button[4]"        # --- CHANGE 2021.12.24
        "/html/body/div[1]/div/div/form/div/button[3]"          # --- CHANGE 2021.12.24
    )
    my_sleep_click(close_button)

    driver.implicitly_wait(3)                           # --- ADD NEW 2021.12.22
    driver.switch_to.window(driver.window_handles[0])   # --- ADD NEW 2021.12.22 要素ゼロのウィンドウへ制御を戻す！

    # ②画面右上「サインアウト」ボタン（非常出口アイコン）をクリック
    logger.info(  # --- ADD NEW 共通部品 20212.01.04
        "サインアウトアイコンクリック"
    )
    sign_out_button = driver.find_element_by_xpath(
        "/html/body/div[2]/nav/div/div[2]/ul/li[6]/form/a"
    )
    my_sleep_click(sign_out_button)
    time.sleep(2)
    # ③【ログイン】ボタンをラストにクリック
    logger.info(  # --- ADD NEW 共通部品 20212.01.04
        "サインイン画面へ戻るボタンクリック"
    )
    sign_in_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    )
    my_sleep_click(sign_in_button)
    time.sleep(2)

# def judgeItemCode(wkItem, wkBearDeptCd):
#    # ※品目は飲食のみである
#    if wkItem == '飲食':
#        # 部課名に開発を含有する
#        # ﾏｰｹﾃｨﾝｸﾞ部 PIS課 / ﾏｰｹﾃｨﾝｸﾞ部ﾌﾟﾛﾓｰｼｮﾝ課 / 開発部 Web開発課 # 開発部共通 # i-FILTER課 / 開発１課 # m-FILTER課 / 開発２課
#        # iFBC/i-ﾌｨﾙﾀｰ課 / 開発３課 # FinalCode課 / 開発４課 # ﾆｭｰﾌﾟﾛﾀﾞｸﾄ課 / 開発５課 # Iﾗﾎﾞ開発課 # 研究開発課# 海外開発課
#         if wkBearDeptCd == 671 \
#             or wkBearDeptCd == 1000 \
#             or wkBearDeptCd == 1119 \
#             or wkBearDeptCd == 1120 \
#             or wkBearDeptCd == 1121 \
#             or wkBearDeptCd == 1122 \
#             or wkBearDeptCd == 1123 \
#             or wkBearDeptCd == 1294 \
#             or wkBearDeptCd == 1900 \
#             or wkBearDeptCd == 1901:
#                 return 'i10r-50201'
#        else:
#             return 'i10r-70802'
#     else:
#         logger.info("品目コード該当なし")
#         sys.exit()

def entry_procedure():

    reloadBrowser()     # --- 2021.12.02 Chromeブラウザー不具合につき追加
    # click_element("/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button")
    # 2021.11.27 UPDATE Start
    logger.info(  # --- ADD NEW 共通部品 20212.01.04
        "PJ.xlsxファイルオープン開始"
    )
    book = openpyxl.load_workbook("C:\XLSX\XM_SLIP\PJ.xlsx")    # Book名がPJ
    sheet = book['PJ']                                          # Sheet名がPJ
    logger.info(  # --- ADD NEW 共通部品 20212.01.04
        "PJ.xlsxファイルオープン完了"
    )

    i = 2   # 1行目は見出し行
    with concurrent.futures.ProcessPoolExecutor() as executor:
        while sheet.cell(row=i, column=1).value != '' or sheet.cell(row=i, column=1).value is not None:
            pr = PaymentRequest(sheet.cell(row=i, column=9).value,      # ※L列[]CREATORを緊急避難で取得
                                '0Nu4M0%4N0',                           # ※共通パスワード。2021.11.29時点未定！ 決まったら書直す事！
                                sheet.cell(row=i, column=79).value,     # 目的
                                sheet.cell(row=i, column=48).value,     # 支払希望日
                                sheet.cell(row=i, column=5).value,      # 請求書番号
                                sheet.cell(row=i, column=63).value,     # 支払先コード
                                sheet.cell(row=i, column=79).value,     # 備考A(CA)
                                sheet.cell(row=i, column=86).value,     # 備考B(CH)
                                sheet.cell(row=i, column=80).value,     # 品目
                                sheet.cell(row=i, column=18).value,     # 負担部門コード（※直接エントリーに用いない）
                                # sheet.cell(row=i, column=10).value,   # 負担部門コード（※直接エントリーに用いない）
                                sheet.cell(row=i, column=83).value,     # 内容
                                sheet.cell(row=i, column=41).value,     # 費用
                                sheet.cell(row=i, column=28).value,     # 税込金額
                                sheet.cell(row=i, column=54).value,     # ※消費税は2021.11.29時点未定！ 決まったら書直す事！
                                sheet.cell(row=i, column=50).value,     # 源泉税計算対象区分
                                sheet.cell(row=i, column=256).value,    # 源泉税額
                                sheet.cell(row=i, column=16).value,     # 負担部門コード
                                sheet.cell(row=i, column=81).value,     # ※分析コードは2021.11.29時点未定！ 決まったら書直す事！
                                sheet.cell(row=i, column=109).value,    # 証憑PDFファイル絶対パスその1
                                sheet.cell(row=i, column=110).value,    # 証憑PDFファイル絶対パスその2
                                sheet.cell(row=i, column=111).value,    # 証憑PDFファイル絶対パスその3
                                sheet.cell(row=i, column=112).value,    # 証憑PDFファイル絶対パスその4
                                sheet.cell(row=i, column=113).value,    # 証憑PDFファイル絶対パスその5
                                sheet.cell(row=i, column=114).value,    # 証憑PDFファイル絶対パスその6
                                sheet.cell(row=i, column=115).value,    # 証憑PDFファイル絶対パスその7
                                sheet.cell(row=i, column=116).value,    # 証憑PDFファイル絶対パスその8
                                sheet.cell(row=i, column=117).value,    # 証憑PDFファイル絶対パスその9
                                sheet.cell(row=i, column=118).value,     # 証憑PDFファイル絶対パスその10
                                sheet.cell(row=i, column=131).value,  # IID --- (1)
                                sheet.cell(row=i, column=132).value,  # DOCNO
                                sheet.cell(row=i, column=133).value,  # FORMNAME
                                sheet.cell(row=i, column=134).value,  # CREATED
                                sheet.cell(row=i, column=135).value,  # ANAME
                                sheet.cell(row=i, column=136).value,  # UID
                                sheet.cell(row=i, column=137).value,  # EMPCD
                                sheet.cell(row=i, column=138).value,  # ENAME
                                sheet.cell(row=i, column=139).value,  # DEPTID
                                sheet.cell(row=i, column=140).value,  # DFULLNAME
                                sheet.cell(row=i, column=141).value,  # POST
                                sheet.cell(row=i, column=142).value,  # TS
                                sheet.cell(row=i, column=144).value,  # IID --- (2)
                                sheet.cell(row=i, column=145).value,  # DOCNO
                                sheet.cell(row=i, column=146).value,  # FORMNAME
                                sheet.cell(row=i, column=147).value,  # CREATED
                                sheet.cell(row=i, column=148).value,  # ANAME
                                sheet.cell(row=i, column=149).value,  # UID
                                sheet.cell(row=i, column=150).value,  # EMPCD
                                sheet.cell(row=i, column=151).value,  # ENAME
                                sheet.cell(row=i, column=152).value,  # DEPTID
                                sheet.cell(row=i, column=153).value,  # DFULLNAME
                                sheet.cell(row=i, column=154).value,  # POST
                                sheet.cell(row=i, column=155).value,  # TS
                                sheet.cell(row=i, column=157).value,  # IID --- (3)
                                sheet.cell(row=i, column=158).value,  # DOCNO
                                sheet.cell(row=i, column=159).value,  # FORMNAME
                                sheet.cell(row=i, column=160).value,  # CREATED
                                sheet.cell(row=i, column=161).value,  # ANAME
                                sheet.cell(row=i, column=162).value,  # UID
                                sheet.cell(row=i, column=163).value,  # EMPCD
                                sheet.cell(row=i, column=164).value,  # ENAME
                                sheet.cell(row=i, column=165).value,  # DEPTID
                                sheet.cell(row=i, column=166).value,  # DFULLNAME
                                sheet.cell(row=i, column=167).value,  # POST
                                sheet.cell(row=i, column=168).value,  # TS
                                sheet.cell(row=i, column=170).value,  # IID --- (4)
                                sheet.cell(row=i, column=171).value,  # DOCNO
                                sheet.cell(row=i, column=172).value,  # FORMNAME
                                sheet.cell(row=i, column=173).value,  # CREATED
                                sheet.cell(row=i, column=174).value,  # ANAME
                                sheet.cell(row=i, column=175).value,  # UID
                                sheet.cell(row=i, column=176).value,  # EMPCD
                                sheet.cell(row=i, column=177).value,  # ENAME
                                sheet.cell(row=i, column=178).value,  # DEPTID
                                sheet.cell(row=i, column=179).value,  # DFULLNAME
                                sheet.cell(row=i, column=180).value,  # POST
                                sheet.cell(row=i, column=181).value,  # TS
                                sheet.cell(row=i, column=183).value,  # IID --- (5)
                                sheet.cell(row=i, column=184).value,  # DOCNO
                                sheet.cell(row=i, column=185).value,  # FORMNAME
                                sheet.cell(row=i, column=186).value,  # CREATED
                                sheet.cell(row=i, column=187).value,  # ANAME
                                sheet.cell(row=i, column=188).value,  # UID
                                sheet.cell(row=i, column=189).value,  # EMPCD
                                sheet.cell(row=i, column=190).value,  # ENAME
                                sheet.cell(row=i, column=191).value,  # DEPTID
                                sheet.cell(row=i, column=192).value,  # DFULLNAME
                                sheet.cell(row=i, column=193).value,  # POST
                                sheet.cell(row=i, column=194).value,  # TS
                                sheet.cell(row=i, column=196).value,  # IID --- (6)
                                sheet.cell(row=i, column=197).value,  # DOCNO
                                sheet.cell(row=i, column=198).value,  # FORMNAME
                                sheet.cell(row=i, column=199).value,  # CREATED
                                sheet.cell(row=i, column=200).value,  # ANAME
                                sheet.cell(row=i, column=201).value,  # UID
                                sheet.cell(row=i, column=202).value,  # EMPCD
                                sheet.cell(row=i, column=203).value,  # ENAME
                                sheet.cell(row=i, column=204).value,  # DEPTID
                                sheet.cell(row=i, column=205).value,  # DFULLNAME
                                sheet.cell(row=i, column=206).value,  # POST
                                sheet.cell(row=i, column=207).value,  # TS
                                sheet.cell(row=i, column=209).value,  # IID --- (7)
                                sheet.cell(row=i, column=210).value,  # DOCNO
                                sheet.cell(row=i, column=211).value,  # FORMNAME
                                sheet.cell(row=i, column=212).value,  # CREATED
                                sheet.cell(row=i, column=213).value,  # ANAME
                                sheet.cell(row=i, column=214).value,  # UID
                                sheet.cell(row=i, column=215).value,  # EMPCD
                                sheet.cell(row=i, column=216).value,  # ENAME
                                sheet.cell(row=i, column=217).value,  # DEPTID
                                sheet.cell(row=i, column=218).value,  # DFULLNAME
                                sheet.cell(row=i, column=219).value,  # POST
                                sheet.cell(row=i, column=220).value,  # TS
                                sheet.cell(row=i, column=222).value,  # IID --- (8)
                                sheet.cell(row=i, column=223).value,  # DOCNO
                                sheet.cell(row=i, column=224).value,  # FORMNAME
                                sheet.cell(row=i, column=225).value,  # CREATED
                                sheet.cell(row=i, column=226).value,  # ANAME
                                sheet.cell(row=i, column=227).value,  # UID
                                sheet.cell(row=i, column=228).value,  # EMPCD
                                sheet.cell(row=i, column=229).value,  # ENAME
                                sheet.cell(row=i, column=230).value,  # DEPTID
                                sheet.cell(row=i, column=231).value,  # DFULLNAME
                                sheet.cell(row=i, column=232).value,  # POST
                                sheet.cell(row=i, column=233).value,  # TS
                                sheet.cell(row=i, column=235).value,  # IID --- (9)
                                sheet.cell(row=i, column=236).value,  # DOCNO
                                sheet.cell(row=i, column=237).value,  # FORMNAME
                                sheet.cell(row=i, column=238).value,  # CREATED
                                sheet.cell(row=i, column=239).value,  # ANAME
                                sheet.cell(row=i, column=240).value,  # UID
                                sheet.cell(row=i, column=241).value,  # EMPCD
                                sheet.cell(row=i, column=242).value,  # ENAME
                                sheet.cell(row=i, column=243).value,  # DEPTID
                                sheet.cell(row=i, column=244).value,  # DFULLNAME
                                sheet.cell(row=i, column=245).value,  # POST
                                sheet.cell(row=i, column=246).value,  # TS
                                sheet.cell(row=i, column=248).value,  # IID --- (10)
                                sheet.cell(row=i, column=249).value,  # DOCNO
                                sheet.cell(row=i, column=250).value,  # FORMNAME
                                sheet.cell(row=i, column=251).value,  # CREATED
                                sheet.cell(row=i, column=252).value,  # ANAME
                                sheet.cell(row=i, column=253).value,  # UID
                                sheet.cell(row=i, column=254).value,  # EMPCD
                                sheet.cell(row=i, column=255).value,  # ENAME
                                sheet.cell(row=i, column=256).value,  # DEPTID
                                sheet.cell(row=i, column=257).value,  # DFULLNAME
                                sheet.cell(row=i, column=258).value,  # POST
                                sheet.cell(row=i, column=259).value,  # TS
                                sheet.cell(row=i, column=261).value,  # IID --- (11)
                                sheet.cell(row=i, column=262).value,  # DOCNO
                                sheet.cell(row=i, column=263).value,  # FORMNAME
                                sheet.cell(row=i, column=264).value,  # CREATED
                                sheet.cell(row=i, column=265).value,  # ANAME
                                sheet.cell(row=i, column=266).value,  # UID
                                sheet.cell(row=i, column=267).value,  # EMPCD
                                sheet.cell(row=i, column=268).value,  # ENAME
                                sheet.cell(row=i, column=269).value,  # DEPTID
                                sheet.cell(row=i, column=270).value,  # DFULLNAME
                                sheet.cell(row=i, column=271).value,  # POST
                                sheet.cell(row=i, column=272).value,  # TS
                                sheet.cell(row=i, column=274).value,  # IID --- (12)
                                sheet.cell(row=i, column=275).value,  # DOCNO
                                sheet.cell(row=i, column=276).value,  # FORMNAME
                                sheet.cell(row=i, column=277).value,  # CREATED
                                sheet.cell(row=i, column=278).value,  # ANAME
                                sheet.cell(row=i, column=279).value,  # UID
                                sheet.cell(row=i, column=280).value,  # EMPCD
                                sheet.cell(row=i, column=281).value,  # ENAME
                                sheet.cell(row=i, column=282).value,  # DEPTID
                                sheet.cell(row=i, column=283).value,  # DFULLNAME
                                sheet.cell(row=i, column=284).value,  # POST
                                sheet.cell(row=i, column=285).value,  # TS
                                sheet.cell(row=i, column=287).value,  # IID --- (13)
                                sheet.cell(row=i, column=288).value,  # DOCNO
                                sheet.cell(row=i, column=289).value,  # FORMNAME
                                sheet.cell(row=i, column=290).value,  # CREATED
                                sheet.cell(row=i, column=291).value,  # ANAME
                                sheet.cell(row=i, column=292).value,  # UID
                                sheet.cell(row=i, column=293).value,  # EMPCD
                                sheet.cell(row=i, column=294).value,  # ENAME
                                sheet.cell(row=i, column=295).value,  # DEPTID
                                sheet.cell(row=i, column=296).value,  # DFULLNAME
                                sheet.cell(row=i, column=297).value,  # POST
                                sheet.cell(row=i, column=298).value,  # TS
                                sheet.cell(row=i, column=300).value,  # IID --- (14)
                                sheet.cell(row=i, column=301).value,  # DOCNO
                                sheet.cell(row=i, column=302).value,  # FORMNAME
                                sheet.cell(row=i, column=303).value,  # CREATED
                                sheet.cell(row=i, column=304).value,  # ANAME
                                sheet.cell(row=i, column=305).value,  # UID
                                sheet.cell(row=i, column=306).value,  # EMPCD
                                sheet.cell(row=i, column=307).value,  # ENAME
                                sheet.cell(row=i, column=308).value,  # DEPTID
                                sheet.cell(row=i, column=309).value,  # DFULLNAME
                                sheet.cell(row=i, column=310).value,  # POST
                                sheet.cell(row=i, column=311).value,  # TS
                                sheet.cell(row=i, column=313).value,  # IID --- (15)
                                sheet.cell(row=i, column=314).value,  # DOCNO
                                sheet.cell(row=i, column=315).value,  # FORMNAME
                                sheet.cell(row=i, column=316).value,  # CREATED
                                sheet.cell(row=i, column=317).value,  # ANAME
                                sheet.cell(row=i, column=318).value,  # UID
                                sheet.cell(row=i, column=319).value,  # EMPCD
                                sheet.cell(row=i, column=320).value,  # ENAME
                                sheet.cell(row=i, column=321).value,  # DEPTID
                                sheet.cell(row=i, column=322).value,  # DFULLNAME
                                sheet.cell(row=i, column=323).value,  # POST
                                sheet.cell(row=i, column=324).value,  # TS
                                sheet.cell(row=i, column=326).value,  # IID --- (16)
                                sheet.cell(row=i, column=327).value,  # DOCNO
                                sheet.cell(row=i, column=328).value,  # FORMNAME
                                sheet.cell(row=i, column=329).value,  # CREATED
                                sheet.cell(row=i, column=330).value,  # ANAME
                                sheet.cell(row=i, column=331).value,  # UID
                                sheet.cell(row=i, column=332).value,  # EMPCD
                                sheet.cell(row=i, column=333).value,  # ENAME
                                sheet.cell(row=i, column=334).value,  # DEPTID
                                sheet.cell(row=i, column=335).value,  # DFULLNAME
                                sheet.cell(row=i, column=336).value,  # POST
                                sheet.cell(row=i, column=337).value  # TS
            )

            # 引数: 左からテナントID、職員番号、パスワード
            gen = pr.getSignInInfo()
            # signin_procedure(gen.__next__(), gen.__next__(), gen.__next__())   <----- サインイン方法確定後有効化せよ！！
            gen = None

            # 下記は検証時における、テナントID直下の【ログイン】ボタン。本稼働時に上記サインイン処理が有効化されたら無効化せよ！！
            click_element(
                "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
            )

            click_element(      # 起票ボタンをクリック
                "/html/body/div[2]/div[1]/ul[1]/li[4]/a"
            )

            logger.info(  # --- ADD NEW 共通部品 20212.01.04
                "●【経費】支払依頼書ラジオボタン選択"
            )
            click_element(      # ●支払依頼書を選択
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[5]/label[6]/div/div[1]/input"
            )

            driver.implicitly_wait(5)   # 2021.11.27 --- 10から5へ減数

            click_element(      # 次へボタンをクリック
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[3]/button"
            )

            driver.implicitly_wait(5)   # 2021.11.27 --- 10から5へ減数
            driver.switch_to.window(driver.window_handles[1])
            # create_button = driver.find_element_by_xpath('//*[text()="起票する"]')    # 先任者において既に無効化されていた

            create_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[3]/button[2]"
            )
            my_sleep_click(create_button)

            # PURPOSE = "目的"  --- DELETE 2021.11.29 - 前任者による定数エントリー
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "目的: " + str(pr.getAim())
            )
            purpose_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div/input"
            )
            # purpose_field.send_keys(PURPOSE)      # --- DELETE 2021.11.29 - 前任者による定数エントリー
            purpose_field.send_keys(pr.getAim())    # --- UPDATE 2021.11.29 - ゲッターより値を返す

            # 支払希望日（支払期日）
            # PAYMENT_DESIRE_DATE = "2021-08-02"    # --- DELETE 2021.11.29 - 前任者による定数エントリー
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "希望日: " + str(pr.getWishDate())
            )
            payment_desire_date_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div[1]/div/input"
            )
            # payment_desire_date_field.send_keys(PAYMENT_DESIRE_DATE)      # --- DELETE 2021.11.29 - 前任者による定数エントリー
            payment_desire_date_field.send_keys(pr.getWishDate())           # --- UPDATE 2021.11.29 - ゲッターより値を返す
                                                                            # --- ※当該フィールドは全数が空白値である

            # TODO
            # 請求書番号    # --- 前任者によるサボタージュ
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "請求書番号: " + str(pr.getInvoiceNo())
            )
            invoice_no_textbox = driver.find_element_by_xpath(              # --- ADD NEW 2021.11.29 - Start
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div[2]/input"
            )
            invoice_no_textbox.send_keys(pr.getInvoiceNo())                 # --- ADD NEW 2021.11.29 - End
            invoice_no_textbox.send_keys(
                Keys.TAB            # --- 消しゴムアイコンのボタン
                + Keys.TAB          # --- iframeオープンボタン
                + Keys.ENTER        # --- iframeオープンボタンへ改行キーを打鍵
            )
            # Search Payee  # --- 前任者によるコメント
            # PAYEE = "支払先"      # --- DELETE 2021.11.29 - 前任者による定数エントリー
            # payee_search_button = driver.find_element_by_xpath(   # --- DELETE 2021.11.29 Start
            #     "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[4]/div/div/span/button[2]"
            # )
            # my_sleep_click(payee_search_button)

            # time.sleep(5)

            # 取引先の検索
            # time.sleep(5)
            # driver.implicitly_wait(10)    # --- DELETE 2021.11.29 Start
            iframe = driver.find_element_by_xpath('//*[@id="zk_iFrameDialog"]/iframe')
            driver.switch_to.frame(iframe)

            deal_code_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[1]/input"
            )
            # EMPLOYEE_CODE = "10011"                   # --- DELETE 2021.11.29 - 前任者による定数エントリー
            # deal_code_field.send_keys(EMPLOYEE_CODE)  # --- DELETE 2021.11.29 - 前任者による定数エントリー
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "支払先コード: " + str(pr.getPaymentDestCd())
            )
            # deal_code_field.send_keys(pr.getPaymentDestCd())    # --- UPDATE 2021.11.29 支払先コードエントリー
            deal_code_field.send_keys(                              # --- UPDATE 2021.12.24 外部へ処理を任せる --- Start
                pmd.returnPaymentDestCd(pr.getPaymentDestCd())
            )                                                       # --- UPDATE 2021.12.24 外部へ処理を任せる --- End
            # deal_code_field.send_keys(Keys.ENTER)
            deal_code_field.send_keys(  # --- ADD 2021.11.29 Start
                Keys.TAB
                + Keys.TAB
                + Keys.ENTER
            )                           # --- ADD 2021.11.29 End
            time.sleep(2)   # 引数5であったが2へ減数
            found_row = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[1]/div[3]/label/div"           # --- DELETE 2021.11.29
                # "/html/body/div[1]/div/div[1]/div[3]/label/div/div[1]"      # --- UPDATE 2021.11.29
            )
            my_sleep_click(found_row)
            driver.switch_to.default_content()

            # NOTE_TEXT = "テスト"      # --- DELETE 2021.11.29 - 前任者による定数エントリー
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "備考: " + str(pr.getRemarks())
            )
            note_text_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[6]/div/textarea"
            )
            # note_text_field.send_keys(NOTE_TEXT)          # --- DELETE 2021.11.29
            # note_text_field.send_keys(pr.getRemarks())      # --- UPDATE 2021.11.29
            gen = pr.getRemarks()                           # --- UPDATE 2021.12.21
            crlf = '\r\n'                                   # --- UPDATE 2021.12.21
            note_text_field.send_keys(
                    str(gen.__next__()) + crlf + str(gen.__next__()) + crlf + str(gen.__next__())
            )    # 三つ目は決裁情報
            time.sleep(3)                                   # 5から3へ減数
            gen = None
            note_text_field.send_keys(                      # --- ADD NEW 2021.11.30 Start
                Keys.TAB
                + Keys.TAB
                + Keys.TAB
                + Keys.ENTER
            )                                               # --- ADD NEW 2021.11.30 End

            time.sleep(2)   # 必須の待機時間。無くすと実行時例外が発生する。
            detail_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[2]/div/div/span/button[2]"
            )
            my_sleep_click(detail_button)

            # Switch to frame
            driver.implicitly_wait(3)       # 10から3へ減数
            iframe = driver.find_element_by_xpath('//*[@id="zk_iFrameDialog"]/iframe')
            driver.switch_to.frame(iframe)
            time.sleep(2)                   # 5から2へ減数

            # Find details
            # ITEM_CODE = "i10r-80101"  # --- DELETE 2021.11.29 - 前任者による定数エントリー
            # gen = pr.getItem()      # yieldからGenerater経由で値を受取る    # --- DELETE 2021.12.25
            gen = itm.returnItem(pr.getItem())      # --- ADD NEW 2021.12.25 外部処理経由
            wkXpath = gen.__next__()        # 表示テーブルの何段目をクリックするか、当該段数を返する
            wkItemCode = gen.__next__()
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "段数: " + str(wkXpath) + " 品目コード: " + str(wkItemCode)
            )
            # wkItemCode = judgeItemCode(gen.__next__(), gen.__next__())
            # --- ADD NEW 品目コード選別関数（※飲食かつ●●の条件が正しいか要確認）
            # item_code_fld.send_keys(ITEM_CODE + Keys.ENTER)   # --- DELETE 2021.11.29
            # item_code_fld.send_keys(wkItemCode)               # --- UPDATE 2021.11.29
            item_code_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[1]/input"
            )
            item_code_fld.send_keys(
                wkItemCode
            )                     # --- UPDATE 2021.12.20

            item_code_fld.send_keys(        # --- UPDATE 2021.11.30 Start
                Keys.TAB
                + Keys.TAB
                + Keys.TAB
                + Keys.ENTER    # 虫眼鏡アイコン検索ボタンをクリック
            )                               # --- UPDATE 2021.11.30 End

            time.sleep(3)   # 引数5を3へ減数

            # すべて開くボタンをクリック
            open_all = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[1]/div[2]/button"
            )
            my_sleep_click(open_all)
            time.sleep(2)

            # item_row = driver.find_element_by_xpath(          # --- Total Abolition 2021.12.25 --- Start
            #    # "/html/body/div[1]/div/div[1]/div[4]/div/div/div/label/div"
            #    "/html/body/div[1]/div/div[1]/div[4]/div/div/div/div/label/div/div[1]"
            #    # str(wkXpath)      # --- UPDATE 2021.12.20 XPAHTを返す（※2, 3, 4, 5, 7段のパターンが存在）
            # )                                                 # --- Total Abolition 2021.12.25 --- Start

            if wkXpath == 2:        # 2段目
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/label[1]/div/div[1]"
                )
            elif wkXpath == 3:        # 3段目
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/div/label/div/div[1]"
                )
            elif wkXpath == 4:        # 4段目
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/div/div/label/div/div[1]"
                )
            elif wkXpath == 5:        # 5段目
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/div/div/div/label/div/div[1]"
                )
            elif wkXpath == 7:        # 7段目
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/div[2]/div/label/div/div[1]"
                )
            my_sleep_click(item_row)
            time.sleep(2)

            # 元のウィンドウへ制御を戻す
            driver.switch_to.default_content()

            time.sleep(2)   # 引数5を2へ減数
            # CONTENT = "内容"  # --- DELETE 2021.11.30 前任者による定数エントリ－
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "内容: " + str(pr.getContent())
            )
            content_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[3]/div/input"
            )
            # content_fld.send_keys(CONTENT)        # --- DELETE 2021.11.30
            content_fld.send_keys(pr.getContent())  # --- UPDATE 2021.11.30 --- UPDATE 2022.01.06

            # COST = "1000"     # --- DELETE 2021.11.30 前任者による定数エントリ－
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "費用: " + str(pr.getCost())
            )
            cost_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[1]/input"
            )
            # cost_fld.send_keys(COST)              # --- DELETE 2021.11.30
            cost_fld.send_keys(pr.getCost())        # 費用

            # 以降、前任者がスルーした①税込金額、②消費税、③源泉税計算対象区分、④源泉税額、⑤負担部門、⑥分析コードをエントリー【開始】
            tax_included_amount_textbox = driver.find_element_by_xpath(     # 税込金額 --- 2021.11.30 ADD
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[2]/input"
            )
            # tax_included_amount_textbox.clear()
            # time.sleep(2)
            # tax_included_amount_textbox.send_keys(pr.getTaxIncludedAmount()) # PEND 2021.11.30

            # consumption_tax_textbox = driver.find_element_by_xpath(         # 消費税 --- 2021.11.30 ADD
            #    "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[4]/input"
            # )
            # if pr.getConsumptionTax() != 0:         # ！！ 本当にクリアして値をエントリーするのか、要確認
            #    consumption_tax_textbox.clear()
            # time.sleep(2)
            # consumption_tax_textbox.send_keys(
            #        pr.getConsumptionTax()
            # )

            logger.info(  # --- ADD NEW 個別 20212.01.04
                "源泉税計算対象区分: " + str(pr.getWithholdingTaxCalclationTargetClassification())
            )
            withholding_tax_calclation_target_classification_textbox = driver.find_element_by_xpath(    # 源泉税計算対象区分
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[5]/div[1]/select"
            )
            withholding_tax_calclation_target_classification_textbox.send_keys(
                pr.getWithholdingTaxCalclationTargetClassification()
            )

            logger.info(  # --- ADD NEW 個別 20212.01.04
                "源泉税額: " + str(pr.getWithholdingTaxAmount())
            )
            withholding_tax_amount_textbox = driver.find_element_by_xpath(  # 源泉税額
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[5]/div[2]/div/input"
            )
            withholding_tax_amount_textbox.send_keys(pr.getWithholdingTaxAmount())

            withholding_tax_amount_textbox.send_keys(
                Keys.TAB        # 電卓アイコンボタン
                + Keys.TAB      # 負担部門の消しゴムアイコンボタン
                + Keys.TAB      # 負担部門のiframe起動ボタンへカーソル移動
                + Keys.ENTER    # iframe起動ボタンを改行キーで打鍵
            )

            driver.implicitly_wait(3)
            # 制御をiframeへ渡す
            iframe = driver.find_element_by_xpath('//*[@id="zk_iFrameDialog"]/iframe')
            driver.switch_to.frame(iframe)
            time.sleep(2)

            logger.info(  # --- ADD NEW 個別 20212.01.04
                "支払部署コード: " + str(pr.getBurdenDepartmentCode())
            )
            burden_department_code_textbox = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[1]/input"
            )
            # burden_department_code_textbox.send_keys(pr.getBurdenDepartmentCode())    # --- INVALID 2021.12.25
            burden_department_code_textbox.send_keys(                                   # --- ADD NEW 2021.12.25 --- Start
                dpt.returnDepartment(pr.getBurdenDepartmentCode())
            )                                                                           # --- ADD NEW 2021.12.25 --- End
            burden_department_code_textbox.send_keys(
                Keys.TAB
                + Keys.TAB
                + Keys.ENTER
            )
            time.sleep(2)

            burden_department_table_record = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[1]/div[3]/label/div/div[1]"
            )
            my_sleep_click(burden_department_table_record)
            time.sleep(2)

            # 元のウィンドウへ制御を戻す
            driver.switch_to.default_content()

            analysis_code_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[6]/div[2]/div/span/button[2]"
            )
            my_sleep_click(analysis_code_button)

            driver.implicitly_wait(3)
            # 制御をiframeへ渡す
            iframe = driver.find_element_by_xpath('//*[@id="zk_iFrameDialog"]/iframe')
            driver.switch_to.frame(iframe)
            time.sleep(2)

            logger.info(  # --- ADD NEW 個別 20212.01.04
                "分析コードコード（ママ）: " + str(pr.getAnalysisCode())
            )
            analysis_code_textbox = driver.find_element_by_xpath(   # 分析コードコード（ママ）
                "/html/body/div[1]/div/form/div/div[1]/input"
            )
            analysis_code_textbox.send_keys(pr.getAnalysisCode())

            analysis_code_textbox.send_keys(
                Keys.TAB        # 分析コード名テキストボックス
                + Keys.TAB      # 検索ボタンへカーソルが移動する
                + Keys.ENTER    # 検索ボタンへ改行キーを打鍵
            )
            time.sleep(2)

            analysis_code_table = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[1]/div[3]/label/div/div[1]"
            )
            my_sleep_click(analysis_code_table)
            time.sleep(2)

            # 元のウィンドウへ制御を戻す
            driver.switch_to.default_content()
            # 以上、前任者がスルーした①税込金額、②消費税、③源泉税計算対象区分、④源泉税額、￥⑤負担部門、⑥分析コードをエントリー【終了】

            confirm_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[5]/div/button[1]"
            )
            my_sleep_click(confirm_button)
            driver.save_screenshot(f'C:/Users/digiworker_biz_02/PycharmProjects/pythonProject/ScreenShot/PJ/{i}_確定ボタンクリック.png')

            # No attach file for this application <--- 前任者による誤認。証憑ファイル提出は存在する。 --- 2021.11.30
            gen = pr.getVoucherPDFs()
            submit_voucher_files(gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__()
            )                                   # --- ADD NEW 2021.11.30 - 2021.12.01

            final_register_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[11]/button[1]"
            )
            click_submit_button(driver, logger, "Validation", final_register_button)
            # --- 2022.01.05 ADD NEW 【申請】ボタンクリック後、却下される可能性がある為、スクリーンショットを取得、保存する
            driver.save_screenshot(f'C:/Users/digiworker_biz_02/PycharmProjects/pythonProject/ScreenShot/EV/{i}_申請ボタンクリック.png')

            # --- ADD NEW 2021.12.23 通知チェックボックスのレ点を解除
            notificationCheckBox = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/form/div[3]/label/input[1]"
            )
            my_sleep_click(notificationCheckBox)

            comment_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/form/textarea"                                 # --- RESTORE 2021.12.24
            )
            comment_field.send_keys(MIGRATION_COMMENT)  # --- ※前任者による備考テキストボックスへのコメント追記。本番実行時は削除せよ

            logger.info(  # --- ADD NEW 個別 20212.01.04
                "実行ボタンクリック"
            )
            final_submit_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/form/div[4]/button[1]"
            )
            my_sleep_click(final_submit_button)

            # TODO uncomment below after testing
            # click_main(driver, logger, 'Submission', final_register_button)   # <--- 左記は不可

            # サインアウト処理
            sign_out_procedure()
        
            # ガーベジコレクター
            # if i % 100 == 0:  --- INVALID 2022.01.05
            gc.collect()
            # PJシートの次行へ移動
            i += 1

def submit_voucher_files(vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9):     # --- ADD NEW 2021.11.30
    logger.info(            # --- ADD NEW 共通部品 20212.01.04
        "添付ファイル - 選択ボタンクリック"
    )
    attach_file_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div[2]/div/form/div[5]/div/ul/li[3]/a"
    )
    my_sleep_click(attach_file_button)
    # 配列へ格納 --- ※証憑はPDFのみとは限らない
    va = [vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9]

    j = 0   # ゼロオリジン
    if va[j] != "N/A":
        while va[j] is not None and j < 10:
            if j == 0:  # 2回目以降は不要である為
                select_file_button = driver.find_element_by_xpath(
                # 添付ファイルタブに所在する添付ファイル-選択ボタン - 2021.11.15 コメント記入
                    "/html/body/div[1]/div[2]/div/form/div[5]/div/div/div["
                    "3]/div/div/div/div[2]/div[2]/div/span/button"
                )
                my_sleep_click(select_file_button)

            logger.info(  # --- ADD NEW 共通部品 20212.01.04
                "証憑ファイルパス: " + va[j]
            )
            upload_file_button = driver.find_element_by_xpath(
                # 次画面にある「添付ファイルを選択してください」→選択ボタンをクリック
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[2]/div[2]"
            )
            upload_file_button.click()
            time.sleep(2)
            pyperclip.copy(va[j])                      # Clip Boardへコピーし、
            pg.hotkey('ctrl', 'v')              # Pasteする。
            pg.press("enter")
            time.sleep(2)
            add_green_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[3]/div/a"
            )
            my_sleep_click(add_green_button)
            time.sleep(2)
            # 右隣セルの値へ移動
            j += 1
    else:
        logger.info(  # --- ADD NEW 共通部品 20212.01.04
            "対象となる証憑ファイルは無し"
        )

    # 証憑提出後、確定ボタンクリック
    if j != 0:  # --- ADD NEW 2021.12.24 --- 証憑が一つ以上あった場合のみ確定ボタンをクリック
        logger.info(  # --- ADD NEW 共通部品 20212.01.04
            "添付ファイルタブ - 確定ボタンクリック"
        )
        final_register_button = driver.find_element_by_xpath(
            # "/html/body/div[1]/div[2]/div/form/div[11]/button[1]"
            "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"
        )
        my_sleep_click(final_register_button)

def main():
    # エントリー手続き
    entry_procedure()
    logger.info("Robot completed")

main()