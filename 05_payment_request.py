# 【経費】支払依頼書
import logging
# --- 2021.11.29 ADD Start
import openpyxl

import sys
# import pyautogui    # --- 2021.11.30 Add New（※証憑ファイル提出が発生する為）
import pyautogui as pg  # --- UPDATE 2021.12.23
pg.FAILSAFE = False     # --- UPDATE 2021.12.23
from datetime import datetime     # --- 2021.11.30 Add New（本日日付取得の為）
import pyperclip    # 2021.12.01 クリップボードコピー追加 (Nakabayashi)
import gc           # 2021.12.01 ガーベジコレクター追加 (Nakabayashi)
import concurrent.futures   # --- ADD 2021.12.02
import re                   # --- ADD 2021.12.02

from selenium import webdriver
driver = webdriver.Chrome(
    executable_path="C:/Users/digiworker_biz_02/PycharmProjects/pythonProject/Chrome/chromedriver.exe"
)
# --- 2021.11.29 ADD End
import paymentdestination as pmd     # --- ADD NEW 2021.12.24 PM
import item as itm                   # --- ADD NEW 2021.12.25 AM
import department as dpt             # --- ADD NEW 2021.12.25 PM
# from chrome_driver_dl import get_latest_driver    # 2021.11.27 DELETE
from common import *

# Gets or creates a logger
logger = logging.getLogger("02")
# set log level
logger.setLevel(logging.INFO)
# define file handler and set formatter
file_handler = logging.FileHandler("../Logs/05_payment_request.log")
formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")
file_handler.setFormatter(formatter)
# add file handler to logger
logger.addHandler(file_handler)

# driver = webdriver.Chrome(get_latest_driver())    # --- DELETE 2021.11.30
driver.get(MFZC_URL)
# driver.fullscreen_window()

file_handler.setFormatter(formatter)
# add file handler to logger
logger.addHandler(file_handler)

def reloadBrowser():        # --- ADD 2021.12.02 Chrome原因不明不調につき急遽追記 - Start
    try:
        time.sleep(2)
        driver.refresh()
        time.sleep(2)
    except Exceptio as e:
        logger.info("Chromeブラウザーをリロードできず。ドライバーを確認せよ")
        print("Chromeブラウザーをリロードできず。ドライバーを確認せよ")
        # driver.quit()
    finally:
        pass                # --- ADD 2021.12.02 Chrome原因不明不調につき急遽追記 - End

def click_element(web_el):
    my_sleep_click(driver.find_element_by_xpath(web_el))

class PaymentRequest:
    def __init__(self,
                 wkUserId,
                 wkPassWd,
                 wkAim,
                 wkWishDate,
                 wkInvoiceNo,
                 wkPaymentDestCd,
                 wkRemarksA,
                 wkRemarksB,
                 wkItem,
                 wkBearDeptId,
                 wkContent,
                 wkCost,
                 wkTaxIncludedAmount,
                 wkConsumptionTax,
                 wkWithholdingTaxCalclationTargetClassification,
                 wkWithholdingTaxAmount,
                 wkBurdenDepartmentCode,
                 wkAnalysisCode,
                 wkVoucherPDF0,
                 wkVoucherPDF1,
                 wkVoucherPDF2,
                 wkVoucherPDF3,
                 wkVoucherPDF4,
                 wkVoucherPDF5,
                 wkVoucherPDF6,
                 wkVoucherPDF7,
                 wkVoucherPDF8,
                 wkVoucherPDF9
            ):
            self.userId = wkUserId                    # ユーザーID            →I列[9]:EMPCD
            self.passWd = wkPassWd                    # パスワード            →ナシ
            self.aim = wkAim                          # 目的                 →CA列[79]:SS13
            self.wishDate = wkWishDate                # 支払希望日            →AV列[48]:WISH_DATE
            self.invoiceNo = wkInvoiceNo              # 請求書番号            →E列[5]:DOCNO
            self.paymentDestCd = wkPaymentDestCd      # 支払先コード          →BK列[63]:CUSTCD
            self.remarksA = wkRemarksA                  # 備考A              →CA列[79]:SS13
            self.remarksB = wkRemarksB                  # 備考B              →CH列[86]:SS20
            self.item = wkItem                        # 品目                 →CB列[80]:SS14
            self.bearDeptId = wkBearDeptId            # 支払部門コード         →R列[18]:BEARDEPTID
            self.content = wkContent                  # 内容                 →CE列[83]:SS17
            self.cost = wkCost                        # 費用                 →AO列[41]:PAY_PRICE
            self.taxIncludedAmount = wkTaxIncludedAmount    # 税込金額        →AB列[28]:SUBTOTAL_1
            self.consumptionTax = wkConsumptionTax          # 消費税          →？
            self.withholdingTaxCalclationTargetClassification\
                = wkWithholdingTaxCalclationTargetClassification    # 源泉税計算対象区分 →AX列[50]:PAY_GEN_FLG
            self.withholdingTaxAmount = wkWithholdingTaxAmount  # 源泉税額              →？
            self.burdenDepartmentCode = wkBurdenDepartmentCode  # 負担部門コード         →P列[16]:BEARDEPTIC_C
            self.analysisCode = wkAnalysisCode         # 分析コード              →？
            self.voucherPDF0 = wkVoucherPDF0           # PDF0                  →DE列[109]
            self.voucherPDF1 = wkVoucherPDF1           # PDF1                  →DF列[110]
            self.voucherPDF2 = wkVoucherPDF2           # PDF2                  →DG列[111]
            self.voucherPDF3 = wkVoucherPDF3           # PDF3                  →DH列[112]
            self.voucherPDF4 = wkVoucherPDF4           # PDF4                  →DI列[113]
            self.voucherPDF5 = wkVoucherPDF5           # PDF5                  →DJ列[114]
            self.voucherPDF6 = wkVoucherPDF6           # PDF6                  →DK列[115]
            self.voucherPDF7 = wkVoucherPDF7           # PDF7                  →DL列[116]
            self.voucherPDF8 = wkVoucherPDF8           # PDF8                  →DM列[117]
            self.voucherPDF9 = wkVoucherPDF9           # PDF9                  →DN列[118]

    def getSignInInfo(self):
        yield 100
        yield self.userID
        yield self.passWd
    
    def getAim(self):
        if self.aim is not None:
            return self.aim
        else:
            return '移行元データにおいてブランクであった。'   # --- エントリー必須項目である為、回避措置を講じた
    
    def getWishDate(self):
        # if self.wishDate != 0 and self.wishDate is not None:  # --- 2021.11.30 PEND Start
        #     return self.wishDate
        # else:
        #     return datetime.today()                           # --- 2021.11.30 PEND End
        today = datetime.now()              # --- 支払希望日がカラである為、本日日付で代替。
        return today.strftime('%Y-%m-%d')

    def getInvoiceNo(self):
        return self.invoiceNo

    def getPaymentDestCd(self):
        return self.paymentDestCd         # --- 2021.12.24 VALID

    def getRemarks(self):
        # if self.remarks is not None:                  # --- 2021.12.21 DELETE Start
        #    return self.remarks
        # else:
        #    # return ''
        #    return '元データにおいて備考は空白であった。'    # --- 2021.12.21 DELETE End
        if self.remarksA is None:
            self.remarksA = ""
        if self.remarksB is None:
            self.remarksB = ""
        yield self.remarksA     # --- 2021.012.21 UPDATE
        yield self.remarksB     # --- 2021.012.21 UPDATE
    
    def getItem(self):  # 品目
        return self.item           # --- 2021.12.25 REMAKE --- 外部処理に任せる

    def getContent(self):
        return self.content
    
    def getCost(self):
        return self.cost

    def getTaxIncludedAmount(self):
        return self.taxIncludedAmount
    
    def getConsumptionTax(self):
        if self.consumptionTax is None:
            return 0
        else:
            return self.consumptionTax

    def getWithholdingTaxCalclationTargetClassification(self):
        if self.withholdingTaxCalclationTargetClassification != 0:
            return "税込金額"
        else:
            return "税抜金額"

    def getWithholdingTaxAmount(self):
        if self.withholdingTaxAmount is None:
            return 0
        else:
            return self.withholdingTaxAmount

    def getBurdenDepartmentCode(self):
        return self.burdenDepartmentCode
    
    def getAnalysisCode(self):
        if self.analysisCode is not None:
            return self.analysisCode
        else:
            return '006'    # 2021.11.30 分析コード不明につき、「その他」オールマイティーコード006を返す
    
    def getVoucherPDFs(self):   #PDFその他証憑ファイルの絶対パス情報のまとまりをyieldで返す→next()を使って配列へ格納する
        yield self.voucherPDF0
        yield self.voucherPDF1
        yield self.voucherPDF2
        yield self.voucherPDF3
        yield self.voucherPDF4
        yield self.voucherPDF5
        yield self.voucherPDF6
        yield self.voucherPDF7
        yield self.voucherPDF8
        yield self.voucherPDF9

def signin_procedure(tenantId, empId, passWd):
    # テナントID = 100
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[1]/input"
    )
    first_input.send_keys(tenantId)
    # 職員コード
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[2]/input"
    )
    first_input.send_keys(empId)
    # パスワード（※2021.11.17時点、方針未決）
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[3]/input"
    )
    first_input.send_keys(passWd)
    # ログインボタンをクリック
    click_element(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[5]/button"
    )
    time.sleep(2)

def sign_out_procedure():    # 2021.12.15
    # ①申請完了後の「閉じる」ボタンをクリック
    close_button = driver.find_element_by_xpath(
        # "/html/body/div[1]/div/div/form/div/button[4]"        # --- CHANGE 2021.12.24
        "/html/body/div[1]/div/div/form/div/button[3]"          # --- CHANGE 2021.12.24
    )
    my_sleep_click(close_button)

    driver.implicitly_wait(3)                           # --- ADD NEW 2021.12.22
    driver.switch_to.window(driver.window_handles[0])   # --- ADD NEW 2021.12.22 要素ゼロのウィンドウへ制御を戻す！

    # ②画面右上「サインアウト」ボタン（非常出口アイコン）をクリック
    sign_out_button = driver.find_element_by_xpath(
        "/html/body/div[2]/nav/div/div[2]/ul/li[6]/form/a"
    )
    my_sleep_click(sign_out_button)
    time.sleep(2)
    # ③【ログイン】ボタンをラストにクリック
    sign_in_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    )
    my_sleep_click(sign_in_button)
    time.sleep(2)

# def judgeItemCode(wkItem, wkBearDeptCd):
#    # ※品目は飲食のみである
#    if wkItem == '飲食':
#        # 部課名に開発を含有する
#        # ﾏｰｹﾃｨﾝｸﾞ部 PIS課 / ﾏｰｹﾃｨﾝｸﾞ部ﾌﾟﾛﾓｰｼｮﾝ課 / 開発部 Web開発課 # 開発部共通 # i-FILTER課 / 開発１課 # m-FILTER課 / 開発２課
#        # iFBC/i-ﾌｨﾙﾀｰ課 / 開発３課 # FinalCode課 / 開発４課 # ﾆｭｰﾌﾟﾛﾀﾞｸﾄ課 / 開発５課 # Iﾗﾎﾞ開発課 # 研究開発課# 海外開発課
#         if wkBearDeptCd == 671 \
#             or wkBearDeptCd == 1000 \
#             or wkBearDeptCd == 1119 \
#             or wkBearDeptCd == 1120 \
#             or wkBearDeptCd == 1121 \
#             or wkBearDeptCd == 1122 \
#             or wkBearDeptCd == 1123 \
#             or wkBearDeptCd == 1294 \
#             or wkBearDeptCd == 1900 \
#             or wkBearDeptCd == 1901:
#                 return 'i10r-50201'
#        else:
#             return 'i10r-70802'
#     else:
#         logger.info("品目コード該当なし")
#         sys.exit()

def entry_procedure():

    reloadBrowser()     # --- 2021.12.02 Chromeブラウザー不具合につき追加
    # click_element("/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button")
    # 2021.11.27 UPDATE Start
    book = openpyxl.load_workbook("C:\XLSX\XM_SLIP\PJ.xlsx")    # Book名がPJ
    sheet = book['PJ']                                          # Sheet名がPJ

    i = 8   # 1行目は見出し行
    with concurrent.futures.ProcessPoolExecutor() as executor:
        while sheet.cell(row=i, column=1).value != '' or sheet.cell(row=i, column=1).value is not None:
            pr = PaymentRequest(sheet.cell(row=i, column=9).value,      # ※L列[]CREATORを緊急避難で取得
                                '0Nu4M0%4N0',                           # ※共通パスワード。2021.11.29時点未定！ 決まったら書直す事！
                                sheet.cell(row=i, column=79).value,     # 目的
                                sheet.cell(row=i, column=48).value,     # 支払希望日
                                sheet.cell(row=i, column=5).value,      # 請求書番号
                                sheet.cell(row=i, column=63).value,     # 支払先コード
                                sheet.cell(row=i, column=79).value,     # 備考A(CA)
                                sheet.cell(row=i, column=86).value,     # 備考B(CH)
                                sheet.cell(row=i, column=80).value,     # 品目
                                sheet.cell(row=i, column=18).value,     # 負担部門コード（※直接エントリーに用いない）
                                # sheet.cell(row=i, column=10).value,   # 負担部門コード（※直接エントリーに用いない）
                                sheet.cell(row=i, column=83).value,     # 内容
                                sheet.cell(row=i, column=41).value,     # 費用
                                sheet.cell(row=i, column=28).value,     # 税込金額
                                sheet.cell(row=i, column=54).value,     # ※消費税は2021.11.29時点未定！ 決まったら書直す事！
                                sheet.cell(row=i, column=50).value,     # 源泉税計算対象区分
                                sheet.cell(row=i, column=256).value,    # 源泉税額
                                sheet.cell(row=i, column=16).value,     # 負担部門コード
                                sheet.cell(row=i, column=81).value,     # ※分析コードは2021.11.29時点未定！ 決まったら書直す事！
                                sheet.cell(row=i, column=109).value,    # 証憑PDFファイル絶対パスその1
                                sheet.cell(row=i, column=110).value,    # 証憑PDFファイル絶対パスその2
                                sheet.cell(row=i, column=111).value,    # 証憑PDFファイル絶対パスその3
                                sheet.cell(row=i, column=112).value,    # 証憑PDFファイル絶対パスその4
                                sheet.cell(row=i, column=113).value,    # 証憑PDFファイル絶対パスその5
                                sheet.cell(row=i, column=114).value,    # 証憑PDFファイル絶対パスその6
                                sheet.cell(row=i, column=115).value,    # 証憑PDFファイル絶対パスその7
                                sheet.cell(row=i, column=116).value,    # 証憑PDFファイル絶対パスその8
                                sheet.cell(row=i, column=117).value,    # 証憑PDFファイル絶対パスその9
                                sheet.cell(row=i, column=118).value     # 証憑PDFファイル絶対パスその10
            )

            # 引数: 左からテナントID、職員番号、パスワード
            gen = pr.getSignInInfo()
            # signin_procedure(gen.__next__(), gen.__next__(), gen.__next__())   <----- サインイン方法確定後有効化せよ！！
            gen = None

            # 下記は検証時における、テナントID直下の【ログイン】ボタン。本稼働時に上記サインイン処理が有効化されたら無効化せよ！！
            click_element(
                "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
            )

            click_element(      # 起票ボタンをクリック
                "/html/body/div[2]/div[1]/ul[1]/li[4]/a"
            )

            click_element(      # ●支払依頼書を選択
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[5]/label[6]/div/div[1]/input"
            )

            driver.implicitly_wait(5)   # 2021.11.27 --- 10から5へ減数

            click_element(      # 次へボタンをクリック
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[3]/button"
            )

            driver.implicitly_wait(5)   # 2021.11.27 --- 10から5へ減数
            driver.switch_to.window(driver.window_handles[1])
            # create_button = driver.find_element_by_xpath('//*[text()="起票する"]')    # 先任者において既に無効化されていた

            create_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[3]/button[2]"
            )
            my_sleep_click(create_button)

            # PURPOSE = "目的"  --- DELETE 2021.11.29 - 前任者による定数エントリー
            purpose_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div/input"
            )
            # purpose_field.send_keys(PURPOSE)      # --- DELETE 2021.11.29 - 前任者による定数エントリー
            purpose_field.send_keys(pr.getAim())    # --- UPDATE 2021.11.29 - ゲッターより値を返す

            # 支払希望日（支払期日）
            # PAYMENT_DESIRE_DATE = "2021-08-02"    # --- DELETE 2021.11.29 - 前任者による定数エントリー
            payment_desire_date_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div[1]/div/input"
            )
            # payment_desire_date_field.send_keys(PAYMENT_DESIRE_DATE)      # --- DELETE 2021.11.29 - 前任者による定数エントリー
            payment_desire_date_field.send_keys(pr.getWishDate())           # --- UPDATE 2021.11.29 - ゲッターより値を返す
                                                                            # --- ※当該フィールドは全数が空白値である

            # TODO
            # 請求書番号    # --- 前任者によるサボタージュ
            invoice_no_textbox = driver.find_element_by_xpath(              # --- ADD NEW 2021.11.29 - Start
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div[2]/input"
            )
            invoice_no_textbox.send_keys(pr.getInvoiceNo())                 # --- ADD NEW 2021.11.29 - End
            invoice_no_textbox.send_keys(
                Keys.TAB            # --- 消しゴムアイコンのボタン
                + Keys.TAB          # --- iframeオープンボタン
                + Keys.ENTER        # --- iframeオープンボタンへ改行キーを打鍵
            )
            # Search Payee  # --- 前任者によるコメント
            # PAYEE = "支払先"      # --- DELETE 2021.11.29 - 前任者による定数エントリー
            # payee_search_button = driver.find_element_by_xpath(   # --- DELETE 2021.11.29 Start
            #     "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[4]/div/div/span/button[2]"
            # )
            # my_sleep_click(payee_search_button)

            # time.sleep(5)

            # 取引先の検索
            # time.sleep(5)
            # driver.implicitly_wait(10)    # --- DELETE 2021.11.29 Start
            iframe = driver.find_element_by_xpath('//*[@id="zk_iFrameDialog"]/iframe')
            driver.switch_to.frame(iframe)

            deal_code_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[1]/input"
            )
            # EMPLOYEE_CODE = "10011"                   # --- DELETE 2021.11.29 - 前任者による定数エントリー
            # deal_code_field.send_keys(EMPLOYEE_CODE)  # --- DELETE 2021.11.29 - 前任者による定数エントリー
            # deal_code_field.send_keys(pr.getPaymentDestCd())    # --- UPDATE 2021.11.29 支払先コードエントリー
            deal_code_field.send_keys(                              # --- UPDATE 2021.12.24 外部へ処理を任せる --- Start
                pmd.returnPaymentDestCd(pr.getPaymentDestCd())
            )                                                       # --- UPDATE 2021.12.24 外部へ処理を任せる --- End
            # deal_code_field.send_keys(Keys.ENTER)
            deal_code_field.send_keys(  # --- ADD 2021.11.29 Start
                Keys.TAB
                + Keys.TAB
                + Keys.ENTER
            )                           # --- ADD 2021.11.29 End
            time.sleep(2)   # 引数5であったが2へ減数
            found_row = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[1]/div[3]/label/div"           # --- DELETE 2021.11.29
                # "/html/body/div[1]/div/div[1]/div[3]/label/div/div[1]"      # --- UPDATE 2021.11.29
            )
            my_sleep_click(found_row)
            driver.switch_to.default_content()

            # NOTE_TEXT = "テスト"      # --- DELETE 2021.11.29 - 前任者による定数エントリー

            note_text_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[6]/div/textarea"
            )
            # note_text_field.send_keys(NOTE_TEXT)          # --- DELETE 2021.11.29
            # note_text_field.send_keys(pr.getRemarks())      # --- UPDATE 2021.11.29
            gen = pr.getRemarks()                           # --- UPDATE 2021.12.21
            crlf = '\r\n'                                   # --- UPDATE 2021.12.21
            note_text_field.send_keys(str(gen.__next__()) + crlf + str(gen.__next__()))
            time.sleep(3)                                   # 5から3へ減数
            gen = None
            note_text_field.send_keys(                      # --- ADD NEW 2021.11.30 Start
                Keys.TAB
                + Keys.TAB
                + Keys.TAB
                + Keys.ENTER
            )                                               # --- ADD NEW 2021.11.30 End

            time.sleep(2)   # 必須の待機時間。無くすと実行時例外が発生する。
            detail_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[2]/div/div/span/button[2]"
            )
            my_sleep_click(detail_button)

            # Switch to frame
            driver.implicitly_wait(3)       # 10から3へ減数
            iframe = driver.find_element_by_xpath('//*[@id="zk_iFrameDialog"]/iframe')
            driver.switch_to.frame(iframe)
            time.sleep(2)                   # 5から2へ減数

            # Find details
            # ITEM_CODE = "i10r-80101"  # --- DELETE 2021.11.29 - 前任者による定数エントリー
            # gen = pr.getItem()      # yieldからGenerater経由で値を受取る    # --- DELETE 2021.12.25
            gen = itm.returnItem(pr.getItem())      # --- ADD NEW 2021.12.25 外部処理経由
            wkXpath = gen.__next__()        # 表示テーブルの何段目をクリックするか、当該段数を返する
            wkItemCode = gen.__next__()
            # wkItemCode = judgeItemCode(gen.__next__(), gen.__next__())
            # --- ADD NEW 品目コード選別関数（※飲食かつ●●の条件が正しいか要確認）
            # item_code_fld.send_keys(ITEM_CODE + Keys.ENTER)   # --- DELETE 2021.11.29
            # item_code_fld.send_keys(wkItemCode)               # --- UPDATE 2021.11.29
            item_code_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[1]/input"
            )
            item_code_fld.send_keys(
                wkItemCode
            )                     # --- UPDATE 2021.12.20

            item_code_fld.send_keys(        # --- UPDATE 2021.11.30 Start
                Keys.TAB
                + Keys.TAB
                + Keys.TAB
                + Keys.ENTER    # 虫眼鏡アイコン検索ボタンをクリック
            )                               # --- UPDATE 2021.11.30 End

            time.sleep(3)   # 引数5を3へ減数

            # すべて開くボタンをクリック
            open_all = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[1]/div[2]/button"
            )
            my_sleep_click(open_all)
            time.sleep(2)

            # item_row = driver.find_element_by_xpath(          # --- Total Abolition 2021.12.25 --- Start
            #    # "/html/body/div[1]/div/div[1]/div[4]/div/div/div/label/div"
            #    "/html/body/div[1]/div/div[1]/div[4]/div/div/div/div/label/div/div[1]"
            #    # str(wkXpath)      # --- UPDATE 2021.12.20 XPAHTを返す（※2, 3, 4, 5, 7段のパターンが存在）
            # )                                                 # --- Total Abolition 2021.12.25 --- Start

            if wkXpath == 2:        # 2段目
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/label[1]/div/div[1]"
                )
            elif wkXpath == 3:        # 3段目
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/div/label/div/div[1]"
                )
            elif wkXpath == 4:        # 4段目
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/div/div/label/div/div[1]"
                )
            elif wkXpath == 5:        # 5段目
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/div/div/div/label/div/div[1]"
                )
            elif wkXpath == 7:        # 7段目
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/div[2]/div/label/div/div[1]"
                )
            my_sleep_click(item_row)
            time.sleep(2)

            # 元のウィンドウへ制御を戻す
            driver.switch_to.default_content()

            time.sleep(2)   # 引数5を2へ減数
            # CONTENT = "内容"  # --- DELETE 2021.11.30 前任者による定数エントリ－
            content_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[3]/div/input"
            )
            # content_fld.send_keys(CONTENT)        # --- DELETE 2021.11.30
            content_fld.send_keys(pr.getContent())  # --- UPDATE 2021.11.30

            # COST = "1000"     # --- DELETE 2021.11.30 前任者による定数エントリ－
            cost_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[1]/input"
            )
            # cost_fld.send_keys(COST)              # --- DELETE 2021.11.30
            cost_fld.send_keys(pr.getCost())        # 費用

            # 以降、前任者がスルーした①税込金額、②消費税、③源泉税計算対象区分、④源泉税額、⑤負担部門、⑥分析コードをエントリー【開始】
            tax_included_amount_textbox = driver.find_element_by_xpath(     # 税込金額 --- 2021.11.30 ADD
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[2]/input"
            )
            # tax_included_amount_textbox.clear()
            # time.sleep(2)
            # tax_included_amount_textbox.send_keys(pr.getTaxIncludedAmount()) # PEND 2021.11.30

            # consumption_tax_textbox = driver.find_element_by_xpath(         # 消費税 --- 2021.11.30 ADD
            #    "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[4]/input"
            # )
            # if pr.getConsumptionTax() != 0:         # ！！ 本当にクリアして値をエントリーするのか、要確認
            #    consumption_tax_textbox.clear()
            # time.sleep(2)
            # consumption_tax_textbox.send_keys(
            #        pr.getConsumptionTax()
            # )

            withholding_tax_calclation_target_classification_textbox = driver.find_element_by_xpath(    # 源泉税計算対象区分
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[5]/div[1]/select"
            )
            withholding_tax_calclation_target_classification_textbox.send_keys(
                pr.getWithholdingTaxCalclationTargetClassification()
            )

            withholding_tax_amount_textbox = driver.find_element_by_xpath(  # 源泉税額
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[5]/div[2]/div/input"
            )
            withholding_tax_amount_textbox.send_keys(pr.getWithholdingTaxAmount())

            withholding_tax_amount_textbox.send_keys(
                Keys.TAB        # 電卓アイコンボタン
                + Keys.TAB      # 負担部門の消しゴムアイコンボタン
                + Keys.TAB      # 負担部門のiframe起動ボタンへカーソル移動
                + Keys.ENTER    # iframe起動ボタンを改行キーで打鍵
            )

            driver.implicitly_wait(3)
            # 制御をiframeへ渡す
            iframe = driver.find_element_by_xpath('//*[@id="zk_iFrameDialog"]/iframe')
            driver.switch_to.frame(iframe)
            time.sleep(2)

            burden_department_code_textbox = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[1]/input"
            )
            # burden_department_code_textbox.send_keys(pr.getBurdenDepartmentCode())    # --- INVALID 2021.12.25
            burden_department_code_textbox.send_keys(                                   # --- ADD NEW 2021.12.25 --- Start
                dpt.returnDepartment(pr.getBurdenDepartmentCode())
            )                                                                           # --- ADD NEW 2021.12.25 --- End
            burden_department_code_textbox.send_keys(
                Keys.TAB
                + Keys.TAB
                + Keys.ENTER
            )
            time.sleep(2)

            burden_department_table_record = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[1]/div[3]/label/div/div[1]"
            )
            my_sleep_click(burden_department_table_record)
            time.sleep(2)

            # 元のウィンドウへ制御を戻す
            driver.switch_to.default_content()

            analysis_code_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[6]/div[2]/div/span/button[2]"
            )
            my_sleep_click(analysis_code_button)

            driver.implicitly_wait(3)
            # 制御をiframeへ渡す
            iframe = driver.find_element_by_xpath('//*[@id="zk_iFrameDialog"]/iframe')
            driver.switch_to.frame(iframe)
            time.sleep(2)

            analysis_code_textbox = driver.find_element_by_xpath(   # 分析コードコード（ママ）
                "/html/body/div[1]/div/form/div/div[1]/input"
            )
            analysis_code_textbox.send_keys(pr.getAnalysisCode())

            analysis_code_textbox.send_keys(
                Keys.TAB        # 分析コード名テキストボックス
                + Keys.TAB      # 検索ボタンへカーソルが移動する
                + Keys.ENTER    # 検索ボタンへ改行キーを打鍵
            )
            time.sleep(2)

            analysis_code_table = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[1]/div[3]/label/div/div[1]"
            )
            my_sleep_click(analysis_code_table)
            time.sleep(2)

            # 元のウィンドウへ制御を戻す
            driver.switch_to.default_content()
            # 以上、前任者がスルーした①税込金額、②消費税、③源泉税計算対象区分、④源泉税額、￥⑤負担部門、⑥分析コードをエントリー【終了】

            confirm_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[5]/div/button[1]"
            )

            my_sleep_click(confirm_button)

            # No attach file for this application <--- 前任者による誤認。証憑ファイル提出は存在する。 --- 2021.11.30
            gen = pr.getVoucherPDFs()
            submit_voucher_files(gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__()
            )                                   # --- ADD NEW 2021.11.30 - 2021.12.01

            final_register_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[11]/button[1]"
            )
            click_submit_button(driver, logger, "Validation", final_register_button)

            # --- ADD NEW 2021.12.23 通知チェックボックスのレ点を解除
            notificationCheckBox = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/form/div[3]/label/input[1]"
            )
            my_sleep_click(notificationCheckBox)

            comment_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/form/textarea"                                 # --- RESTORE 2021.12.24
            )
            comment_field.send_keys(MIGRATION_COMMENT)  # --- ※前任者による備考テキストボックスへのコメント追記。本番実行時は削除せよ

            final_submit_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/form/div[4]/button[1]"
            )
            my_sleep_click(final_submit_button)

            # TODO uncomment below after testing
            # click_main(driver, logger, 'Submission', final_register_button)   # <--- 左記は不可

            # サインアウト処理
            sign_out_procedure()
        
            # ガーベジコレクター
            if i % 100 == 0:
                gc.collect()
            # PJシートの次行へ移動
            i += 1

def submit_voucher_files(vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9):     # --- ADD NEW 2021.11.30
    # 【添付ファイル】タブをクリック
    attach_file_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div[2]/div/form/div[5]/div/ul/li[3]/a"
    )
    my_sleep_click(attach_file_button)
    # 配列へ格納 --- ※証憑はPDFのみとは限らない
    va = [vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9]

    j = 0   # ゼロオリジン
    if va[j] != "N/A":
        while va[j] is not None and j < 10:
            if j == 0:  # 2回目以降は不要である為
                select_file_button = driver.find_element_by_xpath(
                # 添付ファイルタブに所在する添付ファイル-選択ボタン - 2021.11.15 コメント記入
                    "/html/body/div[1]/div[2]/div/form/div[5]/div/div/div["
                    "3]/div/div/div/div[2]/div[2]/div/span/button"
                )
                my_sleep_click(select_file_button)

            upload_file_button = driver.find_element_by_xpath(
                # 次画面にある「添付ファイルを選択してください」→選択ボタンをクリック
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[2]/div[2]"
            )
            upload_file_button.click()
            time.sleep(2)
            pyperclip.copy(va[j])                      # Clip Boardへコピーし、
            pg.hotkey('ctrl', 'v')              # Pasteする。
            pg.press("enter")
            time.sleep(2)
            add_green_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[3]/div/a"
            )
            my_sleep_click(add_green_button)
            time.sleep(2)
            # 右隣セルの値へ移動
            j += 1

    # 証憑提出後、確定ボタンクリック
    if j != 0:  # --- ADD NEW 2021.12.24 --- 証憑が一つ以上あった場合のみ確定ボタンをクリック
        final_register_button = driver.find_element_by_xpath(
            # "/html/body/div[1]/div[2]/div/form/div[11]/button[1]"
            "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"
        )
        my_sleep_click(final_register_button)

def main():
    # エントリー手続き
    entry_procedure()
    logger.info("Robot completed")

main()