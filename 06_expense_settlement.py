# 【経費】経費精算書
import logging
from chrome_driver_dl import get_latest_driver
from common import *
import pyautogui as pg  # --- UPDATE 2021.12.23
pg.FAILSAFE = False     # --- UPDATE 2021.12.23

# driver = webdriver.Chrome("../Chrome/chromedriver.exe")
# driver.get(MFZC_URL)
# driver.fullscreen_window()
# driver.refresh()

# --- ADD 2021.12.09 Start
import openpyxl
import sys
import pyperclip
import gc
from datetime import datetime
import concurrent.futures
import re

from selenium import webdriver
driver = None
driver = webdriver.Chrome(
    executable_path="C:/Users/digiworker_biz_02/PycharmProjects/pythonProject/Chrome/chromedriver.exe"
)
# --- ADD 2021.12.09 End
import accountingsubject as asj                   # --- ADD NEW 2021.12.25 AM
import department as dpt                          # --- ADD NEW 2021.12.26 支払依頼書における（負担部門）部署コードと同じ構造

# Gets or creates a logger
logger = logging.getLogger("05")
# set log level
logger.setLevel(logging.INFO)
# define file handler and set formatter
file_handler = logging.FileHandler("../Logs/01_expense_settlement.log")
formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")
file_handler.setFormatter(formatter)
# add file handler to logger
logger.addHandler(file_handler)

# driver = webdriver.Chrome(get_latest_driver())

driver.get(MFZC_URL)
driver.fullscreen_window()
driver.refresh()

logger.info("Started")

class Expense:
        def __init__(
                self,
                wkUserId,
                wkPassWd,
                wkPurposeOfUse,
                wkRemark1,
                wkRemark2,
                wkAccrualDate,              # 発生日
                wkItemType,                 # 明細種別
                wkContents,                 # 内容
                wkItemSelection,            # 品目選択
                wkPaymentDestination,       # 支払先
                wkCost,                     # 費用
                wkTaxIncludedAmount,        # 税込金額
                wkConsumptionTax,           # 消費税額
                wkBurdenDepartment,         # 負担部門
                wkVoucher0,
                wkVoucher1,
                wkVoucher2,
                wkVoucher3,
                wkVoucher4,
                wkVoucher5,
                wkVoucher6,
                wkVoucher7,
                wkVoucher8,
                wkVoucher9
        ):
                self.userId = wkUserId
                self.passWd = wkPassWd
                self.porposeOfUse = wkPurposeOfUse
                self.remark1 = wkRemark1
                self.remark2 = wkRemark2
                self.accrualDate = wkAccrualDate
                self.itemType = wkItemType
                self.contents = wkContents
                self.itemSelection = wkItemSelection
                self.paymentDestination = wkPaymentDestination
                self.cost = wkCost
                self.taxIncludedAmount = wkTaxIncludedAmount
                self.consumptionTax = wkConsumptionTax
                self.burdenDepartment = wkBurdenDepartment
                self.voucher0 = wkVoucher0
                self.voucher1 = wkVoucher1
                self.voucher2 = wkVoucher2
                self.voucher3 = wkVoucher3
                self.voucher4 = wkVoucher4
                self.voucher5 = wkVoucher5
                self.voucher6 = wkVoucher6
                self.voucher7 = wkVoucher7
                self.voucher8 = wkVoucher8
                self.voucher9 = wkVoucher9

        def getSignInInfo(self):
            yield 100
            yield self.userID
            yield self.passWd

        def getPorposeOfUse(self):
            return self.porposeOfUse

        def getRemark(self):
            if self.remark1 is not None:
                yield self.remark1
            else:
                yield ''
            if self.remark2 is not None:
                yield self.remark2
            else:
                yield ''

        def getItemType(self):
            return "精算明細-立替"

        def getContents(self):
            return self.contents

        def getItemSelection(self): # ※IF文右側のコードは旧MFのコード。対（つい）となる新MFZコードをreturnの右手へ書入れる事！！
            return self.itemSelection       # --- CHANGE 2021.12.25

        def getPaymentDestination(self):
            return self.paymentDestination
        
        def getCost(self):
            return self.cost

        def getTaxIncludedAmount(self):
            return self.taxIncludedAmount

        def getConsumptionTax(self):
            return self.consumptionTax

        def getBurdenDepartment(self):      # 支払部門コードであって、支払部門名ではない。※旧MFの支払部門マスターと必ずしも合致しない
            return self.burdenDepartment    # 2021.12.25～26 共通部品化

        def getVoucherFiles(self):
            yield self.voucher0
            yield self.voucher1
            yield self.voucher2
            yield self.voucher3
            yield self.voucher4
            yield self.voucher5
            yield self.voucher6
            yield self.voucher7
            yield self.voucher8
            yield self.voucher9

def click_element(web_el):
    my_sleep_click(driver.find_element_by_xpath(web_el))


def click_main(web_el):
    try:
        web_el.click()
        logger.info("Success")
    except NoSuchElementException:
        logger.error("Error")
    time.sleep(5)

def reloadBrowser():
    try:
        time.sleep(2)
        driver.refresh()
        time.sleep(2)
    except Exception as e:
        logger.info("Chromeブラウザーをリロードできず。ドライバーを確認せよ")
        # driver.quit()
    finally:
        pass

def signin_procedure(tenantId, empId, passWd):
    # テナントID = 100
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[1]/input"
    )
    first_input.send_keys(tenantId)
    # 職員コード
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[2]/input"
    )
    first_input.send_keys(empId)
    # パスワード（※2021.11.17時点、方針未決）
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[3]/input"
    )
    first_input.send_keys(passWd)
    # ログインボタンをクリック
    click_element(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[5]/button"
    )
    time.sleep(3)

def sign_out_procedure():    # 2021.12.15
    # ①申請完了後の「閉じる」ボタンをクリック
    # close_button = driver.find_element_by_xpath(
    #    "/html/body/div[1]/div/div/form/div/button[4]"
    # )
    # my_sleep_click(close_button)

    print("efghi")

    driver.implicitly_wait(3)                           # --- ADD NEW 2021.12.22
    driver.switch_to.window(driver.window_handles[0])   # --- ADD NEW 2021.12.22 要素ゼロのウィンドウへ制御を戻す！

    print("abcd")

    # ②画面右上「サインアウト」ボタン（非常出口アイコン）をクリック
    sign_out_button = driver.find_element_by_xpath(
        "/html/body/div[2]/nav/div/div[2]/ul/li[6]/form/a"
    )
    my_sleep_click(sign_out_button)
    time.sleep(2)
    # ③【ログイン】ボタンをラストにクリック
    sign_in_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    )
    my_sleep_click(sign_in_button)
    time.sleep(2)

def entry_procedure():

    reloadBrowser()
    # click_element(
    #    "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    # )
    # 経費精算申請書の伝票タイプはE1
    book = openpyxl.load_workbook("C:\XLSX\XM_SLIP_DETAIL\E1.xlsx")     # Book名がE1
    sheet = book['E1']                                                  # 経費精算書

    with concurrent.futures.ProcessPoolExecutor() as executor:
        i = 2   # 1行目は見出し行。2行目から開始。A列セルをNull判定しEOFを見極める。
        while sheet.cell(row=i, column=1).value is not None:
            ex = Expense(
                                sheet.cell(row=i, column=87).value,             # CREATOR
                                '0Nu4M0%4N0',                                   # 共通パスワード（※未定）
                                sheet.cell(row=i, column=31).value,             # 利用目的【SS01】（※確証なし）
                                sheet.cell(row=i, column=44).value,             # 備考その1【SS13】（※左記は未定、要確認）
                                sheet.cell(row=i, column=45).value,             # 備考その2【SS14】（※左記は未定、要確認）
                                sheet.cell(row=i, column=256).value,            # 発生日（※処理日で良いか？ 要確認）
                                sheet.cell(row=i, column=256).value,            # 明細種別（※「精算明細-立替」を決打ちする、で良いか？ 要確認）
                                sheet.cell(row=i, column=81).value,             # 内容【ITEMNAME1】←SS01と同じ様子
                                sheet.cell(row=i, column=67).value,             # 品目選択【SS21】（[i01q-001]のような値）
                                sheet.cell(row=i, column=44).value,             # 支払先【SS14】
                                sheet.cell(row=i, column=16).value,             # 費用【PRICE1】←2が税抜額、3が消費税である様子
                                sheet.cell(row=i, column=16).value,             # 税込金額【PRICE1】←費用テキストボックスと何が異なるのか？
                                sheet.cell(row=i, column=18).value,             # 消費税額【PRICE3】
                                sheet.cell(row=i, column=6).value,              # 負担部門【BEARDEPTID】負担部門コード（例: 541→経理課）
                                sheet.cell(row=i, column=90).value,             # 証憑ファイル1
                                sheet.cell(row=i, column=91).value,             # 証憑ファイル2
                                sheet.cell(row=i, column=92).value,             # 証憑ファイル3
                                sheet.cell(row=i, column=93).value,             # 証憑ファイル4
                                sheet.cell(row=i, column=94).value,             # 証憑ファイル5
                                sheet.cell(row=i, column=95).value,             # 証憑ファイル6
                                sheet.cell(row=i, column=96).value,             # 証憑ファイル7
                                sheet.cell(row=i, column=97).value,             # 証憑ファイル8
                                sheet.cell(row=i, column=98).value,             # 証憑ファイル9
                                sheet.cell(row=i, column=99).value              # 証憑ファイル10
            )

            # 引数: 左からテナントID、職員番号、パスワード
            gen = ex.getSignInInfo()
            # signin_procedure(gen.__next__(), gen.__next__(), gen.__next__())   <----- サインイン方法確定後、有効化せよ
            gen = None

            click_element(
               "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
            )

            click_element(
                "/html/body/div[2]/div[1]/ul[1]/li[4]/a"
            )

            click_element(
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[5]/label[5]/div/div[1]/input"
            )

            driver.implicitly_wait(5)       # 引数10を5へ減数

            # print(driver.current_url)     # --- DELETE 2021.12.10 --- 前任者による削除忘れと思料される 

            click_element(
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[3]/button"
            )

            driver.implicitly_wait(5)       # 引数10を5へ減数
            driver.switch_to.window(driver.window_handles[1])
            # create_button = driver.find_element_by_xpath('//*[text()="起票する"]')

            create_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[3]/button[2]"
            )
            my_sleep_click(create_button)

            # PURPOSE = "利用目的"                                  # --- DELETE 2021.12.10 --- 前任者による定数エントリー（※場所移動）
            purpose_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div/input"
            )
            # purpose_fld.send_keys(PURPOSE)                        # --- DELETE 2021.12.10
            purpose_fld.send_keys(ex.getPorposeOfUse())

            # --- ADD NEW 2021.12.13 --- 前任者不作為につき新規追加 --- Start
            gen = ex.getRemark()
            remark_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div/textarea"
            )
            remark_fld.send_keys(
                str(gen.__next__()) + '\r\n' +  str(gen.__next__()) + '\r\n'    # CRLF改行
            )
            gen = None
            # --- ADD NEW 2021.12.13 --- 前任者不作為につき新規追加 --- End

            # DETAIL_TYPE = "精算明細-立替"                         # --- DELETE 2021.12.10 --- 前任者による定数エントリー
            detail_type_select_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div/div[2]/div/div[" "2]/div[2]/select"
            )
            # detail_type_select_field.send_keys(DETAIL_TYPE)       # --- DELETE 2021.12.10
            detail_type_select_field.send_keys(ex.getItemType())    # --- UPDATE 2021.12.10

            detail_type_select_field.send_keys(
                Keys.TAB
                 + Keys.ENTER
            )

            # Now Input fields from here    # <--- 意味不明の前任者コメント
            # CONTENT = "内容（●●購入代、飲食代等）"                # --- DELETE 2021.12.10 --- 前任者による定数エントリー
            content_input = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div["
                "2]/div[1]/input"
            )
            # content_input.send_keys(CONTENT)                      # --- DELETE 2021.12.10
            content_input.send_keys(ex.getContents())               # --- UPDATE 2021.12.10

            main_page = driver.current_window_handle
            # Enter to frame
            content_input.send_keys(
                Keys.TAB
                 + Keys.TAB
                 + Keys.ENTER
            )

            # ITEM_CODE = "i10r-80101"                             # --- DELETE 2021.12.10 --- 前任者による定数エントリー
            item_get_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div["
                "3]/div/div[2]/div[2]/div/span/button[2]"
            )

            driver.implicitly_wait(3)       # 引数10を3へ減数 

            iframe = driver.find_element_by_xpath('//*[@id="zk_iFrameDialog"]/iframe')
            driver.switch_to.frame(iframe)

            item_code_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[1]/input"
            )
            gen = asj.returnAccountingSubject(ex.getItemSelection())     # --- UPDATE 2021.12.26 共通部品使用へ変更。ジールドで値を取得
            wkXpath = gen.__next__()                        # --- テーブルのクリックする段数（変数へ保存）
            wkItemCd = gen.__next__()                       # --- アイテムコード（変数へ保存）
            item_code_fld.send_keys(
                 # ex.getItemSelection()    # --- UPDATE 2021.12.10
                 wkItemCd                   # --- UPDATE 2021.12.26
                 + Keys.TAB                 # --- ADD NEW 2021.12.10 --- Start
                 + Keys.TAB                 #
                 + Keys.TAB                 # --- ADD NEW 2021.12.10 --- End
                 + Keys.ENTER
            )

            time.sleep(3)                   # 引数5を3へ減数
            open_all = driver.find_element_by_xpath(        # すべて開くボタン
                "/html/body/div[1]/div/div[1]/div[2]/button"
            )
            my_sleep_click(open_all)

            # item_row = driver.find_element_by_xpath(                            # --- INVALID 2021.12.26 Start
            #     "/html/body/div[1]/div/div[1]/div[4]/div/div/div/label/div"     #     前任者仕様固定値
            # )                                                                   # --- INVALID 2021.12.26 End
            if wkXpath == 2:
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/label/div/div[1]"
                )
            elif wkXpath == 3:
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/div/label/div/div[1]"
                )
            elif wkXpath == 4:
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/div/div/label/div/div[1]"
                )
            my_sleep_click(item_row)
            # close_button = driver.find_element_by_xpath('/html/body/div[1]/div/div[2]/form/button')   # 前任者による無効化
            # my_sleep_click(close_button)

            driver.switch_to.default_content()

            # PAY_TO = "支払先"                                      # --- DELETE 2021.12.10 --- 前任者による定数エントリー
            pay_to_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div["
                "3]/div[1]/input"
            )
            # pay_to_fld.send_keys(PAY_TO)                          # --- DELETE 2021.12.10
            pay_to_fld.send_keys(ex.getPaymentDestination())        # --- UPDATE 2021.12.10

            # EXPENSE = "1100"                                      # --- DELETE 2021.12.10 --- 前任者による定数エントリー
            expense_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[3]/div[2]/input"
            )
            # expense_fld.send_keys(EXPENSE)                        # --- DELETE 2021.12.10
            expense_fld.send_keys(ex.getCost())                     # --- UPDATE 2021.12.10

            # AMOUNT_WITH_TAX = '1100'                              # --- 前任者による無効化 --- Start
            # amount_with_tax_fld = driver.find_element_by_xpath(
            #   '/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[1]/input'
            # )
            # amount_with_tax_fld.send_keys(AMOUNT_WITH_TAX)        # --- 前任者による無効化 --- End

            # --- ADD NEW 2021.12.13 --- 前任者不作為につき新規追加 --- Start
            expense_fld.send_keys(
                Keys.TAB
            )
            time.sleep(2)
            tax_included_amount_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[1]/input"
            )
            tax_included_amount_fld.clear()
            tax_included_amount_fld.send_keys(ex.getTaxIncludedAmount())
            time.sleep(2)
            tax_included_amount_fld.send_keys(
                Keys.TAB
            )
            time.sleep(2)

            consumption_tax_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[3]/input"
            )
            consumption_tax_fld.clear()
            time.sleep(2)
            consumption_tax_fld.send_keys(ex.getConsumptionTax())
            consumption_tax_fld.send_keys(
                Keys.TAB            # 消しゴムボタン
                + Keys.TAB          # 負担部門ボタン
                + Keys.ENTER        # 改行キー打鍵
            )
            time.sleep(2)
            # iframeへ画面遷移
            iframe = driver.find_element_by_xpath('//*[@id="zk_iFrameDialog"]/iframe')
            driver.switch_to.frame(iframe)
            driver.implicitly_wait(2)

            burden_department_code_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[1]/input"
            )
            # burden_department_code_fld.send_keys(ex.getBurdenDepartment())                        # --- INVALID 2021.12.26
            burden_department_code_fld.send_keys(dpt.returnDepartment(ex.getBurdenDepartment()))    # --- UPDATE 2021.12.26
            time.sleep(2)
            burden_department_code_fld.send_keys(
                Keys.TAB        # 負担部門名テキストボックスへカーソル遷移
                + Keys.TAB      # 虫眼鏡ボタンへカーソル遷移
                + Keys.ENTER    # 改行キー打鍵
            )
            time.sleep(2)
            burden_department_table = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[1]/div[3]/label/div/div[1]"
            )
            my_sleep_click(burden_department_table)
            time.sleep(2)
            driver.switch_to.default_content()          # 制御を親ウィンドウへ戻す
            # --- ADD NEW 2021.12.13 --- 前任者不作為につき新規追加 --- End

            confirm_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[5]/div/button[1]"
            )
            my_sleep_click(confirm_button)
            # Next attach the evidence

            # Attach File function
            # attach_file_button = driver.find_element_by_xpath(
            #     "/html/body/div[1]/div[2]/div/form/div[5]/div/ul/li[3]/a"
            # )
            # my_sleep_click(attach_file_button)
            # EXPENSE_SETTLE_FILE = "C:\Output\MFZ\Test01.pdf"
            # select_file_button = driver.find_element_by_xpath(
            #     "/html/body/div[1]/div[2]/div/form/div[5]/div/div/div["
            #     "3]/div/div/div/div[2]/div[2]/div/span/button"
            # )
            # my_sleep_click(select_file_button)
            # upload_file_button = driver.find_element_by_xpath(
            #     "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[2]/div[2]"
            # )
            # upload_file_button.click()
            # time.sleep(4)
            # pyautogui.write(EXPENSE_SETTLE_FILE)
            # pyautogui.press("enter")
            # time.sleep(4)
            # add_green_button = driver.find_element_by_xpath(
            #     "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[3]/div/a"
            # )
            # my_sleep_click(add_green_button)
            # confirm_button = driver.find_element_by_xpath(
            #     "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"
            # )
            # my_sleep_click(confirm_button)

            gen = ex.getVoucherFiles()                                      # Generatorへ格納し、next()で取得する
            submit_voucher_files(gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__()
            )
            gen = None  # 参照を切離す
            ex = None   # 参照を切離す
            final_register_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[11]/button[1]"
            )
            click_submit_button(driver, logger, "Validation", final_register_button)

            time.sleep(2)

            # --- ADD NEW 2021.12.23 通知チェックボックスのレ点を解除
            notificationCheckBox = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/form/div[3]/label/input[1]"
            )
            my_sleep_click(notificationCheckBox)

            # 前任者による意図的なコメントエントリー。実運用時は除外せよ！！
            comment_field = driver.find_element_by_xpath(
                # "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div/textarea" # --- INVALID 2021.12.26
                "/html/body/div[1]/div/div/form/textarea"   # --- CHANGE 2012.12.26 変更があったのか？
            )
            comment_field.send_keys(MIGRATION_COMMENT)

            # Final submission
            final_register_button = driver.find_element_by_xpath(
                # "/html/body/div[1]/div[2]/div/form/div[11]/button[1]"
                "/html/body/div[1]/div/div/form/div[4]/button[1]"
            )
            # TODO need to uncomment it later
            # click_submit_button(driver, logger, 'Validation', final_register_button)
            my_sleep_click(final_register_button)  # <--- UPDATE 2021.12.23     # 左記を正とする

            final_submit_button = driver.find_element_by_xpath(
                # "/html/body/div[1]/div/div/form/div[4]/button[1]"
                "/html/body/div[1]/div/div/form/div/button[3]"
            )
            my_sleep_click(final_submit_button)  # <--- UPDATE 2021.12.23     # 左記を正とする

            # TODO uncomment below after testing
            # click_main(driver, logger, 'Submission', final_register_button)

            sign_out_procedure()

            # ガーベジコレクター
            if i % 100 == 0:
                gc.collect()
            # T1シートの次行へ移動
            i += 1

def submit_voucher_files(vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9): 
    # 【添付ファイル】タブをクリック
    attach_file_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div[2]/div/form/div[5]/div/ul/li[3]/a"
    )
    my_sleep_click(attach_file_button)
    # 配列へ格納 --- ※証憑はPDFのみとは限らない
    va = [vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9]

    j = 0  # ゼロオリジン
    if va[0] != "N/A":      # 一つ目のパスが無い場合は2つ目以降のパスも存在しないと見做す
        while va[j] is not None and j < 10:
            if j == 0:  # 2回目以降は不要である為
                select_file_button = driver.find_element_by_xpath(
                    # 添付ファイルタブに所在する添付ファイル-選択ボタン - 2021.11.15 コメント記入
                    "/html/body/div[1]/div[2]/div/form/div[5]/div/div/div["
                    "3]/div/div/div/div[2]/div[2]/div/span/button"
                )
                my_sleep_click(select_file_button)

            upload_file_button = driver.find_element_by_xpath(
            # 次画面にある「添付ファイルを選択してください」→選択ボタンをクリック
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[2]/div[2]"
            )
            upload_file_button.click()
            time.sleep(2)
            pyperclip.copy(va[j])                      # Clip Boardへコピーし、
            pg.hotkey('ctrl', 'v')              # Pasteする。
            pg.press("enter")
            time.sleep(2)
            add_green_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[3]/div/a"
            )
            my_sleep_click(add_green_button)
            time.sleep(2)
            # 右隣セルの値へ移動
            j += 1

    # 証憑提出後、確定ボタンクリック。ただし1件以上ある事が前提
    if j != 0:  # --- ADD NEW 2021.12.24 --- 証憑が一つ以上あった場合のみ確定ボタンをクリック
        final_register_button = driver.find_element_by_xpath(
            # "/html/body/div[1]/div[2]/div/form/div[11]/button[1]"
            "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"
        )
        my_sleep_click(final_register_button)

def main():
    entry_procedure()
    logger.info("Robot completed")

main()
