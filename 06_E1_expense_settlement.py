# 【経費】経費精算書
import logging
from chrome_driver_dl import get_latest_driver
from common import *
import pyautogui as pg  # --- UPDATE 2021.12.23
pg.FAILSAFE = False     # --- UPDATE 2021.12.23

# driver = webdriver.Chrome("../Chrome/chromedriver.exe")
# driver.get(MFZC_URL)
# driver.fullscreen_window()
# driver.refresh()

# --- ADD 2021.12.09 Start
import openpyxl
import sys
import pyperclip
import gc
from datetime import datetime
import concurrent.futures
import re

from selenium import webdriver
driver = None
driver = webdriver.Chrome(
    executable_path="C:/Users/digiworker_biz_02/PycharmProjects/pythonProject/Chrome/chromedriver.exe"
)
# --- ADD 2021.12.09 End
import accountingsubject as asj                   # --- ADD NEW 2021.12.25 AM
import department as dpt                          # --- ADD NEW 2021.12.26 支払依頼書における（負担部門）部署コードと同じ構造

# Gets or creates a logger
logger = logging.getLogger("06")                    # --- CHANGE 2022.01.04
# set log level
logger.setLevel(logging.INFO)
# define file handler and set formatter
file_handler = logging.FileHandler("../Logs/06_expense_settlement.log")     # --- CHANGE 2022.01.04
formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")
file_handler.setFormatter(formatter)
# add file handler to logger
logger.addHandler(file_handler)

# driver = webdriver.Chrome(get_latest_driver())

driver.get(MFZC_URL)
driver.fullscreen_window()
driver.refresh()

logger.info("Started")

class Expense:
        def __init__(
                self,
                wkUserId,
                wkPassWd,
                wkPurposeOfUse,
                wkRemark1,
                wkRemark2,
                wkAccrualDate,              # 発生日
                wkItemType,                 # 明細種別
                wkContents,                 # 内容
                wkItemSelection,            # 品目選択
                wkPaymentDestination,       # 支払先
                wkCost,                     # 費用
                wkTaxIncludedAmount,        # 税込金額
                wkConsumptionTax,           # 消費税額
                wkBurdenDepartment,         # 負担部門
                wkVoucher0,
                wkVoucher1,
                wkVoucher2,
                wkVoucher3,
                wkVoucher4,
                wkVoucher5,
                wkVoucher6,
                wkVoucher7,
                wkVoucher8,
                wkVoucher9,
                wkIid01,  # --- ADD NEW 2022.01.11 スコープ変更に伴う追加 x 16 （配列厳しい為、直打ち）
                wkDocno01,
                wkFormname01,
                wkCreated01,
                wkAname01,
                wkUid01,
                wkEmpcd01,
                wkEname01,
                wkDeptid01,
                wkDfullname01,
                wkPost01,
                wkTs01,
                wkIid02,
                wkDocno02,
                wkFormname02,
                wkCreated02,
                wkAname02,
                wkUid02,
                wkEmpcd02,
                wkEname02,
                wkDeptid02,
                wkDfullname02,
                wkPost02,
                wkTs02,
                wkIid03,
                wkDocno03,
                wkFormname03,
                wkCreated03,
                wkAname03,
                wkUid03,
                wkEmpcd03,
                wkEname03,
                wkDeptid03,
                wkDfullname03,
                wkPost03,
                wkTs03,
                wkIid04,
                wkDocno04,
                wkFormname04,
                wkCreated04,
                wkAname04,
                wkUid04,
                wkEmpcd04,
                wkEname04,
                wkDeptid04,
                wkDfullname04,
                wkPost04,
                wkTs04,
                wkIid05,
                wkDocno05,
                wkFormname05,
                wkCreated05,
                wkAname05,
                wkUid05,
                wkEmpcd05,
                wkEname05,
                wkDeptid05,
                wkDfullname05,
                wkPost05,
                wkTs05,
                wkIid06,
                wkDocno06,
                wkFormname06,
                wkCreated06,
                wkAname06,
                wkUid06,
                wkEmpcd06,
                wkEname06,
                wkDeptid06,
                wkDfullname06,
                wkPost06,
                wkTs06,
                wkIid07,
                wkDocno07,
                wkFormname07,
                wkCreated07,
                wkAname07,
                wkUid07,
                wkEmpcd07,
                wkEname07,
                wkDeptid07,
                wkDfullname07,
                wkPost07,
                wkTs07,
                wkIid08,
                wkDocno08,
                wkFormname08,
                wkCreated08,
                wkAname08,
                wkUid08,
                wkEmpcd08,
                wkEname08,
                wkDeptid08,
                wkDfullname08,
                wkPost08,
                wkTs08,
                wkIid09,
                wkDocno09,
                wkFormname09,
                wkCreated09,
                wkAname09,
                wkUid09,
                wkEmpcd09,
                wkEname09,
                wkDeptid09,
                wkDfullname09,
                wkPost09,
                wkTs09,
                wkIid10,
                wkDocno10,
                wkFormname10,
                wkCreated10,
                wkAname10,
                wkUid10,
                wkEmpcd10,
                wkEname10,
                wkDeptid10,
                wkDfullname10,
                wkPost10,
                wkTs10,
                wkIid11,
                wkDocno11,
                wkFormname11,
                wkCreated11,
                wkAname11,
                wkUid11,
                wkEmpcd11,
                wkEname11,
                wkDeptid11,
                wkDfullname11,
                wkPost11,
                wkTs11,
                wkIid12,
                wkDocno12,
                wkFormname12,
                wkCreated12,
                wkAname12,
                wkUid12,
                wkEmpcd12,
                wkEname12,
                wkDeptid12,
                wkDfullname12,
                wkPost12,
                wkTs12,
                wkIid13,
                wkDocno13,
                wkFormname13,
                wkCreated13,
                wkAname13,
                wkUid13,
                wkEmpcd13,
                wkEname13,
                wkDeptid13,
                wkDfullname13,
                wkPost13,
                wkTs13,
                wkIid14,
                wkDocno14,
                wkFormname14,
                wkCreated14,
                wkAname14,
                wkUid14,
                wkEmpcd14,
                wkEname14,
                wkDeptid14,
                wkDfullname14,
                wkPost14,
                wkTs14,
                wkIid15,
                wkDocno15,
                wkFormname15,
                wkCreated15,
                wkAname15,
                wkUid15,
                wkEmpcd15,
                wkEname15,
                wkDeptid15,
                wkDfullname15,
                wkPost15,
                wkTs15,
                wkIid16,
                wkDocno16,
                wkFormname16,
                wkCreated16,
                wkAname16,
                wkUid16,
                wkEmpcd16,
                wkEname16,
                wkDeptid16,
                wkDfullname16,
                wkPost16,
                wkTs16
        ):
                self.userId = wkUserId
                self.passWd = wkPassWd
                self.porposeOfUse = wkPurposeOfUse
                self.remark1 = wkRemark1
                self.remark2 = wkRemark2
                self.accrualDate = wkAccrualDate
                self.itemType = wkItemType
                self.contents = wkContents
                self.itemSelection = wkItemSelection
                self.paymentDestination = wkPaymentDestination
                self.cost = wkCost
                self.taxIncludedAmount = wkTaxIncludedAmount
                self.consumptionTax = wkConsumptionTax
                self.burdenDepartment = wkBurdenDepartment
                self.voucher0 = wkVoucher0
                self.voucher1 = wkVoucher1
                self.voucher2 = wkVoucher2
                self.voucher3 = wkVoucher3
                self.voucher4 = wkVoucher4
                self.voucher5 = wkVoucher5
                self.voucher6 = wkVoucher6
                self.voucher7 = wkVoucher7
                self.voucher8 = wkVoucher8
                self.voucher9 = wkVoucher9
                self.iid01 = wkIid01
                self.docno01 = wkDocno01
                self.formname01 = wkFormname01
                self.createdd01 = wkCreated01
                self.aname01 = wkAname01
                self.uid01 = wkUid01
                self.empcd01 = wkEmpcd01
                self.ename01 = wkEname01
                self.deptid01 = wkDeptid01
                self.dfullname01 = wkDfullname01
                self.post01 = wkPost01
                self.ts01 = wkTs01
                self.iid02 = wkIid02
                self.docno02 = wkDocno02
                self.forname02 = wkFormname02
                self.created02 = wkCreated02
                self.aname02 = wkAname02
                self.uid02 = wkUid02
                self.empcd02 = wkEmpcd02
                self.ename02 = wkEname02
                self.deptid02 = wkDeptid02
                self.dfullname02 = wkDfullname02
                self.post02 = wkPost02
                self.ts02 = wkTs02
                self.iid03 = wkIid03
                self.docno03 = wkDocno03
                self.fornmame03 = wkFormname03
                self.created03 = wkCreated03
                self.aname03 = wkAname03
                self.uid03 = wkUid03
                self.empcd03 = wkEmpcd03
                self.ename03 = wkEname03
                self.deptid03 = wkDeptid03
                self.dfullname03 = wkDfullname03
                self.post03 = wkPost03
                self.ts03 = wkTs03
                self.iid04 = wkIid04
                self.docno04 = wkDocno04
                self.formname04 = wkFormname04
                self.created04 = wkCreated04
                self.aname04 = wkAname04
                self.uid04 = wkUid04
                self.empcd04 = wkEmpcd04
                self.ename04 = wkEname04
                self.deptid04 = wkDeptid04
                self.dfullname04 = wkDfullname04
                self.post04 = wkPost04
                self.ts04 = wkTs04
                self.iid05 = wkIid05
                self.docno05 = wkDocno05
                self.formname05 = wkFormname05
                self.created05 = wkCreated05
                self.aname05 = wkAname05
                self.uid05 = wkUid05
                self.empcd05 = wkEmpcd05
                self.ename05 = wkEname05
                self.deptid05 = wkDeptid05
                self.dfullname05 = wkDfullname05
                self.post05 = wkPost05
                self.ts05 = wkTs05
                self.iid06 = wkIid06
                self.docno06 = wkDocno06
                self.formname06 = wkFormname06
                self.created06 = wkCreated06
                self.aname06 = wkAname06
                self.uid06 = wkUid06
                self.empcd06 = wkEmpcd06
                self.ename06 = wkEname06
                self.deptid06 = wkDeptid06
                self.dfullname06 = wkDfullname06
                self.post06 = wkPost06
                self.ts06 = wkTs06
                self.iid07 = wkIid07
                self.docno07 = wkDocno07
                self.forname07 = wkFormname07
                self.created07 = wkCreated07
                self.aname07 = wkAname07
                self.uid07 = wkUid07
                self.empcd07 = wkEmpcd07
                self.ename07 = wkEname07
                self.deptid07 = wkDeptid07
                self.dfullname07 = wkDfullname07
                self.post07 = wkPost07
                self.ts07 = wkTs07
                self.iid08 = wkIid08
                self.docno08 = wkDocno08
                self.forname08 = wkFormname08
                self.created08 = wkCreated08
                self.aname08 = wkAname08
                self.uid08 = wkUid08
                self.empcd08 = wkEmpcd08
                self.ename08 = wkEname08
                self.deptid08 = wkDeptid08
                self.dfullname08 = wkDfullname08
                self.post08 = wkPost08
                self.ts08 = wkTs08
                self.iid09 = wkIid09
                self.docno09 = wkDocno09
                self.formname09 = wkFormname09
                self.created09 = wkCreated09
                self.aname09 = wkAname09
                self.uid09 = wkUid09
                self.empcd09 = wkEmpcd09
                self.ename09 = wkEname09
                self.deptid09 = wkDeptid09
                self.dfullname09 = wkDfullname09
                self.post09 = wkPost09
                self.ts09 = wkTs09
                self.iid10 = wkIid10
                self.docno10 = wkDocno10
                self.formname10 = wkFormname10
                self.created10 = wkCreated10
                self.aname10 = wkAname10
                self.uid10 = wkUid10
                self.empcd10 = wkEmpcd10
                self.ename10 = wkEname10
                self.deptid10 = wkDeptid10
                self.dfullname10 = wkDfullname10
                self.post10 = wkPost10
                self.ts = wkTs10
                self.iid11 = wkIid11
                self.docno11 = wkDocno11
                self.fromname11 = wkFormname11
                self.created11 = wkCreated11
                self.aname11 = wkAname11
                self.uid11 = wkUid11
                self.empcd11 = wkEmpcd11
                self.ename11 = wkEname11
                self.deptid11 = wkDeptid11
                self.dfullname11 = wkDfullname11
                self.post11 = wkPost11
                self.ts11 = wkTs11
                self.iid12 = wkIid12
                self.docno12 = wkDocno12
                self.fromname12 = wkFormname12
                self.created12 = wkCreated12
                self.aname12 = wkAname12
                self.uid12 = wkUid12
                self.empcd12 = wkEmpcd12
                self.ename12 = wkEname12
                self.deptid12 = wkDeptid12
                self.dfullname12 = wkDfullname12
                self.post12 = wkPost12
                self.ts12 = wkTs12
                self.iid13 = wkIid13
                self.docno13 = wkDocno13
                self.formname13 = wkFormname13
                self.created13 = wkCreated13
                self.aname13 = wkAname13
                self.uid13 = wkUid13
                self.empcd13 = wkEmpcd13
                self.ename13 = wkEname13
                self.deptid13 = wkDeptid13
                self.dfullname13 = wkDfullname13
                self.post13 = wkPost13
                self.ts13 = wkTs13
                self.iid14 = wkIid14
                self.docno14 = wkDocno14
                self.formname14 = wkFormname14
                self.created14 = wkCreated14
                self.aname14 = wkAname14
                self.uid14 = wkUid14
                self.empcd14 = wkEmpcd14
                self.ename14 = wkEname14
                self.deptid14 = wkDeptid14
                self.dfullname14 = wkDfullname14
                self.post14 = wkPost14
                self.ts14 = wkTs14
                self.iid15 = wkIid15
                self.docno15 = wkDocno15
                self.formname15 = wkFormname15
                self.created15 = wkCreated15
                self.aname15 = wkAname15
                self.uid15 = wkUid15
                self.empcd15 = wkEmpcd15
                self.ename15 = wkEname15
                self.deptid15 = wkDeptid15
                self.dfullname15 = wkDfullname15
                self.post15 = wkPost15
                self.ts15 = wkTs15
                self.iid16 = wkIid16
                self.docno16 = wkDocno16
                self.fromname16 = wkFormname16
                self.created16 = wkCreated16
                self.aname16 = wkAname16
                self.uid16 = wkUid16
                self.empcd16 = wkEmpcd16
                self.ename16 = wkEname16
                self.deptid16 = wkDeptid16
                self.dfullname16 = wkDfullname16
                self.post16 = wkPost16
                self.ts16 = wkTs16

        def getSignInInfo(self):
            yield 100
            yield self.userID
            yield self.passWd

        def getPorposeOfUse(self):
            return self.porposeOfUse

        def getRemark(self):
            if self.remark1 is not None:
                yield self.remark1
            else:
                yield ''
            if self.remark2 is not None:
                yield self.remark2
            else:
                yield ''
            approvalInfo = ""
            if self.iid01 is not None and self.iid01 != "":
                approvalInfo += f'決裁:{self.aname01} 従業員ID:{self.uid01} 従業員番号:{self.empcd01} 従業員指名:{self.ename01} ;'
            if self.iid02 is not None and self.iid02 != "":
                approvalInfo += f'決裁:{self.aname02} 従業員ID:{self.uid02} 従業員番号:{self.empcd02} 従業員指名:{self.ename02} ;'
            if self.iid03 is not None and self.iid03 != "":
                approvalInfo += f'決裁:{self.aname03} 従業員ID:{self.uid03} 従業員番号:{self.empcd03} 従業員指名:{self.ename03} ;'
            if self.iid04 is not None and self.iid04 != "":
                approvalInfo += f'決裁:{self.aname04} 従業員ID:{self.uid04} 従業員番号:{self.empcd04} 従業員指名:{self.ename04} ;'
            if self.iid05 is not None and self.iid05 != "":
                approvalInfo += f'決裁:{self.aname05} 従業員ID:{self.uid05} 従業員番号:{self.empcd05} 従業員指名:{self.ename05} ;'
            if self.iid06 is not None and self.iid06 != "":
                approvalInfo += f'決裁:{self.aname06} 従業員ID:{self.uid06} 従業員番号:{self.empcd06} 従業員指名:{self.ename06} ;'
            if self.iid07 is not None and self.iid07 != "":
                approvalInfo += f'決裁:{self.aname07} 従業員ID:{self.uid07} 従業員番号:{self.empcd07} 従業員指名:{self.ename07} ;'
            if self.iid08 is not None and self.iid08 != "":
                approvalInfo += f'決裁:{self.aname08} 従業員ID:{self.uid08} 従業員番号:{self.empcd08} 従業員指名:{self.ename08} ;'
            if self.iid09 is not None and self.iid09 != "":
                approvalInfo += f'決裁:{self.aname09} 従業員ID:{self.uid09} 従業員番号:{self.empcd09} 従業員指名:{self.ename09} ;'
            if self.iid10 is not None and self.iid10 != "":
                approvalInfo += f'決裁:{self.aname10} 従業員ID:{self.uid10} 従業員番号:{self.empcd10} 従業員指名:{self.ename10} ;'
            if self.iid11 is not None and self.iid11 != "":
                approvalInfo += f'決裁:{self.aname11} 従業員ID:{self.uid11} 従業員番号:{self.empcd11} 従業員指名:{self.ename11} ;'
            if self.iid12 is not None and self.iid12 != "":
                approvalInfo += f'決裁:{self.aname12} 従業員ID:{self.uid12} 従業員番号:{self.empcd12} 従業員指名:{self.ename12} ;'
            if self.iid13 is not None and self.iid13 != "":
                approvalInfo += f'決裁:{self.aname13} 従業員ID:{self.uid13} 従業員番号:{self.empcd13} 従業員指名:{self.ename13} ;'
            if self.iid14 is not None and self.iid14 != "":
                approvalInfo += f'決裁:{self.aname14} 従業員ID:{self.uid14} 従業員番号:{self.empcd14} 従業員指名:{self.ename14} ;'
            if self.iid15 is not None and self.iid15 != "":
                approvalInfo += f'決裁:{self.aname15} 従業員ID:{self.uid15} 従業員番号:{self.empcd15} 従業員指名:{self.ename15} ;'
            if self.iid16 is not None and self.iid16 != "":
                approvalInfo += f'決裁:{self.aname16} 従業員ID:{self.uid16} 従業員番号:{self.empcd16} 従業員指名:{self.ename16} ;'
            yield approvalInfo

        def getItemType(self):
            return "精算明細-立替"

        def getContents(self):
            return self.contents

        def getItemSelection(self): # ※IF文右側のコードは旧MFのコード。対（つい）となる新MFZコードをreturnの右手へ書入れる事！！
            return self.itemSelection       # --- CHANGE 2021.12.25

        def getPaymentDestination(self):
            return self.paymentDestination
        
        def getCost(self):
            return self.cost

        def getTaxIncludedAmount(self):
            return self.taxIncludedAmount

        def getConsumptionTax(self):
            return self.consumptionTax

        def getBurdenDepartment(self):      # 支払部門コードであって、支払部門名ではない。※旧MFの支払部門マスターと必ずしも合致しない
            return self.burdenDepartment    # 2021.12.25～26 共通部品化

        def getVoucherFiles(self):
            yield self.voucher0
            yield self.voucher1
            yield self.voucher2
            yield self.voucher3
            yield self.voucher4
            yield self.voucher5
            yield self.voucher6
            yield self.voucher7
            yield self.voucher8
            yield self.voucher9

def click_element(web_el):
    my_sleep_click(driver.find_element_by_xpath(web_el))


def click_main(web_el):
    try:
        web_el.click()
        logger.info("Success")
    except NoSuchElementException:
        logger.error("Error")
    time.sleep(5)

def reloadBrowser():
    try:
        time.sleep(2)
        driver.refresh()
        time.sleep(2)
    except Exception as e:
        logger.info("Chromeブラウザーをリロードできず。ドライバーを確認せよ")
        # driver.quit()
    finally:
        pass

def signin_procedure(tenantId, empId, passWd):
    logger.info(            # --- ADD NEW 共通部品 20212.01.04
        "テナントID: " + str(tenantId) + " 職員番号: " + str(empId) + " 共通パスワード: " + str(passWd)
    )
    # テナントID = 100
    print(str(tenantId))
    tenant_id_textbox = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[1]/input"
    )
    print(str(empId))
    tenant_id_textbox.send_keys(
        str(tenantId)
    )
    # 職員コード
    employee_cd_textbox = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[2]/input"
    )
    employee_cd_textbox.send_keys(
        str(empId)
    )
    # パスワード（※2021.11.17時点、方針未決）
    password_textbox = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[3]/input"
    )
    password_textbox.send_keys(
        str(passWd)
    )
    # ログインボタンをクリック
    logger.info(            # --- ADD NEW 共通部品 20212.01.04
        "ログインボタンクリック"
    )
    click_element(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[4]/button"
    )
    time.sleep(2)

def sign_out_procedure():    # 2021.12.15
    # ①申請完了後の「閉じる」ボタンをクリック <--- 経費精算書のみ特殊
    # logger.info(  # --- ADD NEW 共通部品 20212.01.04
    #    "閉じるボタンクリック"
    # )
    # close_button = driver.find_element_by_xpath(
    #    "/html/body/div[1]/div/div/form/div/button[4]"
    # )
    # my_sleep_click(close_button)

    driver.implicitly_wait(3)                           # --- ADD NEW 2021.12.22
    driver.switch_to.window(driver.window_handles[0])   # --- ADD NEW 2021.12.22 要素ゼロのウィンドウへ制御を戻す！

    # ②画面右上「サインアウト」ボタン（非常出口アイコン）をクリック
    logger.info(  # --- ADD NEW 共通部品 20212.01.04
        "サインアウトアイコンクリック"
    )
    sign_out_button = driver.find_element_by_xpath(
        "/html/body/div[2]/nav/div/div[2]/ul/li[6]/form/a"
    )
    my_sleep_click(sign_out_button)
    time.sleep(2)
    # ③【ログイン】ボタンをラストにクリック
    logger.info(  # --- ADD NEW 共通部品 20212.01.04
        "サインイン画面へ戻るボタンクリック"
    )
    sign_in_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    )
    my_sleep_click(sign_in_button)
    time.sleep(2)

def entry_procedure():

    reloadBrowser()
    # click_element(
    #    "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    # )
    logger.info(  # --- ADD NEW 共通部品 20212.01.04
        "E1.xlsxファイルオープン開始"
    )
    # 経費精算申請書の伝票タイプはE1
    book = openpyxl.load_workbook("C:\XLSX\XM_SLIP_DETAIL\E1.xlsx")     # Book名がE1
    sheet = book['E1']                                                  # 経費精算書
    logger.info(  # --- ADD NEW 共通部品 20212.01.04
        "E1.xlsxファイルオープン完了"
    )

    with concurrent.futures.ProcessPoolExecutor() as executor:
        i = 3   # 1行目は見出し行。2行目から開始。A列セルをNull判定しEOFを見極める。
        while sheet.cell(row=i, column=1).value is not None:
            ex = Expense(
                                sheet.cell(row=i, column=87).value,             # CREATOR
                                '0Nu4M0%4N0',                                   # 共通パスワード（※未定）
                                sheet.cell(row=i, column=31).value,             # 利用目的【SS01】（※確証なし）
                                sheet.cell(row=i, column=44).value,             # 備考その1【SS13】（※左記は未定、要確認）
                                sheet.cell(row=i, column=45).value,             # 備考その2【SS14】（※左記は未定、要確認）
                                sheet.cell(row=i, column=256).value,            # 発生日（※処理日で良いか？ 要確認）
                                sheet.cell(row=i, column=256).value,            # 明細種別（※「精算明細-立替」を決打ちする、で良いか？ 要確認）
                                sheet.cell(row=i, column=81).value,             # 内容【ITEMNAME1】←SS01と同じ様子
                                sheet.cell(row=i, column=67).value,             # 品目選択【SS21】（[i01q-001]のような値）
                                sheet.cell(row=i, column=44).value,             # 支払先【SS14】
                                sheet.cell(row=i, column=16).value,             # 費用【PRICE1】←2が税抜額、3が消費税である様子
                                sheet.cell(row=i, column=16).value,             # 税込金額【PRICE1】←費用テキストボックスと何が異なるのか？
                                sheet.cell(row=i, column=18).value,             # 消費税額【PRICE3】
                                sheet.cell(row=i, column=6).value,              # 負担部門【BEARDEPTID】負担部門コード（例: 541→経理課）
                                sheet.cell(row=i, column=90).value,             # 証憑ファイル1
                                sheet.cell(row=i, column=91).value,             # 証憑ファイル2
                                sheet.cell(row=i, column=92).value,             # 証憑ファイル3
                                sheet.cell(row=i, column=93).value,             # 証憑ファイル4
                                sheet.cell(row=i, column=94).value,             # 証憑ファイル5
                                sheet.cell(row=i, column=95).value,             # 証憑ファイル6
                                sheet.cell(row=i, column=96).value,             # 証憑ファイル7
                                sheet.cell(row=i, column=97).value,             # 証憑ファイル8
                                sheet.cell(row=i, column=98).value,             # 証憑ファイル9
                                sheet.cell(row=i, column=99).value,              # 証憑ファイル10
                                sheet.cell(row=i, column=131).value,  # IID --- (1)
                                sheet.cell(row=i, column=132).value,  # DOCNO
                                sheet.cell(row=i, column=133).value,  # FORMNAME
                                sheet.cell(row=i, column=134).value,  # CREATED
                                sheet.cell(row=i, column=135).value,  # ANAME
                                sheet.cell(row=i, column=136).value,  # UID
                                sheet.cell(row=i, column=137).value,  # EMPCD
                                sheet.cell(row=i, column=138).value,  # ENAME
                                sheet.cell(row=i, column=139).value,  # DEPTID
                                sheet.cell(row=i, column=140).value,  # DFULLNAME
                                sheet.cell(row=i, column=141).value,  # POST
                                sheet.cell(row=i, column=142).value,  # TS
                                sheet.cell(row=i, column=144).value,  # IID --- (2)
                                sheet.cell(row=i, column=145).value,  # DOCNO
                                sheet.cell(row=i, column=146).value,  # FORMNAME
                                sheet.cell(row=i, column=147).value,  # CREATED
                                sheet.cell(row=i, column=148).value,  # ANAME
                                sheet.cell(row=i, column=149).value,  # UID
                                sheet.cell(row=i, column=150).value,  # EMPCD
                                sheet.cell(row=i, column=151).value,  # ENAME
                                sheet.cell(row=i, column=152).value,  # DEPTID
                                sheet.cell(row=i, column=153).value,  # DFULLNAME
                                sheet.cell(row=i, column=154).value,  # POST
                                sheet.cell(row=i, column=155).value,  # TS
                                sheet.cell(row=i, column=157).value,  # IID --- (3)
                                sheet.cell(row=i, column=158).value,  # DOCNO
                                sheet.cell(row=i, column=159).value,  # FORMNAME
                                sheet.cell(row=i, column=160).value,  # CREATED
                                sheet.cell(row=i, column=161).value,  # ANAME
                                sheet.cell(row=i, column=162).value,  # UID
                                sheet.cell(row=i, column=163).value,  # EMPCD
                                sheet.cell(row=i, column=164).value,  # ENAME
                                sheet.cell(row=i, column=165).value,  # DEPTID
                                sheet.cell(row=i, column=166).value,  # DFULLNAME
                                sheet.cell(row=i, column=167).value,  # POST
                                sheet.cell(row=i, column=168).value,  # TS
                                sheet.cell(row=i, column=170).value,  # IID --- (4)
                                sheet.cell(row=i, column=171).value,  # DOCNO
                                sheet.cell(row=i, column=172).value,  # FORMNAME
                                sheet.cell(row=i, column=173).value,  # CREATED
                                sheet.cell(row=i, column=174).value,  # ANAME
                                sheet.cell(row=i, column=175).value,  # UID
                                sheet.cell(row=i, column=176).value,  # EMPCD
                                sheet.cell(row=i, column=177).value,  # ENAME
                                sheet.cell(row=i, column=178).value,  # DEPTID
                                sheet.cell(row=i, column=179).value,  # DFULLNAME
                                sheet.cell(row=i, column=180).value,  # POST
                                sheet.cell(row=i, column=181).value,  # TS
                                sheet.cell(row=i, column=183).value,  # IID --- (5)
                                sheet.cell(row=i, column=184).value,  # DOCNO
                                sheet.cell(row=i, column=185).value,  # FORMNAME
                                sheet.cell(row=i, column=186).value,  # CREATED
                                sheet.cell(row=i, column=187).value,  # ANAME
                                sheet.cell(row=i, column=188).value,  # UID
                                sheet.cell(row=i, column=189).value,  # EMPCD
                                sheet.cell(row=i, column=190).value,  # ENAME
                                sheet.cell(row=i, column=191).value,  # DEPTID
                                sheet.cell(row=i, column=192).value,  # DFULLNAME
                                sheet.cell(row=i, column=193).value,  # POST
                                sheet.cell(row=i, column=194).value,  # TS
                                sheet.cell(row=i, column=196).value,  # IID --- (6)
                                sheet.cell(row=i, column=197).value,  # DOCNO
                                sheet.cell(row=i, column=198).value,  # FORMNAME
                                sheet.cell(row=i, column=199).value,  # CREATED
                                sheet.cell(row=i, column=200).value,  # ANAME
                                sheet.cell(row=i, column=201).value,  # UID
                                sheet.cell(row=i, column=202).value,  # EMPCD
                                sheet.cell(row=i, column=203).value,  # ENAME
                                sheet.cell(row=i, column=204).value,  # DEPTID
                                sheet.cell(row=i, column=205).value,  # DFULLNAME
                                sheet.cell(row=i, column=206).value,  # POST
                                sheet.cell(row=i, column=207).value,  # TS
                                sheet.cell(row=i, column=209).value,  # IID --- (7)
                                sheet.cell(row=i, column=210).value,  # DOCNO
                                sheet.cell(row=i, column=211).value,  # FORMNAME
                                sheet.cell(row=i, column=212).value,  # CREATED
                                sheet.cell(row=i, column=213).value,  # ANAME
                                sheet.cell(row=i, column=214).value,  # UID
                                sheet.cell(row=i, column=215).value,  # EMPCD
                                sheet.cell(row=i, column=216).value,  # ENAME
                                sheet.cell(row=i, column=217).value,  # DEPTID
                                sheet.cell(row=i, column=218).value,  # DFULLNAME
                                sheet.cell(row=i, column=219).value,  # POST
                                sheet.cell(row=i, column=220).value,  # TS
                                sheet.cell(row=i, column=222).value,  # IID --- (8)
                                sheet.cell(row=i, column=223).value,  # DOCNO
                                sheet.cell(row=i, column=224).value,  # FORMNAME
                                sheet.cell(row=i, column=225).value,  # CREATED
                                sheet.cell(row=i, column=226).value,  # ANAME
                                sheet.cell(row=i, column=227).value,  # UID
                                sheet.cell(row=i, column=228).value,  # EMPCD
                                sheet.cell(row=i, column=229).value,  # ENAME
                                sheet.cell(row=i, column=230).value,  # DEPTID
                                sheet.cell(row=i, column=231).value,  # DFULLNAME
                                sheet.cell(row=i, column=232).value,  # POST
                                sheet.cell(row=i, column=233).value,  # TS
                                sheet.cell(row=i, column=235).value,  # IID --- (9)
                                sheet.cell(row=i, column=236).value,  # DOCNO
                                sheet.cell(row=i, column=237).value,  # FORMNAME
                                sheet.cell(row=i, column=238).value,  # CREATED
                                sheet.cell(row=i, column=239).value,  # ANAME
                                sheet.cell(row=i, column=240).value,  # UID
                                sheet.cell(row=i, column=241).value,  # EMPCD
                                sheet.cell(row=i, column=242).value,  # ENAME
                                sheet.cell(row=i, column=243).value,  # DEPTID
                                sheet.cell(row=i, column=244).value,  # DFULLNAME
                                sheet.cell(row=i, column=245).value,  # POST
                                sheet.cell(row=i, column=246).value,  # TS
                                sheet.cell(row=i, column=248).value,  # IID --- (10)
                                sheet.cell(row=i, column=249).value,  # DOCNO
                                sheet.cell(row=i, column=250).value,  # FORMNAME
                                sheet.cell(row=i, column=251).value,  # CREATED
                                sheet.cell(row=i, column=252).value,  # ANAME
                                sheet.cell(row=i, column=253).value,  # UID
                                sheet.cell(row=i, column=254).value,  # EMPCD
                                sheet.cell(row=i, column=255).value,  # ENAME
                                sheet.cell(row=i, column=256).value,  # DEPTID
                                sheet.cell(row=i, column=257).value,  # DFULLNAME
                                sheet.cell(row=i, column=258).value,  # POST
                                sheet.cell(row=i, column=259).value,  # TS
                                sheet.cell(row=i, column=261).value,  # IID --- (11)
                                sheet.cell(row=i, column=262).value,  # DOCNO
                                sheet.cell(row=i, column=263).value,  # FORMNAME
                                sheet.cell(row=i, column=264).value,  # CREATED
                                sheet.cell(row=i, column=265).value,  # ANAME
                                sheet.cell(row=i, column=266).value,  # UID
                                sheet.cell(row=i, column=267).value,  # EMPCD
                                sheet.cell(row=i, column=268).value,  # ENAME
                                sheet.cell(row=i, column=269).value,  # DEPTID
                                sheet.cell(row=i, column=270).value,  # DFULLNAME
                                sheet.cell(row=i, column=271).value,  # POST
                                sheet.cell(row=i, column=272).value,  # TS
                                sheet.cell(row=i, column=274).value,  # IID --- (12)
                                sheet.cell(row=i, column=275).value,  # DOCNO
                                sheet.cell(row=i, column=276).value,  # FORMNAME
                                sheet.cell(row=i, column=277).value,  # CREATED
                                sheet.cell(row=i, column=278).value,  # ANAME
                                sheet.cell(row=i, column=279).value,  # UID
                                sheet.cell(row=i, column=280).value,  # EMPCD
                                sheet.cell(row=i, column=281).value,  # ENAME
                                sheet.cell(row=i, column=282).value,  # DEPTID
                                sheet.cell(row=i, column=283).value,  # DFULLNAME
                                sheet.cell(row=i, column=284).value,  # POST
                                sheet.cell(row=i, column=285).value,  # TS
                                sheet.cell(row=i, column=287).value,  # IID --- (13)
                                sheet.cell(row=i, column=288).value,  # DOCNO
                                sheet.cell(row=i, column=289).value,  # FORMNAME
                                sheet.cell(row=i, column=290).value,  # CREATED
                                sheet.cell(row=i, column=291).value,  # ANAME
                                sheet.cell(row=i, column=292).value,  # UID
                                sheet.cell(row=i, column=293).value,  # EMPCD
                                sheet.cell(row=i, column=294).value,  # ENAME
                                sheet.cell(row=i, column=295).value,  # DEPTID
                                sheet.cell(row=i, column=296).value,  # DFULLNAME
                                sheet.cell(row=i, column=297).value,  # POST
                                sheet.cell(row=i, column=298).value,  # TS
                                sheet.cell(row=i, column=300).value,  # IID --- (14)
                                sheet.cell(row=i, column=301).value,  # DOCNO
                                sheet.cell(row=i, column=302).value,  # FORMNAME
                                sheet.cell(row=i, column=303).value,  # CREATED
                                sheet.cell(row=i, column=304).value,  # ANAME
                                sheet.cell(row=i, column=305).value,  # UID
                                sheet.cell(row=i, column=306).value,  # EMPCD
                                sheet.cell(row=i, column=307).value,  # ENAME
                                sheet.cell(row=i, column=308).value,  # DEPTID
                                sheet.cell(row=i, column=309).value,  # DFULLNAME
                                sheet.cell(row=i, column=310).value,  # POST
                                sheet.cell(row=i, column=311).value,  # TS
                                sheet.cell(row=i, column=313).value,  # IID --- (15)
                                sheet.cell(row=i, column=314).value,  # DOCNO
                                sheet.cell(row=i, column=315).value,  # FORMNAME
                                sheet.cell(row=i, column=316).value,  # CREATED
                                sheet.cell(row=i, column=317).value,  # ANAME
                                sheet.cell(row=i, column=318).value,  # UID
                                sheet.cell(row=i, column=319).value,  # EMPCD
                                sheet.cell(row=i, column=320).value,  # ENAME
                                sheet.cell(row=i, column=321).value,  # DEPTID
                                sheet.cell(row=i, column=322).value,  # DFULLNAME
                                sheet.cell(row=i, column=323).value,  # POST
                                sheet.cell(row=i, column=324).value,  # TS
                                sheet.cell(row=i, column=326).value,  # IID --- (16)
                                sheet.cell(row=i, column=327).value,  # DOCNO
                                sheet.cell(row=i, column=328).value,  # FORMNAME
                                sheet.cell(row=i, column=329).value,  # CREATED
                                sheet.cell(row=i, column=330).value,  # ANAME
                                sheet.cell(row=i, column=331).value,  # UID
                                sheet.cell(row=i, column=332).value,  # EMPCD
                                sheet.cell(row=i, column=333).value,  # ENAME
                                sheet.cell(row=i, column=334).value,  # DEPTID
                                sheet.cell(row=i, column=335).value,  # DFULLNAME
                                sheet.cell(row=i, column=336).value,  # POST
                                sheet.cell(row=i, column=337).value  # TS
            )

            # 引数: 左からテナントID、職員番号、パスワード
            gen = ex.getSignInInfo()
            # signin_procedure(gen.__next__(), gen.__next__(), gen.__next__())   <----- サインイン方法確定後、有効化せよ
            gen = None

            click_element(
               "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
            )

            click_element(
                "/html/body/div[2]/div[1]/ul[1]/li[4]/a"
            )

            logger.info(  # --- ADD NEW 共通部品 20212.01.04
                "●【経費】経費精算書ラジオボタン選択"
            )
            click_element(
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[5]/label[5]/div/div[1]/input"
            )

            driver.implicitly_wait(5)       # 引数10を5へ減数

            # print(driver.current_url)     # --- DELETE 2021.12.10 --- 前任者による削除忘れと思料される 

            click_element(
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[3]/button"
            )

            driver.implicitly_wait(5)       # 引数10を5へ減数
            driver.switch_to.window(driver.window_handles[1])
            # create_button = driver.find_element_by_xpath('//*[text()="起票する"]')

            create_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[3]/button[2]"
            )
            my_sleep_click(create_button)

            # PURPOSE = "利用目的"                                  # --- DELETE 2021.12.10 --- 前任者による定数エントリー（※場所移動）
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "利用目的: " + str(ex.getPorposeOfUse())
            )
            purpose_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div/input"
            )
            # purpose_fld.send_keys(PURPOSE)                        # --- DELETE 2021.12.10
            purpose_fld.send_keys(ex.getPorposeOfUse())

            # --- ADD NEW 2021.12.13 --- 前任者不作為につき新規追加 --- Start
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "備考: " + str(ex.getRemark())
            )
            gen = ex.getRemark()
            remark_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div/textarea"
            )
            remark_fld.send_keys(
                str(gen.__next__()) + '\r\n' + str(gen.__next__()) + '\r\n' + str(gen.__next__())    # CRLF改行
            )
            gen = None
            # --- ADD NEW 2021.12.13 --- 前任者不作為につき新規追加 --- End

            # DETAIL_TYPE = "精算明細-立替"                         # --- DELETE 2021.12.10 --- 前任者による定数エントリー
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "明細: " + str(ex.getItemType())
            )
            detail_type_select_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div/div[2]/div/div[" "2]/div[2]/select"
            )
            # detail_type_select_field.send_keys(DETAIL_TYPE)       # --- DELETE 2021.12.10
            detail_type_select_field.send_keys(ex.getItemType())    # --- UPDATE 2021.12.10

            detail_type_select_field.send_keys(
                Keys.TAB
                 + Keys.ENTER
            )

            # Now Input fields from here    # <--- 意味不明の前任者コメント
            # CONTENT = "内容（●●購入代、飲食代等）"                # --- DELETE 2021.12.10 --- 前任者による定数エントリー
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "内容: " + str(ex.getContents())
            )
            content_input = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div["
                "2]/div[1]/input"
            )
            # content_input.send_keys(CONTENT)                      # --- DELETE 2021.12.10
            content_input.send_keys(ex.getContents())               # --- UPDATE 2021.12.10

            main_page = driver.current_window_handle
            # Enter to frame
            content_input.send_keys(
                Keys.TAB
                 + Keys.TAB
                 + Keys.ENTER
            )

            # ITEM_CODE = "i10r-80101"                             # --- DELETE 2021.12.10 --- 前任者による定数エントリー
            item_get_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div["
                "3]/div/div[2]/div[2]/div/span/button[2]"
            )

            driver.implicitly_wait(3)       # 引数10を3へ減数 

            iframe = driver.find_element_by_xpath('//*[@id="zk_iFrameDialog"]/iframe')
            driver.switch_to.frame(iframe)

            item_code_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[1]/input"
            )
            gen = asj.returnAccountingSubject(ex.getItemSelection())     # --- UPDATE 2021.12.26 共通部品使用へ変更。ジールドで値を取得
            wkXpath = gen.__next__()                        # --- テーブルのクリックする段数（変数へ保存）
            wkItemCd = gen.__next__()                       # --- アイテムコード（変数へ保存）
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "段数: " + str(wkXpath) + " 品目コード: " + str(wkItemCd)
            )
            item_code_fld.send_keys(
                 # ex.getItemSelection()    # --- UPDATE 2021.12.10
                 wkItemCd                   # --- UPDATE 2021.12.26
                 + Keys.TAB                 # --- ADD NEW 2021.12.10 --- Start
                 + Keys.TAB                 #
                 + Keys.TAB                 # --- ADD NEW 2021.12.10 --- End
                 + Keys.ENTER
            )

            time.sleep(3)                   # 引数5を3へ減数
            open_all = driver.find_element_by_xpath(        # すべて開くボタン
                "/html/body/div[1]/div/div[1]/div[2]/button"
            )
            my_sleep_click(open_all)

            # item_row = driver.find_element_by_xpath(                            # --- INVALID 2021.12.26 Start
            #     "/html/body/div[1]/div/div[1]/div[4]/div/div/div/label/div"     #     前任者仕様固定値
            # )                                                                   # --- INVALID 2021.12.26 End
            if wkXpath == 2:
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/label/div/div[1]"
                )
            elif wkXpath == 3:
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/div/label/div/div[1]"
                )
            elif wkXpath == 4:
                item_row = driver.find_element_by_xpath(
                    "/html/body/div[1]/div/div[1]/div[4]/div/div/div/div/label/div/div[1]"
                )
            my_sleep_click(item_row)
            # close_button = driver.find_element_by_xpath('/html/body/div[1]/div/div[2]/form/button')   # 前任者による無効化
            # my_sleep_click(close_button)

            driver.switch_to.default_content()

            # PAY_TO = "支払先"                                      # --- DELETE 2021.12.10 --- 前任者による定数エントリー
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "支払先: " + str(ex.getPaymentDestination())
            )
            pay_to_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div["
                "3]/div[1]/input"
            )
            # pay_to_fld.send_keys(PAY_TO)                          # --- DELETE 2021.12.10
            pay_to_fld.send_keys(ex.getPaymentDestination())        # --- UPDATE 2021.12.10

            # EXPENSE = "1100"                                      # --- DELETE 2021.12.10 --- 前任者による定数エントリー
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "料金: " + str(ex.getCost())
            )
            expense_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[3]/div[2]/input"
            )
            # expense_fld.send_keys(EXPENSE)                        # --- DELETE 2021.12.10
            expense_fld.send_keys(ex.getCost())                     # --- UPDATE 2021.12.10

            # AMOUNT_WITH_TAX = '1100'                              # --- 前任者による無効化 --- Start
            # amount_with_tax_fld = driver.find_element_by_xpath(
            #   '/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[1]/input'
            # )
            # amount_with_tax_fld.send_keys(AMOUNT_WITH_TAX)        # --- 前任者による無効化 --- End

            # --- ADD NEW 2021.12.13 --- 前任者不作為につき新規追加 --- Start
            expense_fld.send_keys(
                Keys.TAB
            )
            time.sleep(2)

            tax_included_amount_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[1]/input"
            )
            tax_included_amount_fld.clear()
            tax_included_amount_fld.send_keys(ex.getTaxIncludedAmount())
            time.sleep(2)
            tax_included_amount_fld.send_keys(
                Keys.TAB
            )
            time.sleep(2)
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "消費税: " + str(ex.getConsumptionTax())
            )
            consumption_tax_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[3]/input"
            )
            consumption_tax_fld.clear()
            time.sleep(2)
            consumption_tax_fld.send_keys(ex.getConsumptionTax())
            consumption_tax_fld.send_keys(
                Keys.TAB            # 消しゴムボタン
                + Keys.TAB          # 負担部門ボタン
                + Keys.ENTER        # 改行キー打鍵
            )
            time.sleep(2)
            # iframeへ画面遷移
            iframe = driver.find_element_by_xpath('//*[@id="zk_iFrameDialog"]/iframe')
            driver.switch_to.frame(iframe)
            driver.implicitly_wait(2)

            logger.info(  # --- ADD NEW 個別 20212.01.04
                "旧MF 支払部門コード: " + str(ex.getBurdenDepartment())
            )
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "新MFZ支払部門コード: " + str(dpt.returnDepartment(ex.getBurdenDepartment()))
            )
            burden_department_code_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[1]/input"
            )
            # burden_department_code_fld.send_keys(ex.getBurdenDepartment())                        # --- INVALID 2021.12.26
            burden_department_code_fld.send_keys(dpt.returnDepartment(ex.getBurdenDepartment()))    # --- UPDATE 2021.12.26
            time.sleep(2)
            burden_department_code_fld.send_keys(
                Keys.TAB        # 負担部門名テキストボックスへカーソル遷移
                + Keys.TAB      # 虫眼鏡ボタンへカーソル遷移
                + Keys.ENTER    # 改行キー打鍵
            )
            time.sleep(2)
            burden_department_table = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[1]/div[3]/label/div/div[1]"
            )
            my_sleep_click(burden_department_table)
            time.sleep(2)
            driver.switch_to.default_content()          # 制御を親ウィンドウへ戻す
            # --- ADD NEW 2021.12.13 --- 前任者不作為につき新規追加 --- End

            confirm_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[5]/div/button[1]"
            )
            my_sleep_click(confirm_button)
            time.sleep(2)
            driver.save_screenshot(f'C:/Users/digiworker_biz_02/PycharmProjects/pythonProject/ScreenShot/E1/{i}_確定ボタンクリック.png')

            # Next attach the evidence

            # Attach File function
            # attach_file_button = driver.find_element_by_xpath(
            #     "/html/body/div[1]/div[2]/div/form/div[5]/div/ul/li[3]/a"
            # )
            # my_sleep_click(attach_file_button)
            # EXPENSE_SETTLE_FILE = "C:\Output\MFZ\Test01.pdf"
            # select_file_button = driver.find_element_by_xpath(
            #     "/html/body/div[1]/div[2]/div/form/div[5]/div/div/div["
            #     "3]/div/div/div/div[2]/div[2]/div/span/button"
            # )
            # my_sleep_click(select_file_button)
            # upload_file_button = driver.find_element_by_xpath(
            #     "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[2]/div[2]"
            # )
            # upload_file_button.click()
            # time.sleep(4)
            # pyautogui.write(EXPENSE_SETTLE_FILE)
            # pyautogui.press("enter")
            # time.sleep(4)
            # add_green_button = driver.find_element_by_xpath(
            #     "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[3]/div/a"
            # )
            # my_sleep_click(add_green_button)
            # confirm_button = driver.find_element_by_xpath(
            #     "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"
            # )
            # my_sleep_click(confirm_button)

            gen = ex.getVoucherFiles()                                      # Generatorへ格納し、next()で取得する
            submit_voucher_files(gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__()
            )
            gen = None  # 参照を切離す
            ex = None   # 参照を切離す
            final_register_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[11]/button[1]"
            )
            click_submit_button(driver, logger, "Validation", final_register_button)
            # --- 2022.01.05 ADD NEW 【申請】ボタンクリック後、却下される可能性がある為、スクリーンショットを取得、保存する
            driver.save_screenshot(f'C:/Users/digiworker_biz_02/PycharmProjects/pythonProject/ScreenShot/E1/{i}_申請ボタンクリック.png')
            time.sleep(2)

            # --- ADD NEW 2021.12.23 通知チェックボックスのレ点を解除
            notificationCheckBox = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/form/div[3]/label/input[1]"
            )
            my_sleep_click(notificationCheckBox)

            # 前任者による意図的なコメントエントリー。実運用時は除外せよ！！
            comment_field = driver.find_element_by_xpath(
                # "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div/textarea" # --- INVALID 2021.12.26
                "/html/body/div[1]/div/div/form/textarea"   # --- CHANGE 2012.12.26 変更があったのか？
            )
            comment_field.send_keys(MIGRATION_COMMENT)

            # Final submission
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "実行ボタンクリック"
            )
            final_register_button = driver.find_element_by_xpath(
                # "/html/body/div[1]/div[2]/div/form/div[11]/button[1]"
                "/html/body/div[1]/div/div/form/div[4]/button[1]"
            )
            # TODO need to uncomment it later
            # click_submit_button(driver, logger, 'Validation', final_register_button)
            my_sleep_click(final_register_button)  # <--- UPDATE 2021.12.23     # 左記を正とする

            final_submit_button = driver.find_element_by_xpath(
                # "/html/body/div[1]/div/div/form/div[4]/button[1]"
                "/html/body/div[1]/div/div/form/div/button[3]"
            )
            my_sleep_click(final_submit_button)  # <--- UPDATE 2021.12.23     # 左記を正とする

            # TODO uncomment below after testing
            # click_main(driver, logger, 'Submission', final_register_button)

            sign_out_procedure()

            # ガーベジコレクター
            # if i % 100 == 0:  --- INVALID 2022.01.05
            gc.collect()
            # T1シートの次行へ移動
            i += 1

def submit_voucher_files(vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9): 
    logger.info(            # --- ADD NEW 共通部品 20212.01.04
        "添付ファイル - 選択ボタンクリック"
    )
    attach_file_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div[2]/div/form/div[5]/div/ul/li[3]/a"
    )
    my_sleep_click(attach_file_button)
    # 配列へ格納 --- ※証憑はPDFのみとは限らない
    va = [vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9]

    j = 0  # ゼロオリジン
    if va[0] != "N/A":      # 一つ目のパスが無い場合は2つ目以降のパスも存在しないと見做す
        while va[j] is not None and j < 10:
            if j == 0:  # 2回目以降は不要である為
                select_file_button = driver.find_element_by_xpath(
                    # 添付ファイルタブに所在する添付ファイル-選択ボタン - 2021.11.15 コメント記入
                    "/html/body/div[1]/div[2]/div/form/div[5]/div/div/div["
                    "3]/div/div/div/div[2]/div[2]/div/span/button"
                )
                my_sleep_click(select_file_button)

            logger.info(  # --- ADD NEW 共通部品 20212.01.04
                "証憑ファイルパス: " + va[j]
            )
            upload_file_button = driver.find_element_by_xpath(
            # 次画面にある「添付ファイルを選択してください」→選択ボタンをクリック
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[2]/div[2]"
            )
            upload_file_button.click()
            time.sleep(2)
            pyperclip.copy(va[j])                      # Clip Boardへコピーし、
            pg.hotkey('ctrl', 'v')              # Pasteする。
            pg.press("enter")
            time.sleep(2)
            add_green_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[3]/div/a"
            )
            my_sleep_click(add_green_button)
            time.sleep(2)
            # 右隣セルの値へ移動
            j += 1
    else:
        logger.info(  # --- ADD NEW 共通部品 20212.01.04
            "対象となる証憑ファイルは無し"
        )

    # 証憑提出後、確定ボタンクリック。ただし1件以上ある事が前提
    if j != 0:  # --- ADD NEW 2021.12.24 --- 証憑が一つ以上あった場合のみ確定ボタンをクリック
        logger.info(  # --- ADD NEW 共通部品 20212.01.04
            "添付ファイルタブ - 確定ボタンクリック"
        )
        final_register_button = driver.find_element_by_xpath(
            # "/html/body/div[1]/div[2]/div/form/div[11]/button[1]"
            "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"
        )
        my_sleep_click(final_register_button)

def main():
    entry_procedure()
    logger.info("Robot completed")

main()
