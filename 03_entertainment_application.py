# 【経費】交際費申請書
import logging
import pyautogui as pg  # --- UPDATE 2021.12.23
pg.FAILSAFE = False     # --- UPDATE 2021.12.23

from chrome_driver_dl import get_latest_driver
from common import *

# --- ADD 2021.12.01 Start
import openpyxl
import sys
import pyperclip
import gc
from datetime import datetime
import concurrent.futures   # --- ADD 2021.12.02
import re                   # --- ADD 2021.12.02

from selenium import webdriver
driver = webdriver.Chrome(
    executable_path="C:/Users/digiworker_biz_02/PycharmProjects/pythonProject/Chrome/chromedriver.exe"
)
# --- ADD 2021.12.01 End

# Gets or creates a logger
logger = logging.getLogger("03")
# set log level
logger.setLevel(logging.INFO)
# define file handler and set formatter
file_handler = logging.FileHandler("../Logs/03_entertainment_application.log")
formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")
file_handler.setFormatter(formatter)
# add file handler to logger
logger.info("Started")
# driver = webdriver.Chrome(get_latest_driver())    # --- 2021.12.01 Delete
driver.get(MFZC_URL)
# driver.fullscreen_window()                        # ---　有効化してはならない！！
file_handler.setFormatter(formatter)
# add file handler to logger
logger.addHandler(file_handler)

def reloadBrowser():        # --- ADD 2021.12.02 Chrome原因不明不調につき急遽追記 - Start
    try:
        time.sleep(2)
        driver.refresh()
        time.sleep(2)
    except Exception as e:
        logger.info("Chromeブラウザーをリロードできない実行時例外。ドライバーを確認せよ")
        logger.info(e)
        # driver.quit()
    finally:
        pass                # --- ADD 2021.12.02 Chrome原因不明不調につき急遽追記 - End

class EntertainmentExpence:
    def __init__(self,
                wkUserId,
                wkPassWd,
                wkOurParticipantsAffiliation,
                wkOurParticipants,
                wkParticipants,
                wkPosition,
                wkPublicEnemyCd,
                wkEventDate,
                wkNumberOfOurCompany,
                wkNumberOfRecipients,
                wkTotalNumberOfPeople,
                wkAim,
                wkMainContent,
                wkPaymentDestPublicEnemyCd,
                wkStoreName,
                wkPaymentAddress,
                wkScheduledProvisionalPayment,
                wkTemporaryPaymentType,
                wkExpectedPaymentAmount,
                wkDesiredTemporaryPaymentDate,
                wkScheduledTemporaryPaymentSettlementDate,
                wkRemarkA,
                wkRemarkB,
                wkRemarkC,
                wkVoucher0,
                wkVoucher1,
                wkVoucher2,
                wkVoucher3,
                wkVoucher4,
                wkVoucher5,
                wkVoucher6,
                wkVoucher7,
                wkVoucher8,
                wkVoucher9           
        ):
        self.userID = wkUserId
        self.passWd = wkPassWd
        self.ourParticipantsAffiliation = wkOurParticipantsAffiliation
        self.ourParticipants = wkOurParticipants
        self.participants = wkParticipants
        self.position = wkPosition
        self.publicEnemyCd = wkPublicEnemyCd
        self.eventDate = wkEventDate
        self.numberOfOurCompany = wkNumberOfOurCompany
        self.numberOfRecipients = wkNumberOfRecipients
        self.totalNumberOfPeople = wkTotalNumberOfPeople
        self.aim = wkAim
        self.mainContent = wkMainContent
        self.paymentDestPublicEnemyCd = wkPaymentDestPublicEnemyCd
        self.storeName = wkStoreName
        self.paymentAddress = wkPaymentAddress
        self.scheduledProvisionalPayment = wkScheduledProvisionalPayment
        self.temporaryPaymentType = wkTemporaryPaymentType
        self.expectedPaymentAmount = wkExpectedPaymentAmount
        self.desiredTemporaryPaymentDate = wkDesiredTemporaryPaymentDate
        self.scheduledTemporaryPaymentSettlementDate = wkScheduledTemporaryPaymentSettlementDate
        self.remarkA = wkRemarkA
        self.remarkB = wkRemarkB
        self.remarkC = wkRemarkC
        self.voucher0 = wkVoucher0
        self.voucher1 = wkVoucher1
        self.voucher2 = wkVoucher2
        self.voucher3 = wkVoucher3
        self.voucher4 = wkVoucher4
        self.voucher5 = wkVoucher5
        self.voucher6 = wkVoucher6
        self.voucher7 = wkVoucher7
        self.voucher8 = wkVoucher8
        self.voucher9 = wkVoucher9

    def getSignInInfo(self):
        yield 100
        yield self.userID
        yield self.passWd

    def getOurParticipants(self):
        if self.ourParticipantsAffiliation is None:
            yield ''
        else:
            yield self.ourParticipantsAffiliation
        if self.ourParticipants is None:
            yield ''
        else:
            yield self.ourParticipants
    
    def getParticipants(self):
        if self.participants is None or self.participants == '':
            return 'N/A'
        else:
            return self.participants

    def getPosition(self):
        # おきゃくさまーポストは省略できない！！ --- 2021.12.22（おきゃくさまはー（復唱）、かみさまー（復唱）！　こころからのー（復唱）、えがおでー（復唱）！）
        # if (self.position in '-') == True or (self.position in '一般') == True or (self.position in '主任') == True:
        #          return str('課長未満・他一般等')
        # elif (self.position in '課長') == True or (self.position in '部長') == True or  (self.position in '所長') == True:
        #          return str('部課長クラス・顧問等')
        # elif (self.position in '社長') == True:       # 元データはこの限り
        #        return str('取締役・執行役員クラス')
        # else:
        #     return str('課長未満・他一般等')
        return str('課長未満・他一般等')     # --- CHANGE 2021.12.23 --- 必ずいずれかを指定せねばならない
        # return ''       # --- CHANGE 2021.12.21 --- 何もしない

    def getPublicEnemyCd(self):
        if self.publicEnemyCd is not None:
            return self.publicEnemyCd
        else:
            return 'N/A'      # --- CHANGE 2021.12.27
            # return '　'          # --- CHANGE 2021.12.27 --- 全角スペースであっても不可
    
    def getEventDate(self):
        if self.eventDate is None or self.eventDate == 0:
            today = datetime.now()
            return today.strftime('%Y-%m-%d')           # 本日日付を戻す
        else:
            # Excelセルから時分秒を取得して仕舞う為、当該時分秒を除外する処理を施す
            return datetime.strptime(str(self.eventDate).replace('-',''), '%Y%m%d %H:%M:%S').strftime('%Y-%m-%d')

    def getNumberOfOurCompany(self):
        # if self.numberOfOurCompany is None or self.numberOfOurCompany == 0:
        #    return 1    # 必ず1名居ると見做す
        #else:
        #    return self.numberOfOurCompany
        return self.numberOfOurCompany.count(';') + 1   # --- UPDATE 2021.12.23 --- 例: ;をふたつ含有→三人とカウント

    def setNumberOfOurCompany(self, num):               # 会社参加者のSetter
        self.numberOfOurCompany = num

    def getNumberOfRecipients(self):
        if self.numberOfRecipients is None or self.numberOfRecipients == 0:
            return 1    # 必ず1名居ると見做す
        else:
            return self.numberOfRecipients
    
    def setNumberOfRecipients(self, num):               # 先方参加者のSetter
        self.numberOfRecipients = num

    def getTotalNumberOfPeople(self):
        if self.totalNumberOfPeople is None or self.totalNumberOfPeople == 0:
            return int(self.numberOfOurCompany) + int(self.numberOfRecipients)  # 合算を返す
        else:
            return self.totalNumberOfPeople             # ※合算未満の場合は妥当でなくなる。納品先要聴取事項
    
    def getAim(self):
        return str(self.aim)[0:49]      # --- CHANGE 2021.12.23 --- 50文字制限対応
    
    def getMainContent(self):
        return self.mainContent
    
    def getPaymentDestPublicEnemyCd(self):
        if self.paymentDestPublicEnemyCd is None:
            return 'N/A'      # --- CHANGE 2021.12.27
            # return '　'          # --- CHANGE 2021.12.27 --- 全角スペースであっても不可
        else:
            return self.paymentDestPublicEnemyCd
    
    def getStoreName(self):
        if self.storeName is None:
            return 'N/A'      # --- CHANGE 2021.12.27
            # return '　'          # --- CHANGE 2021.12.27 --- 全角スペースであっても不可
        else:
            return self.storeName
    
    def getPaymentAddress(self):
        if self.paymentAddress is None or self.paymentAddress != "":
            return 'N/A'      # --- CHANGE 2021.12.27
            # return '　'        # --- CHANGE 2021.12.27 --- 全角スペースであっても不可
        else:
            return self.paymentAddress

    def getScheduledProvisionalPayment(self):
        return self.scheduledProvisionalPayment

    def getTemporaryPaymentType(self):
        # if self.temporaryPaymentType == 0:    # --- CHANGE 2021.12.27
        if self.temporaryPaymentType == 1:      # --- CHANGE 2021.12.27
            return str('現金')                # ※旧MFにあっては(0)なし、(1)現金、(2)銀行振込。新MFZにあっては「現金」、「振込」のみ。
        else:                                # ※納品先へ聴取し要確認！！
            return str('振込')                # 1以外を振込と仮定

    def getExpectedPaymentAmount(self):
        return self.expectedPaymentAmount

    def getDesiredTemporaryPaymentDate(self):
        if self.desiredTemporaryPaymentDate is None or self.desiredTemporaryPaymentDate == 0:
            today = datetime.now()
            return today.strftime('%Y-%m-%d')           # 本日日付を戻す
        else:
            # Excelセルから時分秒を取得して仕舞う為、当該時分秒を除外する処理を施す
            return datetime.strptime(str(self.desiredTemporaryPaymentDate).replace('-',''), '%Y%m%d %H:%M:%S').strftime('%Y-%m-%d')

    def getScheduledTemporaryPaymentSettlementDate(self):
        if self.scheduledTemporaryPaymentSettlementDate is None or self.scheduledTemporaryPaymentSettlementDate == 0:
            today = datetime.now()
            return today.strftime('%Y-%m-%d')           # 本日日付を戻す
        else:
            # Excelセルから時分秒を取得して仕舞う為、当該時分秒を除外する処理を施す
            return datetime.strptime(str(self.scheduledTemporaryPaymentSettlementDate).replace('-',''), '%Y%m%d %H:%M:%S').strftime('%Y-%m-%d')

    def getRemark(self):
        if self.remarkA is None:
            self.remarkA = ''
        if self.remarkB is None:
            self.remarkB = ''
        if self.remarkC is None:
            self.remarkC = ''
        # yield str(self.remarkA)
        # yield str(self.remarkB)
        # yield str(self.remarkC)
        crlf = '\r\n'
        return str(str(self.remarkA) + crlf + str(self.remarkB) + crlf + str(self.remarkC))[0:199]

    def getVoucherFiles(self):
        yield self.voucher0
        yield self.voucher1
        yield self.voucher2
        yield self.voucher3
        yield self.voucher4
        yield self.voucher5
        yield self.voucher6
        yield self.voucher7
        yield self.voucher8
        yield self.voucher9

def signin_procedure(tenantId, empId, passWd):
    # テナントID = 100
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[1]/input"
    )
    first_input.send_keys(tenantId)
    # 職員番号
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[2]/input"
    )
    first_input.send_keys(empId)
    # パスワード（※2021.11.17時点、方針未決）
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[3]/input"
    )
    first_input.send_keys(passWd)
    # ログインボタンをクリック
    click_element(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[5]/button"
    )

def click_element(web_el):
    my_sleep_click(
        driver.find_element_by_xpath(web_el)
    )

def click_main(web_el):
    try:
        web_el.click()
        logger.info("Success")
    except NoSuchElementException:
        logger.error("Error")
    time.sleep(3)           # 引数5を3へ減数

def entry_procedure():

    reloadBrowser()     # --- ADD NEW 2021.12.02
    # click_element(    # --- MOVE 2021.12.23
    #     "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    # )
    # 交際費申請書はSV
    book = openpyxl.load_workbook("C:\XLSX\XM_SLIP\SV.xlsx")    # Book名がSV
    sheet = book['SV']                                          # Sheet名がSV

    with concurrent.futures.ProcessPoolExecutor() as executor:
        i = 2   # 1行目は見出し行。2行目から開始。A列セルをNull判定。Application for entertainment expenses
        while sheet.cell(row=i, column=1).value is not None:
            ee = EntertainmentExpence(sheet.cell(row=i, column=9).value,    # CREATOR
                                '0Nu4M0%4N0',                               # 共通パスワード（※未定）
                                sheet.cell(row=i, column=19).value,         # 参加者所属【BEARDEPTNAME】
                                sheet.cell(row=i, column=13).value,         # 当方参加者【CREATERJ】
                                sheet.cell(row=i, column=82).value,         # 先方情報【SS16】
                                sheet.cell(row=i, column=26).value,         # お客さまーポスト【POST】（！！ 元データのPOSTは納品先の情報であろう。納品先要確認）
                                sheet.cell(row=i, column=256).value,        # 反社番号（不明。反社-1910-0013のような番号）
                                sheet.cell(row=i, column=69).value,         # 開催日（※【SS03】と考えられる）
                                sheet.cell(row=i, column=81).value,         # 当社人数（※未定）【SS15】内のコロン「;」の数プラスワンでカウントする
                                sheet.cell(row=i, column=72).value,         # 先方人数（※未定）【SS06】を仮に設定                        
                                sheet.cell(row=i, column=256).value,        # 合計人数（※当社人数と先方人数の合算で良いと思料される）
                                sheet.cell(row=i, column=83).value,         # 目的【SS17】
                                sheet.cell(row=i, column=79).value,         # 主たる内容【SS13】
                                sheet.cell(row=i, column=125).value,        # 支払先反社番号 <--- 2021.12.21 手動調整
                                sheet.cell(row=i, column=126).value,        # 店舗名 <--- 2021.12.21 手動調整
                                sheet.cell(row=i, column=127).value,        # 支払先住所 <--- 2021.12.21 手動調整
                                sheet.cell(row=i, column=39).value,         # 仮払予定額【TEMP_PRICE】
                                sheet.cell(row=i, column=46).value,         # 仮払種別【PAY_TYPE】
                                sheet.cell(row=i, column=41).value,         # 仮払金額【PAY_PRICE】
                                sheet.cell(row=i, column=69).value,         # 仮払希望日【SS03】（※中身はゼロばかりである）
                                sheet.cell(row=i, column=71).value,         # 仮払精算予定日【SS05】  （※ゼロまたはスペースばかりである）
                                sheet.cell(row=i, column=82).value,         # 備考A【SS16】 --- ADD NEW 2021.12.22
                                sheet.cell(row=i, column=83).value,         # 備考A【SS17】 --- ADD NEW 2021.12.22
                                sheet.cell(row=i, column=86).value,         # 備考C【SS20】（※値を持つ場合もあるが、スペースの場合もある、注意）
                                sheet.cell(row=i, column=109).value,        # 証憑ファイル1
                                sheet.cell(row=i, column=110).value,        # 証憑ファイル2
                                sheet.cell(row=i, column=111).value,        # 証憑ファイル3
                                sheet.cell(row=i, column=112).value,        # 証憑ファイル4
                                sheet.cell(row=i, column=113).value,        # 証憑ファイル5
                                sheet.cell(row=i, column=114).value,        # 証憑ファイル6
                                sheet.cell(row=i, column=115).value,        # 証憑ファイル7
                                sheet.cell(row=i, column=116).value,        # 証憑ファイル8
                                sheet.cell(row=i, column=117).value,        # 証憑ファイル9
                                sheet.cell(row=i, column=118).value         # 証憑ファイル10
            )

            # 引数: 左からテナントID、職員番号、パスワード
            gen = ee.getSignInInfo()
            # signin_procedure(gen.__next__(), gen.__next__(), gen.__next__())   <----- サインイン方法確定後、有効化せよ！！
            gen = None

            # 当該ボタンはテナントID直下のログインボタンであり、サインイン・サインアウトが出来ない、検証時の避難措置である。正稼働時は無効化せよ！！
            click_element(
                "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
            )

            # 左ペイン起票ボタンクリック（仲林コメント）
            click_element(
                "/html/body/div[2]/div[1]/ul[1]/li[4]/a"
            )

            click_element(
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[5]/label[3]/div/div[1]/input"
            )

            driver.implicitly_wait(5)      # 引数10から5へ減数

            click_element(
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[3]/button"
            )

            driver.implicitly_wait(5)       # 引数10から5へ減数
            driver.switch_to.window(driver.window_handles[1])
            # create_button = driver.find_element_by_xpath('//*[text()="起票する"]')    <--- 前任者によって無効化

            create_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[3]/button[2]"
            )
            my_sleep_click(create_button)

            # THIS_COMPANY_PARTICIPANT = "当社参加者（課、氏名"                             # --- DELETE 2021.12.01 前任者による定数エントリー
            this_company_participant_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div/input"
            )
            # this_company_participant_fld.send_keys(THIS_COMPANY_PARTICIPANT)              # --- DELETE 2021.12.01
            gen = ee.getOurParticipants()                                                   # --- UPDATE 2021.12.01
            this_company_participant_fld.send_keys((gen.__next__() + '　' + gen.__next__()).strip())     # --- UPDATE 2021.12.23

            # OTHER_COMPANY_PARTICIPANT = "先方情報（社名・氏名）"                           # --- DELETE 2021.12.01 前任者による定数エントリー
            other_company_participant_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[" "2]/div/div[3]/div[1]/textarea"
            )
            # other_company_participant_fld.send_keys(OTHER_COMPANY_PARTICIPANT)            # --- DELETE 2021.12.01
            other_company_participant_fld.send_keys(str(ee.getParticipants()))              # --- UPDATE 2021.12.01

            # --- 2021.12.21 COMMENT --- 先方側のポストを抽出できぬ（自分側はできる）。ただし処理をスキップできぬ！ --- 2021.12.22
            posision_listbox = driver.find_element_by_xpath(                                # --- ADD NEW 2021.12.02
                 "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div[2]/select"  # --- 前任者は当該リストボックスを処置せず
            )
            posision_listbox.send_keys(ee.getPosition())                                    # --- ADD NEW 2021.12.02

            # OTHER_COMPANY_ANTI_NO = "先方反社番号"                             `            # --- DELETE 2021.12.01 前任者による定数エントリー
            other_company_anti_no_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[" "3]/div[3]/input"
            )
            # other_company_anti_no_fld.send_keys(OTHER_COMPANY_ANTI_NO)                    # --- DELETE 2021.12.01
            other_company_anti_no_fld.send_keys(ee.getPublicEnemyCd())                      # --- UPDATE 2021.12.01

            # DATE_ORGANISED = "2021-08-01"                                                 # --- DELETE 2021.12.01 前任者による定数エントリー
            date_organised_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[4]/div[" "1]/div/input"
            )
            # date_organised_field.send_keys(DATE_ORGANISED)                                # --- DELETE 2021.12.01
            date_organised_field.send_keys(ee.getEventDate())                               # --- UPDATE 2021.12.01

            # THIS_COMPANY_PARTICIPANT_NO = "2"                                             # --- DELETE 2021.12.01 前任者による定数エントリー
            this_company_participant_no_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[4]/div[2]/input"
            )
            # this_company_participant_no_field.send_keys(THIS_COMPANY_PARTICIPANT_NO)      # --- DELETE 2021.12.01
            this_company_participant_no_field.send_keys(ee.getNumberOfOurCompany())         # --- UPDATE 2021.12.01
            ee.setNumberOfOurCompany(ee.getNumberOfOurCompany())                            # --- ゼロの場合イチにするセッター

            # OTHER_COMPANY_PARTICIPANT_NO = "10"                                           # --- DELETE 2021.12.01 前任者による定数エントリー
            other_company_participant_no_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[4]/div[3]/input"
            )
            # other_company_participant_no_field.send_keys(OTHER_COMPANY_PARTICIPANT_NO)    # --- DELETE 2021.12.01 
            other_company_participant_no_field.send_keys(ee.getNumberOfRecipients())        # --- UPDATE 2021.12.01
            ee.setNumberOfRecipients(ee.getNumberOfRecipients())                            # --- ゼロの場合イチにするセッター

            # TOTAL_PARTICIPANT_NO = "12"                                                   # --- DELETE 2021.12.01 前任者による定数エントリー
            total_company_participant_no_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[4]/div[4]/input"
            )
            # total_company_participant_no_field.send_keys(TOTAL_PARTICIPANT_NO)            # --- DELETE 2021.12.01 
            total_company_participant_no_field.send_keys(ee.getTotalNumberOfPeople())       # --- UPDATE 2021.12.01

            # PURPOSE = "目的"                                                              # --- DELETE 2021.12.01 前任者による定数エントリー
            purpose_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[5]/div/input"
            )
            # purpose_fld.send_keys(PURPOSE)                                                # --- DELETE 2021.12.01
            # purpose_fld.send_keys(ee.getAim())                                            # --- UPDATE 2021.12.01
            purpose_fld.send_keys(str(ee.getAim()))

            # ETD_DETAIL = "飲食"                                                           # --- DELETE 2021.12.01 前任者による定数エントリー
            etd_detail_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[6]/div/select"
            )
            # etd_detail_fld.send_keys(ETD_DETAIL)                                          # --- DELETE 2021.12.01
            etd_detail_fld.send_keys(ee.getMainContent())                                   # --- UPDATE 2021.12.01

            # 支払先反社番号 <--- 前任者による記述

            # SHOP_ANTI_NO = "支払先反社番号"                                               # --- DELETE 2021.12.01 前任者による定数エントリー
            shop_anti_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[7]/div[1]/input"
            )
            # shop_anti_fld.send_keys(SHOP_ANTI_NO)                                         # --- DELETE 2021.12.01
            shop_anti_fld.send_keys(ee.getPaymentDestPublicEnemyCd())                       # --- UPDATE 2021.12.01

            # 店舗名/ゴルフ場名 <--- 先任者による記述
            # SHOP = "店舗名/ゴルフ場名"                                                     # --- DELETE 2021.12.01 前任者による定数エントリー
            shop_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[7]/div[2]/input"
            )
            # shop_fld.send_keys(SHOP)                                                      # --- DELETE 2021.12.01
            shop_fld.send_keys(ee.getStoreName())                                           # --- UPDATE 2021.12.01

            # 支払先住所 * <--- 先任者による記述
            # PAY_ADDRESS = "支払先住所"                                                    # --- DELETE 2021.12.01 前任者による定数エントリー
            pay_address_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[8]/div/input"
            )
            # pay_address_fld.send_keys(PAY_ADDRESS)
            pay_address_fld.send_keys(ee.getPaymentAddress())

            # 支払予定額 <--- 先任者による記述
            # PLAN_PAY = "1000"                                                             # --- DELETE 2021.12.01 前任者による定数エントリー
            plan_pay_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[9]/div/input"
            )
            plan_pay_fld.send_keys(ee.getScheduledProvisionalPayment())

            # --- ADD NEW 2021.12.01 前任者は当該リストボックスへ対するエントリーをスキップ
            temp_pay_type_list = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[10]/div[1]/select"
            )
            temp_pay_type_list.send_keys(ee.getTemporaryPaymentType())

            # 仮払金額 <--- 先任者による記述
            # DOWN_PAY = "0"                                                                # --- DELETE 2021.12.01 前任者による定数エントリー
            down_pay_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[10]/div[2]/input"
            )
            # down_pay_fld.send_keys(DOWN_PAY)                                              # --- DELETE 2021.12.01
            down_pay_fld.send_keys(ee.getExpectedPaymentAmount())                           # --- UPDATE 2021.12.01

            # --- ADD NEW 2021.12.02 前任者は当該テキストボックスへ対するエントリーをスキップ
            desired_temp_pay_date_list = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[10]/div[3]/div/input"
            )
            desired_temp_pay_date_list.send_keys(ee.getDesiredTemporaryPaymentDate())

            # --- ADD NEW 2021.12.02 前任者は当該テキストボックスへ対するエントリーをスキップ
            scheduled_temp_pay_settlement_date_list = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[10]/div[4]/div/input"
            )
            scheduled_temp_pay_settlement_date_list.send_keys(ee.getScheduledTemporaryPaymentSettlementDate())

            # --- ADD NEW 2021.12.02 前任者は当該テキストボックスへ対するエントリーをスキップ
            remark_textbox_list = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[11]/div/textarea"
            )
            # gen = ee.getRemark()        # --- UPDATE 2021.12.22 Start
            # crlf = '\r\n'
            # remark_textbox_list.send_keys(gen.__next__() + crlf + gen.__next__() + crlf + gen.__next__())
            remark_textbox_list.send_keys(ee.getRemark())
            # gen = None                  # --- UPDATE 2021.12.22 End

            # Attach File function                                                          # --- 2021.12.01 Invalidate Start
            # attach_file_button = driver.find_element_by_xpath(                            # 添付ファイルタブクリックは別所にて実施
            #    "/html/body/div[1]/div[2]/div/form/div[5]/div/ul/li[3]/a"                  #
            # )                                                                             # 
            # my_sleep_click(attach_file_button)                                            # --- 2021.12.01 Invalidate End

            # ENT_ATTACH_FILE = "C:\Output\MFZ\Test01.pdf"                                  # --- DELETE 2021.12.01 前任者による定数エントリー

            # select_file_button = driver.find_element_by_xpath(                                # --- 2021.12.01 Invalidate Start
            #     "/html/body/div[1]/div[2]/div/form/div[5]/div/div/div["
            #     "3]/div/div/div/div[2]/div[2]/div/span/button"
            # )

            # my_sleep_click(select_file_button)

            # upload_file_button = driver.find_element_by_xpath(
            #     "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[2]/div[2]"
            # )
            # upload_file_button.click()
            # time.sleep(4)
            # pyautogui.write(ENT_ATTACH_FILE)                                                  # 前任者方式はフォルダーパス、ファイルパスに2バイト文字を含有する場合、
            # pyautogui.press("enter")                                                          # 失敗する
            # time.sleep(4)

            # add_green_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[3]/div/a"
            # )

            # my_sleep_click(add_green_button)

            # upload_file_button.send_keys(TRIP_ATTACH_FILE)                                # --- 2021.12.01 Invalidate End

            gen = ee.getVoucherFiles()                                                      # Generatorへ格納し、next()で取得する
            submit_voucher_files(gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__()
            )

            gen = None
            # my_sleep_click(confirm_button)
            ee = None

            # Final submission

            final_register_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[11]/button[1]"
            )

            click_submit_button(driver, logger, "Validation", final_register_button)

            # --- ADD NEW 2021.12.23 通知チェックボックスのレ点を解除
            notificationCheckBox = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/form/div[3]/label/input[1]"
            )
            my_sleep_click(notificationCheckBox)

            comment_field = driver.find_element_by_xpath(   # ！！ 備考テキストボックスである。下のコメントを参照せよ。
                "/html/body/div[1]/div/div/form/textarea"
            )
            comment_field.send_keys(MIGRATION_COMMENT)      # ！！ 前任者によるテストラン用コメント。実運用時は無効化せよ。

            final_submit_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/form/div[4]/button[1]"
            )

            # TODO uncomment below after testing
            # click_main(driver, logger, 'Submission', final_register_button)   # <--- 元IBMの意図が解せぬ
            my_sleep_click(final_submit_button)     # <--- UPDATE 2021.12.22

            # サインアウト処理
            sign_out_procedure()

            # ガーベジコレクター
            if i % 100 == 0:
                gc.collect()
            # SVシートの次行へ移動
            i += 1

def submit_voucher_files(vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9): 
    # 【添付ファイル】タブをクリック
    attach_file_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div[2]/div/form/div[5]/div/ul/li[3]/a"
    )
    my_sleep_click(attach_file_button)
    # 配列へ格納 --- ※証憑はPDFのみとは限らない
    va = [vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9]

    j = 0   # ゼロオリジン
    if va[j] != "N/A":  # 一つ目の配列がN/Aの場合は証憑がない
        while va[j] is not None and j < 10:
            if j == 0:  # 2回目以降は不要である為
                select_file_button = driver.find_element_by_xpath(
                # 添付ファイルタブに所在する添付ファイル-選択ボタン - 2021.11.15 コメント記入
                    "/html/body/div[1]/div[2]/div/form/div[5]/div/div/div["
                    "3]/div/div/div/div[2]/div[2]/div/span/button"
                )
                my_sleep_click(select_file_button)

            upload_file_button = driver.find_element_by_xpath(
                # 次画面にある「添付ファイルを選択してください」→選択ボタンをクリック
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[2]/div[2]"
            )
            upload_file_button.click()
            time.sleep(2)
            pyperclip.copy(va[j])                      # Clip Boardへコピーし、
            pg.hotkey('ctrl', 'v')              # Pasteする。
            pg.press("enter")
            time.sleep(2)
            add_green_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[3]/div/a"
            )
            my_sleep_click(add_green_button)
            time.sleep(2)
            # 右隣セルの値へ移動
            j += 1

        # 証憑ファイルサブミット後の確定ボタン
    if j != 0:  # --- ADD NEW 2021.12.24 --- 証憑が一つ以上あった場合のみ確定ボタンをクリック
        confirm_button = driver.find_element_by_xpath(
            "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"
        )
        my_sleep_click(confirm_button)

def signin_procedure(tenantId, empId, passWd):
    # テナントID = 100
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[1]/input"
    )
    first_input.send_keys(tenantId)
    # 職員コード
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[2]/input"
    )
    first_input.send_keys(empId)
    # パスワード（※2021.11.17時点、方針未決）
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[3]/input"
    )
    first_input.send_keys(passWd)
    # ログインボタンをクリック
    click_element(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[5]/button"
    )
    time.sleep(3)

def sign_out_procedure():    # 2021.12.15
    time.sleep(1)

    # ①申請完了後の「閉じる」ボタンをクリック
    close_button = driver.find_element_by_xpath(
        # "/html/body/div[1]/div/div/form/div/button[4]"    # <--- INVALIDATION 2021.12.22
        "/html/body/div[1]/div/div/form/div/button[2]"      # <--- UPDATE 2021.12.22
    )
    my_sleep_click(close_button)

    driver.implicitly_wait(3)                           # --- ADD NEW 2021.12.22
    driver.switch_to.window(driver.window_handles[0])   # --- ADD NEW 2021.12.22 要素ゼロのウィンドウへ制御を戻す！

    # driver.get("https://mflowz.daj.co.jp/MFZC/100/Logout/Logout")
    try:
        # ②画面右上「サインアウト」ボタン（非常出口アイコン）をクリック
        sign_out_buton = driver.find_element_by_xpath(
            "/html/body/div[2]/nav/div/div[2]/ul/li[6]/form/a"
        )
        my_sleep_click(sign_out_buton)
    except Exception as e:
        logger.info(e)
        # driver.quit()
    finally:
        pass

    time.sleep(1)
    # ③【ログイン】ボタンをラストにクリック
    sign_in_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    )
    my_sleep_click(sign_in_button)
    time.sleep(2)

def main():
    entry_procedure()
    logger.info("Robot completed")

main()
