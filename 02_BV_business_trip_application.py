# 【経費】出張申請書
import logging
import pyautogui as pg  # --- UPDATE 2021.12.23
pg.FAILSAFE = False     # --- UPDATE 2021.12.23

from chrome_driver_dl import get_latest_driver
from common import *

# --- ADD 2021.12.03 Start
import openpyxl
import sys
import pyperclip
import gc
from datetime import datetime
import concurrent.futures
import re

from selenium import webdriver
driver = webdriver.Chrome(
    executable_path="C:/Users/digiworker_biz_02/PycharmProjects/pythonProject/Chrome/chromedriver.exe"
)
# --- ADD 2021.12.03 End

# Gets or creates a logger
logger = logging.getLogger("02")
# set log level
logger.setLevel(logging.INFO)
# define file handler and set formatter
file_handler = logging.FileHandler("../Logs/02_business_trip_application.log")
formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")
file_handler.setFormatter(formatter)
# add file handler to logger
logger.addHandler(file_handler)

# driver = driver = webdriver.Chrome(get_latest_driver())
driver.get(MFZC_URL)
# driver.fullscreen_window()

driver.refresh()
logger.info("Robot Started")

class BusinessTripApplication:
    def __init__(self,
                wkUserId,
                wkPassWd,
                wkBusinessTripType,
                wkBusinessTripDepartment,
                wkBusinessTripArea,
                wkBusinessTripStartDate,
                wkBusinessTripEndDate,
                wkDepartureTime,
                wkGoStraight,
                wkReturnTime,
                wkBounce,                       # 直帰
                wkBusinessDay,
                wkBusinessDestinationA,
                wkBusinessDestinationB,
                wkBusinessContent,
                wkRemark,
                wkExCardNo,
                wkVoucher0,
                wkVoucher1,
                wkVoucher2,
                wkVoucher3,
                wkVoucher4,
                wkVoucher5,
                wkVoucher6,
                wkVoucher7,
                wkVoucher8,
                wkVoucher9,
                wkIid01,        # --- ADD NEW 2022.01.11 スコープ変更に伴う追加 x 16 （配列厳しい為、直打ち）
                wkDocno01,
                wkFormname01,
                wkCreated01,
                wkAname01,
                wkUid01,
                wkEmpcd01,
                wkEname01,
                wkDeptid01,
                wkDfullname01,
                wkPost01,
                wkTs01,
                wkIid02,
                wkDocno02,
                wkFormname02,
                wkCreated02,
                wkAname02,
                wkUid02,
                wkEmpcd02,
                wkEname02,
                wkDeptid02,
                wkDfullname02,
                wkPost02,
                wkTs02,
                wkIid03,
                wkDocno03,
                wkFormname03,
                wkCreated03,
                wkAname03,
                wkUid03,
                wkEmpcd03,
                wkEname03,
                wkDeptid03,
                wkDfullname03,
                wkPost03,
                wkTs03,
                wkIid04,
                wkDocno04,
                wkFormname04,
                wkCreated04,
                wkAname04,
                wkUid04,
                wkEmpcd04,
                wkEname04,
                wkDeptid04,
                wkDfullname04,
                wkPost04,
                wkTs04,
                wkIid05,
                wkDocno05,
                wkFormname05,
                wkCreated05,
                wkAname05,
                wkUid05,
                wkEmpcd05,
                wkEname05,
                wkDeptid05,
                wkDfullname05,
                wkPost05,
                wkTs05,
                wkIid06,
                wkDocno06,
                wkFormname06,
                wkCreated06,
                wkAname06,
                wkUid06,
                wkEmpcd06,
                wkEname06,
                wkDeptid06,
                wkDfullname06,
                wkPost06,
                wkTs06,
                wkIid07,
                wkDocno07,
                wkFormname07,
                wkCreated07,
                wkAname07,
                wkUid07,
                wkEmpcd07,
                wkEname07,
                wkDeptid07,
                wkDfullname07,
                wkPost07,
                wkTs07,
                wkIid08,
                wkDocno08,
                wkFormname08,
                wkCreated08,
                wkAname08,
                wkUid08,
                wkEmpcd08,
                wkEname08,
                wkDeptid08,
                wkDfullname08,
                wkPost08,
                wkTs08,
                wkIid09,
                wkDocno09,
                wkFormname09,
                wkCreated09,
                wkAname09,
                wkUid09,
                wkEmpcd09,
                wkEname09,
                wkDeptid09,
                wkDfullname09,
                wkPost09,
                wkTs09,
                wkIid10,
                wkDocno10,
                wkFormname10,
                wkCreated10,
                wkAname10,
                wkUid10,
                wkEmpcd10,
                wkEname10,
                wkDeptid10,
                wkDfullname10,
                wkPost10,
                wkTs10,
                wkIid11,
                wkDocno11,
                wkFormname11,
                wkCreated11,
                wkAname11,
                wkUid11,
                wkEmpcd11,
                wkEname11,
                wkDeptid11,
                wkDfullname11,
                wkPost11,
                wkTs11,
                wkIid12,
                wkDocno12,
                wkFormname12,
                wkCreated12,
                wkAname12,
                wkUid12,
                wkEmpcd12,
                wkEname12,
                wkDeptid12,
                wkDfullname12,
                wkPost12,
                wkTs12,
                wkIid13,
                wkDocno13,
                wkFormname13,
                wkCreated13,
                wkAname13,
                wkUid13,
                wkEmpcd13,
                wkEname13,
                wkDeptid13,
                wkDfullname13,
                wkPost13,
                wkTs13,
                wkIid14,
                wkDocno14,
                wkFormname14,
                wkCreated14,
                wkAname14,
                wkUid14,
                wkEmpcd14,
                wkEname14,
                wkDeptid14,
                wkDfullname14,
                wkPost14,
                wkTs14,
                wkIid15,
                wkDocno15,
                wkFormname15,
                wkCreated15,
                wkAname15,
                wkUid15,
                wkEmpcd15,
                wkEname15,
                wkDeptid15,
                wkDfullname15,
                wkPost15,
                wkTs15,
                wkIid16,
                wkDocno16,
                wkFormname16,
                wkCreated16,
                wkAname16,
                wkUid16,
                wkEmpcd16,
                wkEname16,
                wkDeptid16,
                wkDfullname16,
                wkPost16,
                wkTs16
    ):
        self.userID = wkUserId
        self.passWd = wkPassWd
        self.businessTripType = wkBusinessTripType
        self.businessTripDepartment = wkBusinessTripDepartment
        self.businessTripArea = wkBusinessTripArea
        self.businessTripStartDate = wkBusinessTripStartDate
        self.businessTripEndDate = wkBusinessTripEndDate
        self.departureTime = wkDepartureTime
        self.goStraight = wkGoStraight
        self.returnTime = wkReturnTime
        self.bounce = wkBounce
        self.businessDay = wkBusinessDay
        self.businessDestinationA = wkBusinessDestinationA
        self.businessDestinationB = wkBusinessDestinationB
        self.businessContent = wkBusinessContent
        self.remark = wkRemark
        self.exCardNo = wkExCardNo
        self.voucher0 = wkVoucher0
        self.voucher1 = wkVoucher1
        self.voucher2 = wkVoucher2
        self.voucher3 = wkVoucher3
        self.voucher4 = wkVoucher4
        self.voucher5 = wkVoucher5
        self.voucher6 = wkVoucher6
        self.voucher7 = wkVoucher7
        self.voucher8 = wkVoucher8
        self.voucher9 = wkVoucher9
        self.iid01 = wkIid01
        self.docno01 = wkDocno01
        self.formname01 = wkFormname01
        self.createdd01 = wkCreated01
        self.aname01 = wkAname01
        self.uid01 = wkUid01
        self.empcd01 = wkEmpcd01
        self.ename01 = wkEname01
        self.deptid01 = wkDeptid01
        self.dfullname01 = wkDfullname01
        self.post01 = wkPost01
        self.ts01 = wkTs01
        self.iid02 = wkIid02
        self.docno02 = wkDocno02
        self.forname02 = wkFormname02
        self.created02 = wkCreated02
        self.aname02 = wkAname02
        self.uid02 = wkUid02
        self.empcd02 = wkEmpcd02
        self.ename02 = wkEname02
        self.deptid02 = wkDeptid02
        self.dfullname02 = wkDfullname02
        self.post02 = wkPost02
        self.ts02 = wkTs02
        self.iid03 = wkIid03
        self.docno03 = wkDocno03
        self.fornmame03 = wkFormname03
        self.created03 = wkCreated03
        self.aname03 = wkAname03
        self.uid03 = wkUid03
        self.empcd03 = wkEmpcd03
        self.ename03 = wkEname03
        self.deptid03 = wkDeptid03
        self.dfullname03 = wkDfullname03
        self.post03 = wkPost03
        self.ts03 = wkTs03
        self.iid04 = wkIid04
        self.docno04 = wkDocno04
        self.formname04 = wkFormname04
        self.created04 = wkCreated04
        self.aname04 = wkAname04
        self.uid04 = wkUid04
        self.empcd04 = wkEmpcd04
        self.ename04 = wkEname04
        self.deptid04 = wkDeptid04
        self.dfullname04 = wkDfullname04
        self.post04 = wkPost04
        self.ts04 = wkTs04
        self.iid05 = wkIid05
        self.docno05 = wkDocno05
        self.formname05 = wkFormname05
        self.created05 = wkCreated05
        self.aname05 = wkAname05
        self.uid05 = wkUid05
        self.empcd05 = wkEmpcd05
        self.ename05 = wkEname05
        self.deptid05 = wkDeptid05
        self.dfullname05 = wkDfullname05
        self.post05 = wkPost05
        self.ts05 = wkTs05
        self.iid06 = wkIid06
        self.docno06 = wkDocno06
        self.formname06 = wkFormname06
        self.created06 = wkCreated06
        self.aname06 = wkAname06
        self.uid06 = wkUid06
        self.empcd06 = wkEmpcd06
        self.ename06 = wkEname06
        self.deptid06 = wkDeptid06
        self.dfullname06 = wkDfullname06
        self.post06 = wkPost06
        self.ts06 = wkTs06
        self.iid07 = wkIid07
        self.docno07 = wkDocno07
        self.forname07 = wkFormname07
        self.created07 = wkCreated07
        self.aname07 = wkAname07
        self.uid07 = wkUid07
        self.empcd07 = wkEmpcd07
        self.ename07 = wkEname07
        self.deptid07 = wkDeptid07
        self.dfullname07 = wkDfullname07
        self.post07 = wkPost07
        self.ts07 = wkTs07
        self.iid08 = wkIid08
        self.docno08 = wkDocno08
        self.forname08 = wkFormname08
        self.created08 = wkCreated08
        self.aname08 = wkAname08
        self.uid08 = wkUid08
        self.empcd08 = wkEmpcd08
        self.ename08 = wkEname08
        self.deptid08 = wkDeptid08
        self.dfullname08 = wkDfullname08
        self.post08 = wkPost08
        self.ts08 = wkTs08
        self.iid09 = wkIid09
        self.docno09 = wkDocno09
        self.formname09 = wkFormname09
        self.created09 = wkCreated09
        self.aname09 = wkAname09
        self.uid09 = wkUid09
        self.empcd09 = wkEmpcd09
        self.ename09 = wkEname09
        self.deptid09 = wkDeptid09
        self.dfullname09 = wkDfullname09
        self.post09 = wkPost09
        self.ts09 = wkTs09
        self.iid10 = wkIid10
        self.docno10 = wkDocno10
        self.formname10 = wkFormname10
        self.created10 = wkCreated10
        self.aname10 = wkAname10
        self.uid10 = wkUid10
        self.empcd10 = wkEmpcd10
        self.ename10 = wkEname10
        self.deptid10 = wkDeptid10
        self.dfullname10 = wkDfullname10
        self.post10 = wkPost10
        self.ts = wkTs10
        self.iid11 = wkIid11
        self.docno11 = wkDocno11
        self.fromname11 = wkFormname11
        self.created11 = wkCreated11
        self.aname11 = wkAname11
        self.uid11 = wkUid11
        self.empcd11 = wkEmpcd11
        self.ename11 = wkEname11
        self.deptid11 = wkDeptid11
        self.dfullname11 = wkDfullname11
        self.post11 = wkPost11
        self.ts11 = wkTs11
        self.iid12 = wkIid12
        self.docno12 = wkDocno12
        self.fromname12 = wkFormname12
        self.created12 = wkCreated12
        self.aname12 = wkAname12
        self.uid12 = wkUid12
        self.empcd12 = wkEmpcd12
        self.ename12 = wkEname12
        self.deptid12 = wkDeptid12
        self.dfullname12 = wkDfullname12
        self.post12 = wkPost12
        self.ts12 = wkTs12
        self.iid13 = wkIid13
        self.docno13 = wkDocno13
        self.formname13 = wkFormname13
        self.created13 = wkCreated13
        self.aname13 = wkAname13
        self.uid13 = wkUid13
        self.empcd13 = wkEmpcd13
        self.ename13 = wkEname13
        self.deptid13 = wkDeptid13
        self.dfullname13 = wkDfullname13
        self.post13 = wkPost13
        self.ts13 = wkTs13
        self.iid14 = wkIid14
        self.docno14 = wkDocno14
        self.formname14 = wkFormname14
        self.created14 = wkCreated14
        self.aname14 = wkAname14
        self.uid14 = wkUid14
        self.empcd14 = wkEmpcd14
        self.ename14 = wkEname14
        self.deptid14 = wkDeptid14
        self.dfullname14 = wkDfullname14
        self.post14 = wkPost14
        self.ts14 = wkTs14
        self.iid15 = wkIid15
        self.docno15 = wkDocno15
        self.formname15 = wkFormname15
        self.created15 = wkCreated15
        self.aname15 = wkAname15
        self.uid15 = wkUid15
        self.empcd15 = wkEmpcd15
        self.ename15 = wkEname15
        self.deptid15 = wkDeptid15
        self.dfullname15 = wkDfullname15
        self.post15 = wkPost15
        self.ts15 = wkTs15
        self.iid16 = wkIid16
        self.docno16 = wkDocno16
        self.fromname16 = wkFormname16
        self.created16 = wkCreated16
        self.aname16 = wkAname16
        self.uid16 = wkUid16
        self.empcd16 = wkEmpcd16
        self.ename16 = wkEname16
        self.deptid16 = wkDeptid16
        self.dfullname16 = wkDfullname16
        self.post16 = wkPost16
        self.ts16 = wkTs16

    def getSignInInfo(self):
        yield 100
        yield self.userID
        yield self.passWd

    def getBusinessTripType(self):
        return self.businessTripType

    # --- UPDATE 2021.12.06 --- 東京本社と思料される所属以外はそのまま返し、東京本社内の部課は東京本社文字列を返す
    def getBusinessTripDepartment(self):
        if self.businessTripDepartment == '北海道営業所' \
            or self.businessTripDepartment == '東北営業所' \
            or self.businessTripDepartment == '東京本社' \
            or self.businessTripDepartment == '中部営業所' \
            or self.businessTripDepartment == '関西営業所' \
            or self.businessTripDepartment == '中四国営業所' \
            or self.businessTripDepartment == '九州営業所':
                return self.businessTripDepartment
        else:
            return '東京本社'

    # --- レコード数が少ないので、対象文字列の左側2文字または3文字で識別させる
    def getBusinessTripArea(self):
        mojiretsu = self.businessTripArea
        if mojiretsu[0:2] == '宮城' or mojiretsu[0:2] == '岩手':
            return '東北'
        elif mojiretsu[0:2] == '栃木' or mojiretsu[0:2] == '静岡' or mojiretsu[0:2] == '越後' or mojiretsu[0:1] == '御茶' or mojiretsu[0:2] == '東京':
            return '関東'
        elif mojiretsu[0:2] == '愛知' or mojiretsu[0:2] == '三重' or mojiretsu[0:2] == '浜松':
            return '中部'
        elif mojiretsu[0:2] == '大阪' or mojiretsu[0:2] == '京都':
            return '関西'
        elif mojiretsu[0:2] == '愛媛' or mojiretsu[0:2] == '鳥取' or mojiretsu[0:2] == '山口' or mojiretsu[0:2] == '広島':
            return '中四国'
        elif mojiretsu[0:2] == '熊本' or mojiretsu[0:3] == '鹿児島' or mojiretsu[0:2] == '宮崎' or mojiretsu[0:3] == '久留米' or mojiretsu[0:2] == '長崎' or mojiretsu[0:2] == '福岡' or mojiretsu[0:2] == '大分':
            return '九州'
        elif mojiretsu[0:2] == '沖縄':
            return '沖縄'
        else:
            return 'アジア・オセアニア'

    def getBusinessTripStartDate(self):
        if self.businessTripStartDate is None or self.businessTripStartDate == 0:
            today = datetime.now()
            return today.strftime('%Y-%m-%d')           # 本日日付を戻す
        else:
            try:
                wkDateTime = datetime.strptime((self.businessTripStartDate), "%Y/%m/%d %H:%M")
                # Excelセルから時分秒を取得して仕舞う為、当該時分秒を除外する処理を施す
                return datetime.strptime(str(self.businessTripStartDate).replace('-', ''), '%Y%m%d %H:%M:%S').strftime('%Y-%m-%d')
            except Exception as e:
                logger.info(e)  # 実行時例外を記録のみする
                today = datetime.now()
                return today.strftime('%Y-%m-%d')  # 本日日付を戻す
            finally:
                pass

    def getBusinessTripEndDate(self):
        if self.businessTripEndDate is None or self.businessTripEndDate == 0:
            today = datetime.now()
            return today.strftime('%Y-%m-%d')           # 本日日付を戻す
        else:
            try:
                wkDateTime = datetime.strptime((self.businessTripEndDate), "%Y/%m/%d %H:%M")
                # Excelセルから時分秒を取得して仕舞う為、当該時分秒を除外する処理を施す
                return datetime.strptime(str(self.businessTripEndDate).replace('-', ''), '%Y%m%d %H:%M:%S').strftime('%Y-%m-%d')
            except Exception as e:
                logger.info(e)  # 実行時例外を記録のみする
                today = datetime.now()
                return today.strftime('%Y-%m-%d')  # 本日日付を戻す
            finally:
                pass

    def getBusinessDay(self):
        if self.businessDay is None or self.businessDay == 0:
            today = datetime.now()
            return today.strftime('%Y-%m-%d')           # 本日日付を戻す
        else:
            # Excelセルから時分秒を取得して仕舞う為、当該時分秒を除外する処理を施す
            # return datetime.strptime(str(self.businessDay).replace('-',''), '%Y%m%d %H:%M:%S').strftime('%Y-%m-%d')   # --- INVALID 2021.12.23
            return str(self.businessDay) [0:49]   # --- REPRACE 2021.12.23。varchar2型と思料され、50文字上限

    def getDepartureTime(self):
        # return datetime.time(self.departureTime)  # 使用不可
        if str(self.departureTime) in ':' == True:  # --- ADD NEW 2021.12.23 --- 時刻でない値が入っていた場合の対応 --- Start
            return str(self.departureTime)[0:5]
        else:
            return '09:00'                           # --- ADD NEW 2021.12.23 --- 時刻でない値が入っていた場合の対応 --- End

    def getGoStraight(self):
        if self.goStraight == 'する':                 # 旧MFにおける当該区分は①「出社する」、②「出社しない」につき、
            # return True                            # 新MFZでは①「直行しない」・②「直行する」へ読替える
            return False                             # --- CHANGE 2021.12.16
        elif self.goStraight == 'しない' or self.goStraight is None:
            # return False                          # --- CHANGE 2021.12.16
            return True                             # --- CHANGE 2021.12.16

    def getReturnTime(self):
        if str(self.returnTime) in ':' == True:     # --- ADD NEW 2021.12.23 --- 時刻でない値が入っていた場合の対応 --- Start
            return str(self.returnTime)[0:5]
        else:
            return '18:00'                          # --- ADD NEW 2021.12.23 --- 時刻でない値が入っていた場合の対応 --- End

    def getBounce(self):
        if self.bounce == 'する':                     # 旧MFにおける当該区分は①「帰社する」、②「帰社しない」につき、
            # return True                            # 新MFZでは①「直帰しない」・②「直帰する」へ読替える
            return False
        elif self.bounce == 'しない' or self.bounce is None:
            # return False                          # --- CHANGE 2021.12.16
            return True                             # --- CHANGE 2021.12.16

    def getBusinessDestination(self):
        yield str(self.businessDestinationA)
        yield str(self.businessDestinationB)

    def getBusinessContact(self):
        return self.businessContent
    
    def getVoucherFiles(self):
        yield self.voucher0
        yield self.voucher1
        yield self.voucher2
        yield self.voucher3
        yield self.voucher4
        yield self.voucher5
        yield self.voucher6
        yield self.voucher7
        yield self.voucher8
        yield self.voucher9

    def getApprovalInfo(self):      # --- 2022.01.11 ADD NEW 決裁情報を返す
        approvalInfo = ""
        if self.iid01 is not None and self.iid01 != "":
            approvalInfo += f'決裁:{self.aname01} 従業員ID:{self.uid01} 従業員番号:{self.empcd01} 従業員指名:{self.ename01} ;'
        if self.iid02 is not None and self.iid02 != "":
            approvalInfo += f'決裁:{self.aname02} 従業員ID:{self.uid02} 従業員番号:{self.empcd02} 従業員指名:{self.ename02} ;'
        if self.iid03 is not None and self.iid03 != "":
            approvalInfo += f'決裁:{self.aname03} 従業員ID:{self.uid03} 従業員番号:{self.empcd03} 従業員指名:{self.ename03} ;'
        if self.iid04 is not None and self.iid04 != "":
            approvalInfo += f'決裁:{self.aname04} 従業員ID:{self.uid04} 従業員番号:{self.empcd04} 従業員指名:{self.ename04} ;'
        if self.iid05 is not None and self.iid05 != "":
            approvalInfo += f'決裁:{self.aname05} 従業員ID:{self.uid05} 従業員番号:{self.empcd05} 従業員指名:{self.ename05} ;'
        if self.iid06 is not None and self.iid06 != "":
            approvalInfo += f'決裁:{self.aname06} 従業員ID:{self.uid06} 従業員番号:{self.empcd06} 従業員指名:{self.ename06} ;'
        if self.iid07 is not None and self.iid07 != "":
            approvalInfo += f'決裁:{self.aname07} 従業員ID:{self.uid07} 従業員番号:{self.empcd07} 従業員指名:{self.ename07} ;'
        if self.iid08 is not None and self.iid08 != "":
            approvalInfo += f'決裁:{self.aname08} 従業員ID:{self.uid08} 従業員番号:{self.empcd08} 従業員指名:{self.ename08} ;'
        if self.iid09 is not None and self.iid09 != "":
            approvalInfo += f'決裁:{self.aname09} 従業員ID:{self.uid09} 従業員番号:{self.empcd09} 従業員指名:{self.ename09} ;'
        if self.iid10 is not None and self.iid10 != "":
            approvalInfo += f'決裁:{self.aname10} 従業員ID:{self.uid10} 従業員番号:{self.empcd10} 従業員指名:{self.ename10} ;'
        if self.iid11 is not None and self.iid11 != "":
            approvalInfo += f'決裁:{self.aname11} 従業員ID:{self.uid11} 従業員番号:{self.empcd11} 従業員指名:{self.ename11} ;'
        if self.iid12 is not None and self.iid12 != "":
            approvalInfo += f'決裁:{self.aname12} 従業員ID:{self.uid12} 従業員番号:{self.empcd12} 従業員指名:{self.ename12} ;'
        if self.iid13 is not None and self.iid13 != "":
            approvalInfo += f'決裁:{self.aname13} 従業員ID:{self.uid13} 従業員番号:{self.empcd13} 従業員指名:{self.ename13} ;'
        if self.iid14 is not None and self.iid14 != "":
            approvalInfo += f'決裁:{self.aname14} 従業員ID:{self.uid14} 従業員番号:{self.empcd14} 従業員指名:{self.ename14} ;'
        if self.iid15 is not None and self.iid15 != "":
            approvalInfo += f'決裁:{self.aname15} 従業員ID:{self.uid15} 従業員番号:{self.empcd15} 従業員指名:{self.ename15} ;'
        if self.iid16 is not None and self.iid16 != "":
            approvalInfo += f'決裁:{self.aname16} 従業員ID:{self.uid16} 従業員番号:{self.empcd16} 従業員指名:{self.ename16} ;'
        return approvalInfo     # まとめて返す

def signin_procedure(tenantId, empId, passWd):
    logger.info(            # --- ADD NEW 共通部品 20212.01.04
        "テナントID: " + tenantId + " 職員番号: " + empId + " 共通パスワード: " + passWd
    )
    # テナントID = 100
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[1]/input"
    )
    first_input.send_keys(tenantId)
    # 職員番号
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[2]/input"
    )
    first_input.send_keys(empId)
    # パスワード（※2021.11.17時点、方針未決）
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[3]/input"
    )
    first_input.send_keys(passWd)
    # ログインボタンをクリック
    logger.info(            # --- ADD NEW 共通部品 20212.01.04
        "ログインボタンクリック"
    )
    click_element(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[5]/button"
    )
    time.sleep(2)       # --- ADD 2022.01.04

def click_element(web_el):
    my_sleep_click(driver.find_element_by_xpath(web_el))

def click_main(web_el):
    try:
        web_el.click()
        logger.info("Success")
    except NoSuchElementException:
        logger.error("Error")
    time.sleep(3)

def reloadBrowser():
    try:
        time.sleep(2)
        driver.refresh()
        time.sleep(2)
    except Exceptio as e:
        logger.info("Chromeブラウザーをリロードできず。ドライバーを確認せよ")
        # print("Chromeブラウザーをリロードできず。ドライバーを確認せよ")
        # driver.quit()
    finally:
        pass

def entry_procedure():

    reloadBrowser()
    # click_element(
    #     "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    # )
    logger.info(  # --- ADD NEW 共通部品 20212.01.04
        "BV.xlsxファイルオープン開始"
    )
    # 交際費申請書の伝票タイプはBV
    book = openpyxl.load_workbook("C:\XLSX\XM_SLIP\BV.xlsx")    # Book名がBV
    sheet = book['BV']                                          # 出張申請書
    logger.info(  # --- ADD NEW 共通部品 20212.01.04
        "BV.xlsxファイルオープン完了"
    )

    with concurrent.futures.ProcessPoolExecutor() as executor:
        i = 2   # 1行目は見出し行。2行目から開始。A列セルをNull判定しEOFを見極める。
        while sheet.cell(row=i, column=1).value is not None:
            bt = BusinessTripApplication(sheet.cell(row=i, column=9).value,     # CREATOR
                                '0Nu4M0%4N0',                                   # 共通パスワード（※未定）
                                sheet.cell(row=i, column=68).value,             # 出張種別【SS02】
                                sheet.cell(row=i, column=11).value,             # 出張者の所属【DEPTNAME】
                                sheet.cell(row=i, column=81).value,             # 出張先地域【SS15】（※納品先要確認）
                                sheet.cell(row=i, column=69).value,             # 出張期間（自）【SS03】
                                sheet.cell(row=i, column=70).value,             # 出張期間（至）【SS04】
                                sheet.cell(row=i, column=72).value,             # 出発時刻【SS06】
                                sheet.cell(row=i, column=71).value,             # 直行チェックボックス【SS05】（※未確定、納品先要確認）
                                sheet.cell(row=i, column=74).value,             # 帰着時刻【SS08】
                                sheet.cell(row=i, column=73).value,             # 直帰チェックボックス【SS07】（※未確定、納品先要確認）
                                sheet.cell(row=i, column=84).value,             # 用務日【SS18】
                                sheet.cell(row=i, column=81).value,             # 用務先A【★SS15】&【SS16】（※未確定、納品先要確認）
                                sheet.cell(row=i, column=82).value,             # 用務先B【SS15】&【★SS16】（※未確定、納品先要確認）
                                sheet.cell(row=i, column=83).value,             # 用務内容【SS17】
                                sheet.cell(row=i, column=86).value,             # 備考【SS20】
                                sheet.cell(row=i, column=256).value,            # EXカード番号（※未確定、納品先要確認）
                                sheet.cell(row=i, column=109).value,            # 証憑ファイル1
                                sheet.cell(row=i, column=110).value,            # 証憑ファイル2
                                sheet.cell(row=i, column=111).value,            # 証憑ファイル3
                                sheet.cell(row=i, column=112).value,            # 証憑ファイル4
                                sheet.cell(row=i, column=113).value,            # 証憑ファイル5
                                sheet.cell(row=i, column=114).value,            # 証憑ファイル6
                                sheet.cell(row=i, column=115).value,            # 証憑ファイル7
                                sheet.cell(row=i, column=116).value,            # 証憑ファイル8
                                sheet.cell(row=i, column=117).value,            # 証憑ファイル9
                                sheet.cell(row=i, column=118).value,             # 証憑ファイル10
                                sheet.cell(row=i, column=131).value,  # IID --- (1)
                                sheet.cell(row=i, column=132).value,  # DOCNO
                                sheet.cell(row=i, column=133).value,  # FORMNAME
                                sheet.cell(row=i, column=134).value,  # CREATED
                                sheet.cell(row=i, column=135).value,  # ANAME
                                sheet.cell(row=i, column=136).value,  # UID
                                sheet.cell(row=i, column=137).value,  # EMPCD
                                sheet.cell(row=i, column=138).value,  # ENAME
                                sheet.cell(row=i, column=139).value,  # DEPTID
                                sheet.cell(row=i, column=140).value,  # DFULLNAME
                                sheet.cell(row=i, column=141).value,  # POST
                                sheet.cell(row=i, column=142).value,  # TS
                                sheet.cell(row=i, column=144).value,  # IID --- (2)
                                sheet.cell(row=i, column=145).value,  # DOCNO
                                sheet.cell(row=i, column=146).value,  # FORMNAME
                                sheet.cell(row=i, column=147).value,  # CREATED
                                sheet.cell(row=i, column=148).value,  # ANAME
                                sheet.cell(row=i, column=149).value,  # UID
                                sheet.cell(row=i, column=150).value,  # EMPCD
                                sheet.cell(row=i, column=151).value,  # ENAME
                                sheet.cell(row=i, column=152).value,  # DEPTID
                                sheet.cell(row=i, column=153).value,  # DFULLNAME
                                sheet.cell(row=i, column=154).value,  # POST
                                sheet.cell(row=i, column=155).value,  # TS
                                sheet.cell(row=i, column=157).value,  # IID --- (3)
                                sheet.cell(row=i, column=158).value,  # DOCNO
                                sheet.cell(row=i, column=159).value,  # FORMNAME
                                sheet.cell(row=i, column=160).value,  # CREATED
                                sheet.cell(row=i, column=161).value,  # ANAME
                                sheet.cell(row=i, column=162).value,  # UID
                                sheet.cell(row=i, column=163).value,  # EMPCD
                                sheet.cell(row=i, column=164).value,  # ENAME
                                sheet.cell(row=i, column=165).value,  # DEPTID
                                sheet.cell(row=i, column=166).value,  # DFULLNAME
                                sheet.cell(row=i, column=167).value,  # POST
                                sheet.cell(row=i, column=168).value,  # TS
                                sheet.cell(row=i, column=170).value,  # IID --- (4)
                                sheet.cell(row=i, column=171).value,  # DOCNO
                                sheet.cell(row=i, column=172).value,  # FORMNAME
                                sheet.cell(row=i, column=173).value,  # CREATED
                                sheet.cell(row=i, column=174).value,  # ANAME
                                sheet.cell(row=i, column=175).value,  # UID
                                sheet.cell(row=i, column=176).value,  # EMPCD
                                sheet.cell(row=i, column=177).value,  # ENAME
                                sheet.cell(row=i, column=178).value,  # DEPTID
                                sheet.cell(row=i, column=179).value,  # DFULLNAME
                                sheet.cell(row=i, column=180).value,  # POST
                                sheet.cell(row=i, column=181).value,  # TS
                                sheet.cell(row=i, column=183).value,  # IID --- (5)
                                sheet.cell(row=i, column=184).value,  # DOCNO
                                sheet.cell(row=i, column=185).value,  # FORMNAME
                                sheet.cell(row=i, column=186).value,  # CREATED
                                sheet.cell(row=i, column=187).value,  # ANAME
                                sheet.cell(row=i, column=188).value,  # UID
                                sheet.cell(row=i, column=189).value,  # EMPCD
                                sheet.cell(row=i, column=190).value,  # ENAME
                                sheet.cell(row=i, column=191).value,  # DEPTID
                                sheet.cell(row=i, column=192).value,  # DFULLNAME
                                sheet.cell(row=i, column=193).value,  # POST
                                sheet.cell(row=i, column=194).value,  # TS
                                sheet.cell(row=i, column=196).value,  # IID --- (6)
                                sheet.cell(row=i, column=197).value,  # DOCNO
                                sheet.cell(row=i, column=198).value,  # FORMNAME
                                sheet.cell(row=i, column=199).value,  # CREATED
                                sheet.cell(row=i, column=200).value,  # ANAME
                                sheet.cell(row=i, column=201).value,  # UID
                                sheet.cell(row=i, column=202).value,  # EMPCD
                                sheet.cell(row=i, column=203).value,  # ENAME
                                sheet.cell(row=i, column=204).value,  # DEPTID
                                sheet.cell(row=i, column=205).value,  # DFULLNAME
                                sheet.cell(row=i, column=206).value,  # POST
                                sheet.cell(row=i, column=207).value,  # TS
                                sheet.cell(row=i, column=209).value,  # IID --- (7)
                                sheet.cell(row=i, column=210).value,  # DOCNO
                                sheet.cell(row=i, column=211).value,  # FORMNAME
                                sheet.cell(row=i, column=212).value,  # CREATED
                                sheet.cell(row=i, column=213).value,  # ANAME
                                sheet.cell(row=i, column=214).value,  # UID
                                sheet.cell(row=i, column=215).value,  # EMPCD
                                sheet.cell(row=i, column=216).value,  # ENAME
                                sheet.cell(row=i, column=217).value,  # DEPTID
                                sheet.cell(row=i, column=218).value,  # DFULLNAME
                                sheet.cell(row=i, column=219).value,  # POST
                                sheet.cell(row=i, column=220).value,  # TS
                                sheet.cell(row=i, column=222).value,  # IID --- (8)
                                sheet.cell(row=i, column=223).value,  # DOCNO
                                sheet.cell(row=i, column=224).value,  # FORMNAME
                                sheet.cell(row=i, column=225).value,  # CREATED
                                sheet.cell(row=i, column=226).value,  # ANAME
                                sheet.cell(row=i, column=227).value,  # UID
                                sheet.cell(row=i, column=228).value,  # EMPCD
                                sheet.cell(row=i, column=229).value,  # ENAME
                                sheet.cell(row=i, column=230).value,  # DEPTID
                                sheet.cell(row=i, column=231).value,  # DFULLNAME
                                sheet.cell(row=i, column=232).value,  # POST
                                sheet.cell(row=i, column=233).value,  # TS
                                sheet.cell(row=i, column=235).value,  # IID --- (9)
                                sheet.cell(row=i, column=236).value,  # DOCNO
                                sheet.cell(row=i, column=237).value,  # FORMNAME
                                sheet.cell(row=i, column=238).value,  # CREATED
                                sheet.cell(row=i, column=239).value,  # ANAME
                                sheet.cell(row=i, column=240).value,  # UID
                                sheet.cell(row=i, column=241).value,  # EMPCD
                                sheet.cell(row=i, column=242).value,  # ENAME
                                sheet.cell(row=i, column=243).value,  # DEPTID
                                sheet.cell(row=i, column=244).value,  # DFULLNAME
                                sheet.cell(row=i, column=245).value,  # POST
                                sheet.cell(row=i, column=246).value,  # TS
                                sheet.cell(row=i, column=248).value,  # IID --- (10)
                                sheet.cell(row=i, column=249).value,  # DOCNO
                                sheet.cell(row=i, column=250).value,  # FORMNAME
                                sheet.cell(row=i, column=251).value,  # CREATED
                                sheet.cell(row=i, column=252).value,  # ANAME
                                sheet.cell(row=i, column=253).value,  # UID
                                sheet.cell(row=i, column=254).value,  # EMPCD
                                sheet.cell(row=i, column=255).value,  # ENAME
                                sheet.cell(row=i, column=256).value,  # DEPTID
                                sheet.cell(row=i, column=257).value,  # DFULLNAME
                                sheet.cell(row=i, column=258).value,  # POST
                                sheet.cell(row=i, column=259).value,  # TS
                                sheet.cell(row=i, column=261).value,  # IID --- (11)
                                sheet.cell(row=i, column=262).value,  # DOCNO
                                sheet.cell(row=i, column=263).value,  # FORMNAME
                                sheet.cell(row=i, column=264).value,  # CREATED
                                sheet.cell(row=i, column=265).value,  # ANAME
                                sheet.cell(row=i, column=266).value,  # UID
                                sheet.cell(row=i, column=267).value,  # EMPCD
                                sheet.cell(row=i, column=268).value,  # ENAME
                                sheet.cell(row=i, column=269).value,  # DEPTID
                                sheet.cell(row=i, column=270).value,  # DFULLNAME
                                sheet.cell(row=i, column=271).value,  # POST
                                sheet.cell(row=i, column=272).value,  # TS
                                sheet.cell(row=i, column=274).value,  # IID --- (12)
                                sheet.cell(row=i, column=275).value,  # DOCNO
                                sheet.cell(row=i, column=276).value,  # FORMNAME
                                sheet.cell(row=i, column=277).value,  # CREATED
                                sheet.cell(row=i, column=278).value,  # ANAME
                                sheet.cell(row=i, column=279).value,  # UID
                                sheet.cell(row=i, column=280).value,  # EMPCD
                                sheet.cell(row=i, column=281).value,  # ENAME
                                sheet.cell(row=i, column=282).value,  # DEPTID
                                sheet.cell(row=i, column=283).value,  # DFULLNAME
                                sheet.cell(row=i, column=284).value,  # POST
                                sheet.cell(row=i, column=285).value,  # TS
                                sheet.cell(row=i, column=287).value,  # IID --- (13)
                                sheet.cell(row=i, column=288).value,  # DOCNO
                                sheet.cell(row=i, column=289).value,  # FORMNAME
                                sheet.cell(row=i, column=290).value,  # CREATED
                                sheet.cell(row=i, column=291).value,  # ANAME
                                sheet.cell(row=i, column=292).value,  # UID
                                sheet.cell(row=i, column=293).value,  # EMPCD
                                sheet.cell(row=i, column=294).value,  # ENAME
                                sheet.cell(row=i, column=295).value,  # DEPTID
                                sheet.cell(row=i, column=296).value,  # DFULLNAME
                                sheet.cell(row=i, column=297).value,  # POST
                                sheet.cell(row=i, column=298).value,  # TS
                                sheet.cell(row=i, column=300).value,  # IID --- (14)
                                sheet.cell(row=i, column=301).value,  # DOCNO
                                sheet.cell(row=i, column=302).value,  # FORMNAME
                                sheet.cell(row=i, column=303).value,  # CREATED
                                sheet.cell(row=i, column=304).value,  # ANAME
                                sheet.cell(row=i, column=305).value,  # UID
                                sheet.cell(row=i, column=306).value,  # EMPCD
                                sheet.cell(row=i, column=307).value,  # ENAME
                                sheet.cell(row=i, column=308).value,  # DEPTID
                                sheet.cell(row=i, column=309).value,  # DFULLNAME
                                sheet.cell(row=i, column=310).value,  # POST
                                sheet.cell(row=i, column=311).value,  # TS
                                sheet.cell(row=i, column=313).value,  # IID --- (15)
                                sheet.cell(row=i, column=314).value,  # DOCNO
                                sheet.cell(row=i, column=315).value,  # FORMNAME
                                sheet.cell(row=i, column=316).value,  # CREATED
                                sheet.cell(row=i, column=317).value,  # ANAME
                                sheet.cell(row=i, column=318).value,  # UID
                                sheet.cell(row=i, column=319).value,  # EMPCD
                                sheet.cell(row=i, column=320).value,  # ENAME
                                sheet.cell(row=i, column=321).value,  # DEPTID
                                sheet.cell(row=i, column=322).value,  # DFULLNAME
                                sheet.cell(row=i, column=323).value,  # POST
                                sheet.cell(row=i, column=324).value,  # TS
                                sheet.cell(row=i, column=326).value,  # IID --- (16)
                                sheet.cell(row=i, column=327).value,  # DOCNO
                                sheet.cell(row=i, column=328).value,  # FORMNAME
                                sheet.cell(row=i, column=329).value,  # CREATED
                                sheet.cell(row=i, column=330).value,  # ANAME
                                sheet.cell(row=i, column=331).value,  # UID
                                sheet.cell(row=i, column=332).value,  # EMPCD
                                sheet.cell(row=i, column=333).value,  # ENAME
                                sheet.cell(row=i, column=334).value,  # DEPTID
                                sheet.cell(row=i, column=335).value,  # DFULLNAME
                                sheet.cell(row=i, column=336).value,  # POST
                                sheet.cell(row=i, column=337).value  # TS
                                )

            # 引数: 左からテナントID、職員番号、パスワード
            gen = bt.getSignInInfo()
            # signin_procedure(gen.__next__(), gen.__next__(), gen.__next__())   <----- サインイン方法確定後、有効化せよ！！
            gen = None

            # 下記はテナントID直下の【サインイン】ボタン。上記サインイン手続き有効化時は無効化せよ！！
            click_element(
                "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
            )

            click_element(
                "/html/body/div[2]/div[1]/ul[1]/li[4]/a"
            )

            logger.info(  # --- ADD NEW 共通部品 20212.01.04
                "●【経費】出張申請書ラジオボタン選択"
            )
            click_element(
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[5]/label[2]/div/div[1]/input"
            )

            driver.implicitly_wait(5)       # 引数10を5へ減数

            click_element(
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[3]/button"
            )

            driver.implicitly_wait(5)       # 引数10を5へ減数
            driver.switch_to.window(driver.window_handles[1])
            # create_button = driver.find_element_by_xpath('//*[text()="起票する"]')

            create_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[3]/button[2]"
            )
            my_sleep_click(create_button)

            logger.info(  # --- ADD NEW 個別 20212.01.04
                "出張種別: " + str(bt.getBusinessTripType())
            )
            # TRIP_TYPE = "普通"                                    # --- DELETE 2021.12.03 前任者による定数エントリー 
            trip_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div[1]/select"
            )
            # my_send_keys(trip_field, TRIP_TYPE)                   # --- DELETE 2021.12.03
            my_send_keys(trip_field, bt.getBusinessTripType())      # --- UPDATE 2021.12.03

            # TRIP_DEPT = "東京"                                    # --- DELETE 2021.12.03 無用の為、無効化 - Start
            # trip_field = driver.find_element_by_xpath(
            #     "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div[2]/select"
            # )
            # my_send_keys(trip_field, TRIP_DEPT)
            #
            # TRIP_AREA = "北海道"
            # trip_area = driver.find_element_by_xpath(
            #     "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div[3]/select"
            # )
            # my_send_keys(trip_area, TRIP_AREA)                    # --- DELETE 2021.12.03 無用の為、無効化 - End

            logger.info(  # --- ADD NEW 個別 20212.01.04
                "所属: " + str(bt.getBusinessTripDepartment())
            )
            department_field = driver.find_element_by_xpath(                    # --- ADD NEW 2021.12.06 --- Start
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div[2]/select"
            )
            my_send_keys(department_field, bt.getBusinessTripDepartment())      # --- ADD NEW 2021.12.06 --- End

            logger.info(  # --- ADD NEW 個別 20212.01.04
                "出張地域: " + str(bt.getBusinessTripArea())
            )
            trip_area_field = driver.find_element_by_xpath(                     # --- ADD NEW 2021.12.06 --- Start
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div[3]/select"
            )
            my_send_keys(trip_area_field, bt.getBusinessTripArea())             # --- ADD NEW 2021.12.06 --- End

            # TRIP_START = "2021-08-01"                             # --- DELETE 2021.12.03 前任者による定数エントリー
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "出発日: " + str(bt.getBusinessTripStartDate())
            )
            trip_start_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div[" "1]/div/input"
            )
            # my_send_keys(trip_start_field, TRIP_START)            # --- DELETE 2021.12.03
            my_send_keys(trip_start_field, 
                            bt.getBusinessTripStartDate())          # --- UPDATE 2021.12.03

            # TRIP_END = "2021-08-10"                               # --- DELETE 2021.12.03 前任者による定数エントリー
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "帰着日: " + str(bt.getBusinessTripEndDate())
            )
            trip_end_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div[" "2]/div/input"
            )
            # my_send_keys(trip_end_field, TRIP_END)                # --- DELETE 2021.12.03
            my_send_keys(trip_end_field, bt.getBusinessTripEndDate())   # --- UPDATE 2021.12.03

            # --- ADD NEW 2021.12.06 ■出発時刻、■直帰区分、■帰社時刻、■直帰区分の４項目は前任者において作為されず、新規作成 --- Start
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "出発時刻: " + str(bt.getDepartureTime())
            )
            departure_time_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[4]/div[1]/input"
            )
            my_send_keys(departure_time_field, bt.getDepartureTime())

            go_straight_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[4]/div[2]/label/input[1]"
            )
            if bt.getGoStraight() == True:
                my_sleep_click(go_straight_field)
                logger.info(  # --- ADD NEW 個別 20212.01.04
                    "直行する"
                )
            else:
                logger.info(  # --- ADD NEW 個別 20212.01.04
                    "直行しない"
                )

            logger.info(  # --- ADD NEW 個別 20212.01.04
                "帰着時刻: " + str(bt.getReturnTime())
            )
            return_time_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[4]/div[3]/input"
            )
            my_send_keys(return_time_field, bt.getReturnTime())

            bounce_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[4]/div[4]/label/input[1]"
            )
            if bt.getBounce() == True:
                my_sleep_click(bounce_field)
                logger.info(  # --- ADD NEW 個別 20212.01.04
                    "直帰する"
                )
            else:
                logger.info(  # --- ADD NEW 個別 20212.01.04
                    "直帰しない"
                )
            # --- ADD NEW 2021.12.06 ■出発時刻、■直帰区分、■帰社時刻、■直帰区分の４項目は前任者において作為されず、新規作成 --- End

            # TRIP_BUSINESS_DAY = "08/09"                           # --- DELETE 2021.12.03 前任者による定数エントリー
            # 前任者は上記変数を用務日だと言いたいらしい。
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "用務日: " + str(bt.getBusinessDay())
            )
            trip_business_days_field = driver.find_element_by_xpath(
                "html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[" "5]/div[1]/input"
            )
            # my_send_keys(trip_business_days_field, TRIP_BUSINESS_DAY)     # --- DELETE 2021.12.03
            my_send_keys(trip_business_days_field, bt.getBusinessDay())     # --- UPDATE 2021.12.03

            # TRIP_BUSINESS_CONTACT = "山田様"                      # --- DELETE 2021.12.03 前任者による定数エントリー
            trip_business_contact_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[" "5]/div[2]/input"
            )
            # my_send_keys(trip_business_contact_field, TRIP_BUSINESS_CONTACT)      # --- DELETE 2021.12.03
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "用務先: " + str(bt.getBusinessDestination())
            )
            gen = bt.getBusinessDestination()
            my_send_keys(trip_business_contact_field,
                         (gen.__next__() + '・' + gen.__next__())
            )                                                                       # --- UPDATE 2021.12.03

            # TRIP_BUSINESS_DETAIL = "●の打合せ」「●の対応"                         # --- DELETE 2021.12.03 前任者による定数エントリー
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "用務内容: " + str(bt.getBusinessContact())
            )
            trip_business_contact_field = driver.find_element_by_xpath(
                "html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[" "6]/div[1]/input"
            )
            # my_send_keys(trip_business_contact_field, TRIP_BUSINESS_DETAIL)       # --- DELETE 2021.12.03
            my_send_keys(trip_business_contact_field, bt.getBusinessContact())

            # --- 2022.01.11 ADD NEW 備考枠を新設（今まで無かった!!） --- Start
            remark_textbox = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[7]/div/textarea"
            )
            my_send_keys(remark_textbox, bt.getApprovalInfo())      # --- 決裁情報
            # --- 2022.01.11 ADD NEW 備考枠を新設（今まで無かった!!） --- End

            attach_file_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[5]/div/ul/li[3]/a"
            )
            my_sleep_click(attach_file_button)

            # TRIP_ATTACH_FILE = "C:\Output\MFZ\Test01.pdf"        # --- DELETE 2021.12.03 前任者による定数エントリー

            gen = bt.getVoucherFiles()                             # Generatorへ格納し、next()で取得する
            submit_voucher_files(gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__()
            )
            gen = None

            time.sleep(2)
            confirm_button = driver.find_element_by_xpath(
                # "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"      # --- 2021.12.27 変更が加わったのか？
                "/html/body/div[1]/div[2]/div/form/div[14]/div[1]/ul/li[1]/button[2]"
            )
            my_sleep_click(confirm_button)
            driver.save_screenshot(f'C:/Users/digiworker_biz_02/PycharmProjects/pythonProject/ScreenShot/BV/{i}_確定ボタンクリック.png')

            bt = None   # オブジェクトへの参照を切離す

            # select_file_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div[2]/div/form/div[5]/div/div/div["
            #    "3]/div/div/div/div[2]/div[2]/div/span/button"
            # )
            # my_sleep_click(select_file_button)

            # upload_file_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[2]/div[2]"
            # )
            # upload_file_button.click()
            # time.sleep(4)
            # pyautogui.write(TRIP_ATTACH_FILE)         # --- 実フォルダーパス、ファイル名に対応できない処理 
            # pyautogui.press("enter")
            # time.sleep(4)

            # add_green_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[3]/div/a"
            # )

            # my_sleep_click(add_green_button)

            # upload_file_button.send_keys(TRIP_ATTACH_FILE)

            # confirm_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"
            # )

            # my_sleep_click(confirm_button)

            # Final submission

            # final_register_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div[2]/div/form/div[11]/button[1]"
            # )
            # click_submit_button(driver, logger, "Validation", final_register_button)

            # --- ADD NEW 2021.12.23 通知チェックボックスのレ点を解除
            notificationCheckBox = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/form/div[3]/label/input[1]"
            )
            my_sleep_click(notificationCheckBox)
            comment_field = driver.find_element_by_xpath(       # 本稼働時は無効化すること！！ --- START
                "/html/body/div[1]/div/div/form/textarea"
            )
            my_send_keys(comment_field, MIGRATION_COMMENT)      # 本稼働時は無効化すること！！ --- END

            final_submit_button = driver.find_element_by_xpath(     # ファイナル・サブミット・ボタン…どない和訳したいのや
                "/html/body/div[1]/div/div/form/div[4]/button[1]"
            )

            # TODO uncomment below after testing
            # click_main(driver, logger, 'Submission', final_register_button)   # <--- 前任者仕様を取消す
            logger.info(  # --- ADD NEW 個別 20212.01.04
                "実行ボタンクリック"
            )
            my_sleep_click(final_submit_button)  # <--- UPDATE 2021.12.23
            # --- 2022.01.05 ADD NEW 【申請】ボタンクリック後、却下される可能性がある為、スクリーンショットを取得、保存する
            driver.save_screenshot(f'C:/Users/digiworker_biz_02/PycharmProjects/pythonProject/ScreenShot/BV/{i}_申請ボタンクリック.png')

            # サインアウト処理
            sign_out_procedure()

            # ガーベジコレクター
            # if i % 100 == 0:  --- INVALID 2022.01.05
            gc.collect()
            # BVシートの次行へ移動
            i += 1

def submit_voucher_files(vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9): 
    logger.info(            # --- ADD NEW 共通部品 20212.01.04
        "添付ファイル - 選択ボタンクリック"
    )
    attach_file_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div[2]/div/form/div[5]/div/ul/li[3]/a"
    )
    my_sleep_click(attach_file_button)
    # 配列へ格納 --- ※証憑はPDFのみとは限らない
    va = [vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9]

    j = 0   # ゼロオリジン
    if va[j] != "N/A":  # 一つ目の配列がN/Aの場合は証憑がない
        while va[j] is not None and j < 10:
            if j == 0:  # 2回目以降は不要である為
                select_file_button = driver.find_element_by_xpath(
                # 添付ファイルタブに所在する添付ファイル-選択ボタン - 2021.11.15 コメント記入
                    "/html/body/div[1]/div[2]/div/form/div[5]/div/div/div["
                    "3]/div/div/div/div[2]/div[2]/div/span/button"
                )
                my_sleep_click(select_file_button)

            logger.info(  # --- ADD NEW 共通部品 20212.01.04
                "証憑ファイルパス: " + va[j]
            )
            time.sleep(1)
            upload_file_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[2]/div[2]"     # 次画面にある「添付ファイルを選択してください」→選択ボタンをクリック
            )
            upload_file_button.click()
            time.sleep(2)
            try:
                pyperclip.copy(va[j])                       # Clip Boardへコピーし、
            except:
                logger.info(  # --- ADD NEW 共通部品 20212.01.04
                    "クリップボードコピーできませんでした！！"
                )
            else:
                logger.info(  # --- ADD NEW 共通部品 20212.01.04
                    "クリップボードコピーできました。"
                )
            try:
                pg.hotkey('ctrl', 'v')                      # Pasteする。
            except:
                logger.info(  # --- ADD NEW 共通部品 20212.01.04
                    "クリップボードペーストできませんでした！！"
                )
            else:
                logger.info(  # --- ADD NEW 共通部品 20212.01.04
                    "クリップボードペーストできました。"
                )
            time.sleep(2)
            try:
                pg.press("enter")
            except:
                logger.info(  # --- ADD NEW 共通部品 20212.01.04
                    "証憑ファイルを添付できませんでした！！"
                )
            else:
                logger.info(  # --- ADD NEW 共通部品 20212.01.04
                    "証憑ファイルを添付できました。"
                )
            time.sleep(2)
            add_green_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[3]/div/a"
            )
            my_sleep_click(add_green_button)
            time.sleep(2)
            # 右隣セルの値へ移動
            j += 1
    else:
        logger.info(  # --- ADD NEW 共通部品 20212.01.04
            "対象となる証憑ファイルは無し"
        )

    if j != 0:  # --- ADD NEW 2021.12.24 --- 証憑が一つ以上あった場合のみ確定ボタンをクリック
        logger.info(  # --- ADD NEW 共通部品 20212.01.04
            "添付ファイルタブ - 確定ボタンクリック"
        )
        confirm_button = driver.find_element_by_xpath(
            "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"
        )
        my_sleep_click(confirm_button)

def sign_out_procedure():    # 2021.12.15
    # ①申請完了後の「閉じる」ボタンをクリック
    logger.info(  # --- ADD NEW 共通部品 20212.01.04
        "閉じるボタンクリック"
    )
    close_button = driver.find_element_by_xpath(
        # "/html/body/div[1]/div/div/form/div/button[4]"    # --- DELETE 2021.12.23
        # "/html/body/div[1]/div/div/form/div/button[3]"    # --- UPDATE 2021.12.23 --- これも違うのだという
        "/html/body/div[1]/div/div/form/div/button[2]"
    )
    close_button.click()    # 当該ソースのみ特殊。要注意

    driver.implicitly_wait(3)                           # --- ADD NEW 2021.12.22
    driver.switch_to.window(driver.window_handles[0])   # --- ADD NEW 2021.12.22 要素ゼロのウィンドウへ制御を戻す！

    # ②画面右上「サインアウト」ボタン（非常出口アイコン）をクリック
    logger.info(  # --- ADD NEW 共通部品 20212.01.04
        "サインアウトアイコンクリック"
    )
    sign_out_button = driver.find_element_by_xpath(
        "/html/body/div[2]/nav/div/div[2]/ul/li[6]/form/a"
    )
    sign_out_button.click()
    time.sleep(2)
    # ③【ログイン】ボタンをラストにクリック
    logger.info(  # --- ADD NEW 共通部品 20212.01.04
        "サインイン画面へ戻るボタンクリック"
    )
    sign_in_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    )
    sign_in_button.click()
    time.sleep(2)

def main():
    entry_procedure()
    logger.info("Robot completed")

main()