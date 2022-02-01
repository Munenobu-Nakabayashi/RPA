# 【経費】出張申請書
import logging
import pyautogui as pg  # --- UPDATE 2021.12.23
pg.FAILSAFE = False     # --- UPDATE 2021.12.23

from chrome_driver_dl import get_latest_driver
from common import *

# --- ADD 2021.12.03 Start
import openpyxl
import sys
import pyperclip
import gc
from datetime import datetime
import concurrent.futures
import re

from selenium import webdriver
driver = webdriver.Chrome(
    executable_path="C:/Users/digiworker_biz_02/PycharmProjects/pythonProject/Chrome/chromedriver.exe"
)
# --- ADD 2021.12.03 End

# Gets or creates a logger
logger = logging.getLogger("02")
# set log level
logger.setLevel(logging.INFO)
# define file handler and set formatter
file_handler = logging.FileHandler("../Logs/02_business_trip_application.log")
formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")
file_handler.setFormatter(formatter)
# add file handler to logger
logger.addHandler(file_handler)

# driver = driver = webdriver.Chrome(get_latest_driver())
driver.get(MFZC_URL)
# driver.fullscreen_window()

driver.refresh()
logger.info("Robot Started")

class BusinessTripApplication:
    def __init__(self,
                wkUserId,
                wkPassWd,
                wkBusinessTripType,
                wkBusinessTripDepartment,
                wkBusinessTripArea,
                wkBusinessTripStartDate,
                wkBusinessTripEndDate,
                wkDepartureTime,
                wkGoStraight,
                wkReturnTime,
                wkBounce,                       # 直帰
                wkBusinessDay,
                wkBusinessDestinationA,
                wkBusinessDestinationB,
                wkBusinessContent,
                wkRemark,
                wkExCardNo,
                wkVoucher0,
                wkVoucher1,
                wkVoucher2,
                wkVoucher3,
                wkVoucher4,
                wkVoucher5,
                wkVoucher6,
                wkVoucher7,
                wkVoucher8,
                wkVoucher9
    ):
        self.userID = wkUserId
        self.passWd = wkPassWd
        self.businessTripType = wkBusinessTripType
        self.businessTripDepartment = wkBusinessTripDepartment
        self.businessTripArea = wkBusinessTripArea
        self.businessTripStartDate = wkBusinessTripStartDate
        self.businessTripEndDate = wkBusinessTripEndDate
        self.departureTime = wkDepartureTime
        self.goStraight = wkGoStraight
        self.returnTime = wkReturnTime
        self.bounce = wkBounce
        self.businessDay = wkBusinessDay
        self.businessDestinationA = wkBusinessDestinationA
        self.businessDestinationB = wkBusinessDestinationB
        self.businessContent = wkBusinessContent
        self.remark = wkRemark
        self.exCardNo = wkExCardNo
        self.voucher0 = wkVoucher0
        self.voucher1 = wkVoucher1
        self.voucher2 = wkVoucher2
        self.voucher3 = wkVoucher3
        self.voucher4 = wkVoucher4
        self.voucher5 = wkVoucher5
        self.voucher6 = wkVoucher6
        self.voucher7 = wkVoucher7
        self.voucher8 = wkVoucher8
        self.voucher9 = wkVoucher9

    def getSignInInfo(self):
        yield 100
        yield self.userID
        yield self.passWd

    def getBusinessTripType(self):
        return self.businessTripType

    # --- UPDATE 2021.12.06 --- 東京本社と思料される所属以外はそのまま返し、東京本社内の部課は東京本社文字列を返す
    def getBusinessTripDepartment(self):
        if self.businessTripDepartment == '北海道営業所' \
            or self.businessTripDepartment == '東北営業所' \
            or self.businessTripDepartment == '東京本社' \
            or self.businessTripDepartment == '中部営業所' \
            or self.businessTripDepartment == '関西営業所' \
            or self.businessTripDepartment == '中四国営業所' \
            or self.businessTripDepartment == '九州営業所':
                return self.businessTripDepartment
        else:
            return '東京本社'

    # --- レコード数が少ないので、対象文字列の左側2文字または3文字で識別させる
    def getBusinessTripArea(self):
        mojiretsu = self.businessTripArea
        if mojiretsu[0:2] == '宮城' or mojiretsu[0:2] == '岩手':
            return '東北'
        elif mojiretsu[0:2] == '栃木' or mojiretsu[0:2] == '静岡' or mojiretsu[0:2] == '越後' or mojiretsu[0:1] == '御茶' or mojiretsu[0:2] == '東京':
            return '関東'
        elif mojiretsu[0:2] == '愛知' or mojiretsu[0:2] == '三重' or mojiretsu[0:2] == '浜松':
            return '中部'
        elif mojiretsu[0:2] == '大阪' or mojiretsu[0:2] == '京都':
            return '関西'
        elif mojiretsu[0:2] == '愛媛' or mojiretsu[0:2] == '鳥取' or mojiretsu[0:2] == '山口' or mojiretsu[0:2] == '広島':
            return '中四国'
        elif mojiretsu[0:2] == '熊本' or mojiretsu[0:3] == '鹿児島' or mojiretsu[0:2] == '宮崎' or mojiretsu[0:3] == '久留米' or mojiretsu[0:2] == '長崎' or mojiretsu[0:2] == '福岡' or mojiretsu[0:2] == '大分':
            return '九州'
        elif mojiretsu[0:2] == '沖縄':
            return '沖縄'
        else:
            return 'アジア・オセアニア'

    def getBusinessTripStartDate(self):
        if self.businessTripStartDate is None or self.businessTripStartDate == 0:
            today = datetime.now()
            return today.strftime('%Y-%m-%d')           # 本日日付を戻す
        else:
            try:
                wkDateTime = datetime.strptime((self.businessTripStartDate), "%Y/%m/%d %H:%M")
                # Excelセルから時分秒を取得して仕舞う為、当該時分秒を除外する処理を施す
                return datetime.strptime(str(self.businessTripStartDate).replace('-', ''), '%Y%m%d %H:%M:%S').strftime('%Y-%m-%d')
            except Exception as e:
                logger.info(e)  # 実行時例外を記録のみする
                today = datetime.now()
                return today.strftime('%Y-%m-%d')  # 本日日付を戻す
            finally:
                pass

    def getBusinessTripEndDate(self):
        if self.businessTripEndDate is None or self.businessTripEndDate == 0:
            today = datetime.now()
            return today.strftime('%Y-%m-%d')           # 本日日付を戻す
        else:
            try:
                wkDateTime = datetime.strptime((self.businessTripEndDate), "%Y/%m/%d %H:%M")
                # Excelセルから時分秒を取得して仕舞う為、当該時分秒を除外する処理を施す
                return datetime.strptime(str(self.businessTripEndDate).replace('-', ''), '%Y%m%d %H:%M:%S').strftime('%Y-%m-%d')
            except Exception as e:
                logger.info(e)  # 実行時例外を記録のみする
                today = datetime.now()
                return today.strftime('%Y-%m-%d')  # 本日日付を戻す
            finally:
                pass

    def getBusinessDay(self):
        if self.businessDay is None or self.businessDay == 0:
            today = datetime.now()
            return today.strftime('%Y-%m-%d')           # 本日日付を戻す
        else:
            # Excelセルから時分秒を取得して仕舞う為、当該時分秒を除外する処理を施す
            # return datetime.strptime(str(self.businessDay).replace('-',''), '%Y%m%d %H:%M:%S').strftime('%Y-%m-%d')   # --- INVALID 2021.12.23
            return str(self.businessDay) [0:49]   # --- REPRACE 2021.12.23。varchar2型と思料され、50文字上限

    def getDepartureTime(self):
        # return datetime.time(self.departureTime)  # 使用不可
        if str(self.departureTime) in ':' == True:  # --- ADD NEW 2021.12.23 --- 時刻でない値が入っていた場合の対応 --- Start
            return str(self.departureTime)[0:5]
        else:
            return '09:00'                           # --- ADD NEW 2021.12.23 --- 時刻でない値が入っていた場合の対応 --- End

    def getGoStraight(self):
        if self.goStraight == 'する':                 # 旧MFにおける当該区分は①「出社する」、②「出社しない」につき、
            # return True                            # 新MFZでは①「直行しない」・②「直行する」へ読替える
            return False                             # --- CHANGE 2021.12.16
        elif self.goStraight == 'しない' or self.goStraight is None:
            # return False                          # --- CHANGE 2021.12.16
            return True                             # --- CHANGE 2021.12.16

    def getReturnTime(self):
        if str(self.returnTime) in ':' == True:     # --- ADD NEW 2021.12.23 --- 時刻でない値が入っていた場合の対応 --- Start
            return str(self.returnTime)[0:5]
        else:
            return '18:00'                          # --- ADD NEW 2021.12.23 --- 時刻でない値が入っていた場合の対応 --- End

    def getBounce(self):
        if self.bounce == 'する':                     # 旧MFにおける当該区分は①「帰社する」、②「帰社しない」につき、
            # return True                            # 新MFZでは①「直帰しない」・②「直帰する」へ読替える
            return False
        elif self.bounce == 'しない' or self.bounce is None:
            # return False                          # --- CHANGE 2021.12.16
            return True                             # --- CHANGE 2021.12.16

    def getBusinessDestination(self):
        yield str(self.businessDestinationA)
        yield str(self.businessDestinationB)


    def getBusinessContact(self):
        return self.businessContent
    
    def getVoucherFiles(self):
        yield self.voucher0
        yield self.voucher1
        yield self.voucher2
        yield self.voucher3
        yield self.voucher4
        yield self.voucher5
        yield self.voucher6
        yield self.voucher7
        yield self.voucher8
        yield self.voucher9
    
def signin_procedure(tenantId, empId, passWd):
    # テナントID = 100
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[1]/input"
    )
    first_input.send_keys(tenantId)
    # 職員番号
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[2]/input"
    )
    first_input.send_keys(empId)
    # パスワード（※2021.11.17時点、方針未決）
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[3]/input"
    )
    first_input.send_keys(passWd)
    # ログインボタンをクリック
    click_element(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[5]/button"
    )

def click_element(web_el):
    my_sleep_click(driver.find_element_by_xpath(web_el))

def click_main(web_el):
    try:
        web_el.click()
        logger.info("Success")
    except NoSuchElementException:
        logger.error("Error")
    time.sleep(3)

def reloadBrowser():
    try:
        time.sleep(2)
        driver.refresh()
        time.sleep(2)
    except Exceptio as e:
        logger.info("Chromeブラウザーをリロードできず。ドライバーを確認せよ")
        # print("Chromeブラウザーをリロードできず。ドライバーを確認せよ")
        # driver.quit()
    finally:
        pass

def entry_procedure():

    reloadBrowser()
    # click_element(
    #     "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    # )

    # 交際費申請書の伝票タイプはBV
    book = openpyxl.load_workbook("C:\XLSX\XM_SLIP\BV.xlsx")    # Book名がBV
    sheet = book['BV']                                          # 出張申請書

    with concurrent.futures.ProcessPoolExecutor() as executor:
        i = 10   # 1行目は見出し行。2行目から開始。A列セルをNull判定しEOFを見極める。
        while sheet.cell(row=i, column=1).value is not None:
            bt = BusinessTripApplication(sheet.cell(row=i, column=9).value,     # CREATOR
                                '0Nu4M0%4N0',                                   # 共通パスワード（※未定）
                                sheet.cell(row=i, column=68).value,             # 出張種別【SS02】
                                sheet.cell(row=i, column=11).value,             # 出張者の所属【DEPTNAME】
                                sheet.cell(row=i, column=81).value,             # 出張先地域【SS15】（※納品先要確認）
                                sheet.cell(row=i, column=69).value,             # 出張期間（自）【SS03】
                                sheet.cell(row=i, column=70).value,             # 出張期間（至）【SS04】
                                sheet.cell(row=i, column=72).value,             # 出発時刻【SS06】
                                sheet.cell(row=i, column=71).value,             # 直行チェックボックス【SS05】（※未確定、納品先要確認）
                                sheet.cell(row=i, column=74).value,             # 帰着時刻【SS08】
                                sheet.cell(row=i, column=73).value,             # 直帰チェックボックス【SS07】（※未確定、納品先要確認）
                                sheet.cell(row=i, column=84).value,             # 用務日【SS18】
                                sheet.cell(row=i, column=81).value,             # 用務先A【★SS15】&【SS16】（※未確定、納品先要確認）
                                sheet.cell(row=i, column=82).value,             # 用務先B【SS15】&【★SS16】（※未確定、納品先要確認）
                                sheet.cell(row=i, column=83).value,             # 用務内容【SS17】
                                sheet.cell(row=i, column=86).value,             # 備考【SS20】
                                sheet.cell(row=i, column=256).value,            # EXカード番号（※未確定、納品先要確認）
                                sheet.cell(row=i, column=109).value,            # 証憑ファイル1
                                sheet.cell(row=i, column=110).value,            # 証憑ファイル2
                                sheet.cell(row=i, column=111).value,            # 証憑ファイル3
                                sheet.cell(row=i, column=112).value,            # 証憑ファイル4
                                sheet.cell(row=i, column=113).value,            # 証憑ファイル5
                                sheet.cell(row=i, column=114).value,            # 証憑ファイル6
                                sheet.cell(row=i, column=115).value,            # 証憑ファイル7
                                sheet.cell(row=i, column=116).value,            # 証憑ファイル8
                                sheet.cell(row=i, column=117).value,            # 証憑ファイル9
                                sheet.cell(row=i, column=118).value             # 証憑ファイル10
            )


            # 引数: 左からテナントID、職員番号、パスワード
            gen = bt.getSignInInfo()
            # signin_procedure(gen.__next__(), gen.__next__(), gen.__next__())   <----- サインイン方法確定後、有効化せよ！！
            gen = None

            # 下記はテナントID直下の【サインイン】ボタン。上記サインイン手続き有効化時は無効化せよ！！
            click_element(
                "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
            )

            click_element(
                "/html/body/div[2]/div[1]/ul[1]/li[4]/a"
            )

            click_element(
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[5]/label[2]/div/div[1]/input"
            )

            driver.implicitly_wait(5)       # 引数10を5へ減数

            click_element(
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[3]/button"
            )

            driver.implicitly_wait(5)       # 引数10を5へ減数
            driver.switch_to.window(driver.window_handles[1])
            # create_button = driver.find_element_by_xpath('//*[text()="起票する"]')

            create_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[3]/button[2]"
            )
            my_sleep_click(create_button)

            # TRIP_TYPE = "普通"                                    # --- DELETE 2021.12.03 前任者による定数エントリー 
            trip_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div[1]/select"
            )
            # my_send_keys(trip_field, TRIP_TYPE)                   # --- DELETE 2021.12.03
            my_send_keys(trip_field, bt.getBusinessTripType())      # --- UPDATE 2021.12.03

            # TRIP_DEPT = "東京"                                    # --- DELETE 2021.12.03 無用の為、無効化 - Start
            # trip_field = driver.find_element_by_xpath(
            #     "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div[2]/select"
            # )
            # my_send_keys(trip_field, TRIP_DEPT)
            #
            # TRIP_AREA = "北海道"
            # trip_area = driver.find_element_by_xpath(
            #     "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div[3]/select"
            # )
            # my_send_keys(trip_area, TRIP_AREA)                    # --- DELETE 2021.12.03 無用の為、無効化 - End

            department_field = driver.find_element_by_xpath(                    # --- ADD NEW 2021.12.06 --- Start
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div[2]/select"
            )
            my_send_keys(department_field, bt.getBusinessTripDepartment())      # --- ADD NEW 2021.12.06 --- End

            trip_area_field = driver.find_element_by_xpath(                     # --- ADD NEW 2021.12.06 --- Start
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div[3]/select"
            )
            my_send_keys(trip_area_field, bt.getBusinessTripArea())             # --- ADD NEW 2021.12.06 --- End

            # TRIP_START = "2021-08-01"                             # --- DELETE 2021.12.03 前任者による定数エントリー 
            trip_start_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div[" "1]/div/input"
            )
            # my_send_keys(trip_start_field, TRIP_START)            # --- DELETE 2021.12.03
            my_send_keys(trip_start_field, 
                            bt.getBusinessTripStartDate())          # --- UPDATE 2021.12.03

            # TRIP_END = "2021-08-10"                               # --- DELETE 2021.12.03 前任者による定数エントリー
            trip_end_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div[" "2]/div/input"
            )
            # my_send_keys(trip_end_field, TRIP_END)                # --- DELETE 2021.12.03
            my_send_keys(trip_end_field, bt.getBusinessTripEndDate())   # --- UPDATE 2021.12.03

            # --- ADD NEW 2021.12.06 ■出発時刻、■直帰区分、■帰社時刻、■直帰区分の４項目は前任者において作為されず、新規作成 --- Start
            departure_time_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[4]/div[1]/input"
            )
            my_send_keys(departure_time_field, bt.getDepartureTime())

            go_straight_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[4]/div[2]/label/input[1]"
            )
            if bt.getGoStraight() == True:
                my_sleep_click(go_straight_field)

            return_time_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[4]/div[3]/input"
            )
            my_send_keys(return_time_field, bt.getReturnTime())

            bounce_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[4]/div[4]/label/input[1]"
            )
            if bt.getBounce() == True:
                my_sleep_click(bounce_field)
            # --- ADD NEW 2021.12.06 ■出発時刻、■直帰区分、■帰社時刻、■直帰区分の４項目は前任者において作為されず、新規作成 --- End

            # TRIP_BUSINESS_DAY = "08/09"                           # --- DELETE 2021.12.03 前任者による定数エントリー
            # 前任者は上記変数を用務日だと言いたいらしい。
            trip_business_days_field = driver.find_element_by_xpath(
                "html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[" "5]/div[1]/input"
            )
            # my_send_keys(trip_business_days_field, TRIP_BUSINESS_DAY)     # --- DELETE 2021.12.03
            my_send_keys(trip_business_days_field, bt.getBusinessDay())     # --- UPDATE 2021.12.03

            # TRIP_BUSINESS_CONTACT = "山田様"                      # --- DELETE 2021.12.03 前任者による定数エントリー
            trip_business_contact_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[" "5]/div[2]/input"
            )
            # my_send_keys(trip_business_contact_field, TRIP_BUSINESS_CONTACT)      # --- DELETE 2021.12.03
            gen = bt.getBusinessDestination()
            my_send_keys(trip_business_contact_field,
                         (gen.__next__() + '・' + gen.__next__())
            )                                                                       # --- UPDATE 2021.12.03

            # TRIP_BUSINESS_DETAIL = "●の打合せ」「●の対応"                         # --- DELETE 2021.12.03 前任者による定数エントリー
            trip_business_contact_field = driver.find_element_by_xpath(
                "html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[" "6]/div[1]/input"
            )
            # my_send_keys(trip_business_contact_field, TRIP_BUSINESS_DETAIL)       # --- DELETE 2021.12.03
            my_send_keys(trip_business_contact_field, bt.getBusinessContact())

            attach_file_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[5]/div/ul/li[3]/a"
            )
            my_sleep_click(attach_file_button)

            # TRIP_ATTACH_FILE = "C:\Output\MFZ\Test01.pdf"        # --- DELETE 2021.12.03 前任者による定数エントリー

            gen = bt.getVoucherFiles()                             # Generatorへ格納し、next()で取得する
            submit_voucher_files(gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__()
            )
            gen = None

            time.sleep(2)
            confirm_button = driver.find_element_by_xpath(
                # "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"      # --- 2021.12.27 変更が加わったのか？
                "/html/body/div[1]/div[2]/div/form/div[14]/div[1]/ul/li[1]/button[2]"
            )
            my_sleep_click(confirm_button)

            bt = None   # オブジェクトへの参照を切離す

            # select_file_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div[2]/div/form/div[5]/div/div/div["
            #    "3]/div/div/div/div[2]/div[2]/div/span/button"
            # )
            # my_sleep_click(select_file_button)

            # upload_file_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[2]/div[2]"
            # )
            # upload_file_button.click()
            # time.sleep(4)
            # pyautogui.write(TRIP_ATTACH_FILE)         # --- 実フォルダーパス、ファイル名に対応できない処理 
            # pyautogui.press("enter")
            # time.sleep(4)

            # add_green_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[3]/div/a"
            # )

            # my_sleep_click(add_green_button)

            # upload_file_button.send_keys(TRIP_ATTACH_FILE)

            # confirm_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"
            # )

            # my_sleep_click(confirm_button)

            # Final submission

            # final_register_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div[2]/div/form/div[11]/button[1]"
            # )
            # click_submit_button(driver, logger, "Validation", final_register_button)

            # --- ADD NEW 2021.12.23 通知チェックボックスのレ点を解除
            notificationCheckBox = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/form/div[3]/label/input[1]"
            )
            my_sleep_click(notificationCheckBox)
            comment_field = driver.find_element_by_xpath(       # 本稼働時は無効化すること！！ --- START
                "/html/body/div[1]/div/div/form/textarea"
            )
            my_send_keys(comment_field, MIGRATION_COMMENT)      # 本稼働時は無効化すること！！ --- END

            final_submit_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/form/div[4]/button[1]"
            )

            # TODO uncomment below after testing
            # click_main(driver, logger, 'Submission', final_register_button)　
            my_sleep_click(final_submit_button)  # <--- UPDATE 2021.12.23

            # サインアウト処理
            sign_out_procedure()  # サインアウトを止めている

            # ガーベジコレクター
            if i % 100 == 0:
                gc.collect()
            # BVシートの次行へ移動
            i += 1

def submit_voucher_files(vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9): 
    # 【添付ファイル】タブをクリック
    attach_file_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div[2]/div/form/div[5]/div/ul/li[3]/a"
    )
    my_sleep_click(attach_file_button)
    # 配列へ格納 --- ※証憑はPDFのみとは限らない
    va = [vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9]

    j = 0   # ゼロオリジン
    if va[j] != "N/A":  # 一つ目の配列がN/Aの場合は証憑がない
        while va[j] is not None and j < 10:
            if j == 0:  # 2回目以降は不要である為
                select_file_button = driver.find_element_by_xpath(
                # 添付ファイルタブに所在する添付ファイル-選択ボタン - 2021.11.15 コメント記入
                    "/html/body/div[1]/div[2]/div/form/div[5]/div/div/div["
                    "3]/div/div/div/div[2]/div[2]/div/span/button"
                )
                my_sleep_click(select_file_button)

            upload_file_button = driver.find_element_by_xpath(
            # 次画面にある「添付ファイルを選択してください」→選択ボタンをクリック
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[2]/div[2]"
            )
            upload_file_button.click()
            time.sleep(2)
            pyperclip.copy(va[j])                      # Clip Boardへコピーし、
            pg.hotkey('ctrl', 'v')              # Pasteする。
            pg.press("enter")
            time.sleep(2)
            add_green_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[3]/div/a"
            )
            my_sleep_click(add_green_button)
            time.sleep(2)
            # 右隣セルの値へ移動
            j += 1

    if j != 0:  # --- ADD NEW 2021.12.24 --- 証憑が一つ以上あった場合のみ確定ボタンをクリック
        confirm_button = driver.find_element_by_xpath(
            "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"
        )
        my_sleep_click(confirm_button)

def signin_procedure(tenantId, empId, passWd):
    # テナントID = 100
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[1]/input"
    )
    first_input.send_keys(tenantId)
    # 職員コード
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[2]/input"
    )
    first_input.send_keys(empId)
    # パスワード（※2021.11.17時点、方針未決）
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[3]/input"
    )
    first_input.send_keys(passWd)
    # ログインボタンをクリック
    click_element(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[5]/button"
    )
    time.sleep(3)

def sign_out_procedure():    # 2021.12.15
    # ①申請完了後の「閉じる」ボタンをクリック
    close_button = driver.find_element_by_xpath(
        # "/html/body/div[1]/div/div/form/div/button[4]"    # --- DELETE 2021.12.23
        # "/html/body/div[1]/div/div/form/div/button[3]"    # --- UPDATE 2021.12.23 --- これも違うのだという
        "/html/body/div[1]/div/div/form/div/button[2]"
    )
    close_button.click()    # 当該ソースのみ特殊。要注意

    driver.implicitly_wait(3)                           # --- ADD NEW 2021.12.22
    driver.switch_to.window(driver.window_handles[0])   # --- ADD NEW 2021.12.22 要素ゼロのウィンドウへ制御を戻す！

    # ②画面右上「サインアウト」ボタン（非常出口アイコン）をクリック
    sign_out_button = driver.find_element_by_xpath(
        "/html/body/div[2]/nav/div/div[2]/ul/li[6]/form/a"
    )
    sign_out_button.click()
    time.sleep(2)
    # ③【ログイン】ボタンをラストにクリック
    sign_in_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    )
    sign_in_button.click()
    time.sleep(2)

def main():
    entry_procedure()
    logger.info("Robot completed")

main()