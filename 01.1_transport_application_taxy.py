# 【経費】交通費精算書（高速代、ガス代、駐車場代等）
import logging
# 2011.11.15 # pip install pandas, openpyxl（※於二号機）
# import pandas as pd   # 2021.11.17 Comment Out (Chandi)
import openpyxl         # 2021.11.17 (Nakabayashi)

import pyautogui as pg  # --- UPDATE 2021.12.23
pg.FAILSAFE = False     # --- UPDATE 2021.12.23
import pyperclip    # 2021.11.25 クリップボードコピー追加 (Nakabayashi)
import gc           # 2021.11.24 ガーベジコレクター追加 (Nakabayashi)

# import fareclassification as fc           # --- 左記は電車賃まわりの交通費。当該プログラムではインポートしない
import highwaytallclassification as htc     # --- ADD NEW 2021.12.28 高速代、ガス代、駐車場代等

import concurrent.futures
from datetime import datetime

from chrome_driver_dl import get_latest_driver
from common import *    # 註: common.pyを外部変数的に取込んでいる --- 2021.11.17

from selenium import webdriver      # 2021.11.21 ADD Start (Nakabayashi)
import department as dpt                          # --- ADD NEW 2021.12.26 支払依頼書における（負担部門）部署コードと同じ構造
driver = webdriver.Chrome(
    executable_path="C:/Users/digiworker_biz_02/PycharmProjects/pythonProject/Chrome/chromedriver.exe"
)                                   # 2021.11.21 ADD End (Nakabayashi)

# Gets or creates a logger          # Global宣言する必要が発生した為、
logger = logging.getLogger("01.1")    # 宣言場所を移動。 --- 2021.11.25 (Nakabayashi)

# TODO 1 ← Chandi

# 個々のユーザ認証の際にユーザID・PW入力の自動化の実装 ← Chandi
# TODO 2 ← Chandi
# ログアウト部分も ← Chandi

# TODO 3 ← Chandi
# 申請に該当するCSVファイルを指定し開く ← Chandi
# csv_file = open_csv('C:\Users\digiworker_biz_02\Desktop\mfz_input\sample_travel.csv') ← Chandi

# TODO 4 ← Chandi
# while row in csv_file.iterrows(): ← Chandi
# travel_application_dict = [[row[0], row[1], row[2], row[3], row[4]]] ← Chandi
#
# travel_application_dict = [["Digital Arts", "Business", "Osaka", "Tokyo", "1000"]] ← Chandi
# travel_data = travel_application_dict[0] ← Chandi

class VoucherPDFs:  # 2021.11.22 --- クラス化し対応 (Nakabayashi)
    def __init__(self,
                 wkvpdf0,
                 wkvpdf1,
                 wkvpdf2,
                 wkvpdf3,
                 wkvpdf4,
                 wkvpdf5,
                 wkvpdf6,
                 wkvpdf7,
                 wkvpdf8,
                 wkvpdf9
    ):
        self.vpdf0 = wkvpdf0
        self.vpdf1 = wkvpdf1
        self.vpdf2 = wkvpdf2
        self.vpdf3 = wkvpdf3
        self.vpdf4 = wkvpdf4
        self.vpdf5 = wkvpdf5
        self.vpdf6 = wkvpdf6
        self.vpdf7 = wkvpdf7
        self.vpdf8 = wkvpdf8
        self.vpdf9 = wkvpdf9

    # Setter（※使用しない）
    def setVpdf(self,
                wkvpdf0,
                wkvpdf1,
                wkvpdf2,
                wkvpdf3,
                wkvpdf4,
                wkvpdf5,
                wkvpdf6,
                wkvpdf7,
                wkvpdf8,
                wkvpdf9
    ):
        self.vpdf0 = wkvpdf0
        self.vpdf1 = wkvpdf1
        self.vpdf2 = wkvpdf2
        self.vpdf3 = wkvpdf3
        self.vpdf4 = wkvpdf4
        self.vpdf5 = wkvpdf5
        self.vpdf6 = wkvpdf6
        self.vpdf7 = wkvpdf7
        self.vpdf8 = wkvpdf8
        self.vpdf9 = wkvpdf9

    # Getter --- 引数に従って証憑ファイルパスを返す
    def getVpdf(self, num):
        if num == 0:
            return self.vpdf0
        elif num == 1:
            return self.vpdf1
        elif num == 2:
            return self.vpdf2
        elif num == 3:
            return self.vpdf3
        elif num == 4:
            return self.vpdf4
        elif num == 5:
            return self.vpdf5
        elif num == 6:
            return self.vpdf6
        elif num == 7:
            return self.vpdf7
        elif num == 8:
            return self.vpdf8
        elif num == 9:
            return self.vpdf9

# 初期処理 - 2021.11.15 (Nakabayashi)
def initiate_program():
    # driver = webdriver.Chrome(get_latest_driver())
    # commom.pyを参照せよ
    driver.get(MFZC_URL)
    driver.fullscreen_window()
    driver.refresh()

#   # Gets or creates a logger          # Global宣言する必要が発生した為、
#   logger = logging.getLogger("01")    # 場所を移動。 --- 2021.11.25 (Nakabayashi)
    # set log level
    logger.setLevel(logging.INFO)
    # define file handler and set formatter
    file_handler = logging.FileHandler("../Logs/01_transport_application.log")
    formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")
    file_handler.setFormatter(formatter)
    # add file handler to logger
    logger.addHandler(file_handler)

    logger.info("Robot started")

def click_element(web_el):
    my_sleep_click(driver.find_element_by_xpath(web_el))

def entry_procedure():
    # ---------- 2021.11.21 Change Start
    # click_element(
    #    "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    # )
    ### click_element("/html/body/div[2]/div[1]/ul[1]/li[4]/a")
    ### click_element(
    ###     "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[5]/label[1]/div/div[1]/input"
    ### )
    # ---------- 2021.11.21 Change End
    # 交通費精算ファイルであるT1.xlsxへアクセス - 2021.11.16
    book = openpyxl.load_workbook("C:\XLSX\XM_SLIP_DETAIL\T3.xlsx")     # Change 2021.12.27
    sheet = book['T3']
    # T1エクセルブックを2行目（1行目は見出し行）から読込み→OpenPyXlを使用する。Pandasは排除。
    i = 2
    while sheet.cell(row=i, column=1).value is not None:
        # ■申請者アカウントによるサインイン処理→CH列[86]]がEmpCD、CI列[87]]が略称(※h-kawadaなど)。なお社員コードはCH列[86]（※D000241など）
        passWord = '0Nu4M0%4N0' # ※共通パスワード決定後、定数部分を変更せよ <--- 無制限パスワード判明後変更
        # 引数: 左からテナントID、職員コード（CI列かCH列か？）、パスワード
        # signin_procedure(100, sheet.cell(row=i, column=86).value, passWord) <----- サインイン方法確定後、有効化せよ！！

        # --- 下記はテナントID直下の【ログイン】ボタンであり、検証時の仮置き。上記サインイン手続き有効化時は無効化せよ！！
        time.sleep(2)   # --- 2021.12.28 ADD NEW
        click_element(
            "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
        )

        # ホーム画面左ペイン - 起票ボタンクリック - 2021.11.15 コメント記入
        click_element(
            "/html/body/div[2]/div[1]/ul[1]/li[4]/a"
        )
        # 起票画面→右ペインテーブル - 経費-【経費】交通費精算の行をクリック（●選択） - 2021.11.15 コメント記入
        click_element(
            "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[5]/label[1]/div/div[1]/input"
        )
        driver.implicitly_wait(5)   # <--- 元は10であったが5へ減数(Nakabayashi) 2021.11.25
        # print(driver.current_url)   # Chandi
        # 次へボタンクリック - 2021.11.15 コメント記入
        click_element(
            "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[3]/button"
        )
        driver.implicitly_wait(5)  # <--- 元は10であったが5へ減数(Nakabayashi) 2021.11.25
        driver.switch_to.window(driver.window_handles[1])
        # create_button = driver.find_element_by_xpath('//*[text()="起票する"]') ← Chandi氏がCommentOutしたもの
        create_button = driver.find_element_by_xpath(
        # 現在位置【申請書選択 > 起票日選択 > 申請方法選択 > 申請書選択 > 起票確認】
        # 「新規起票」画面→「起票する」ボタンをクリック - 2021.11.15 コメント記入
            "/html/body/div[1]/div/form/div/div[3]/button[2]"
        )
        my_sleep_click(create_button)   # 当該関数の所為は[common.py]を参照
        
        # ■「【経費】交通費精算書」初期画面へ遷移→「明細種類」選択
        select_specification_kind()
        
        # ■項目エントリー→以下、引数順を説明。※但し2021.11.17現在において未FIXである。委細はT1.xlsxファイルを参照せよ
        # 1.行き先[AS]、2.要件[AT]、3.出発地[AO]、4.到着地[AP]、5.利用区分コード[BO]、6.料金[P]、7.特急料金[?]（？）、8.片道/往復[AG]、
        # 9.距離[?]（？）、10.税込金額[P]、11.消費税額[Q]、12.詳細経路[AU]、13.負担部門[CC]
        entry_items(sheet.cell(row=i, column=45).value,     # 1.行き先[AS]
                    sheet.cell(row=i, column=46).value,     # 2.要件[AT]
                    sheet.cell(row=i, column=41).value,     # 3.出発地[AO]
                    sheet.cell(row=i, column=42).value,     # 4.到着地[AP]
                    sheet.cell(row=i, column=67).value,     # 5.利用区分コード[AE]
                    sheet.cell(row=i, column=16).value,     # 6.料金[P]
                    sheet.cell(row=i, column=256).value,    # 7.特急料金[?]（？）←不明であるので256と仮置きした。
                    sheet.cell(row=i, column=33).value,     # 8.片道/往復[AG]
                    sheet.cell(row=i, column=256).value,    # 9.距離[?]（？）←不明であるので256と仮置きした。
                    sheet.cell(row=i, column=16).value,     # 10.税込金額[P]
                    sheet.cell(row=i, column=18).value,     # 11.消費税額[R]
                    sheet.cell(row=i, column=47).value,     # 12.詳細経路[AU]
                    sheet.cell(row=i, column=6).value,      # 13.負担部門コード[F]
                    sheet.cell(row=i, column=90).value,     # 14.PDFファイルパス(0)
                    sheet.cell(row=i, column=91).value,     # 15.PDFファイルパス(1)
                    sheet.cell(row=i, column=92).value,     # 16.PDFファイルパス(2)
                    sheet.cell(row=i, column=93).value,     # 17.PDFファイルパス(3)
                    sheet.cell(row=i, column=94).value,     # 18.PDFファイルパス(4)
                    sheet.cell(row=i, column=95).value,     # 19.PDFファイルパス(5)
                    sheet.cell(row=i, column=96).value,     # 20.PDFファイルパス(6)
                    sheet.cell(row=i, column=97).value,     # 21.PDFファイルパス(7)
                    sheet.cell(row=i, column=98).value,     # 22.PDFファイルパス(8)
                    sheet.cell(row=i, column=99).value,     # 23.PDFファイルパス(9)
                    sheet.cell(row=i, column=31).value      # 24.SS01[AE] <--- 備考テキストボックスへ代入する。（例: 高速道路通行料, ガソリン代, 駐車場代）
        )
        
        # サインアウト処理 --- 2021.11.22
        sign_out_procedure()  # サインアウトを止めている
        # ガーベジコレクター --- 2021.11.24
        if i % 100 == 0:
            gc.collect()
        # T1シートの次行へ移動
        i += 1

def signin_procedure(tenantId, empId, passWd):
    # テナントID = 100
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[1]/input"
    )
    first_input.send_keys(tenantId)
    # 職員コード
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[2]/input"
    )
    first_input.send_keys(empId)
    # パスワード（※2021.11.17時点、方針未決）
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[3]/input"
    )
    first_input.send_keys(passWd)
    # ログインボタンをクリック
    click_element(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[5]/button"
    )
    time.sleep(3)

def select_specification_kind():
    # TODO ← Chandi
    # transport_type もCSVの項目で受け取る ← Chandi
    # transport_type = "交通費-立替(実費精算)"    # 左記は電車賃等に係る交通費
    transport_type = "その他費用-立替(実費精算)"      # --- 2021.12.28 UPDATE --- ガス代、高速通行料等
    detail_type_select_field = driver.find_element_by_xpath(
    # 明細種類リストボックス - 2021.11.15 コメント記入
        "/html/body/div[1]/div[2]/div/form/div[7]/div/div[2]/div/div[" "2]/div[2]/select"
    )
    detail_type_select_field.send_keys(transport_type)
    # 「明細入力」ボタンへ改行キーを押下する所為 - 2021.11.18 コメント追加
    detail_type_select_field.send_keys(
        Keys.TAB
        + Keys.ENTER
    )

# Now Input fields from here # Chandi

def entry_items(destination,
                requirements,
                pointOfDeparture,
                pointOfDestination,
                usageClassificationCd,
                price,
                limitedExpressCharge,
                oneWayOrRoundtrip,
                distance,
                taxIncludedAmount,
                consumptionTax,
                detailedRoute,
                burdenDepartmentCode,
                pdf_voucher0,
                pdf_voucher1,
                pdf_voucher2,
                pdf_voucher3,
                pdf_voucher4,
                pdf_voucher5,
                pdf_voucher6,
                pdf_voucher7,
                pdf_voucher8,
                pdf_voucher9,
                remark          # --- 2028.12.28 ADD NEQ
    ):

    first_input = driver.find_element_by_xpath(     # first_inputは前任者による命名
        ### 行き先（顧客名・代理店名・セミナー名等） - 2021.11.15 コメント記入
        "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div["
        "2]/div[1]/input"
    )
    # first_input.send_keys(travel_data[0]) # Chndi
    first_input.send_keys(destination)  # Nakabayashi

    main_page = driver.current_window_handle
    first_input.send_keys(
        Keys.TAB
        # 用件（「●の打合せ」「●の対応」等） *テキストボックス - 2021.11.15 コメント記入
        # + travel_data[1]  # Chandi
        + requirements  # Nakabayashi
        + Keys.TAB
        # 出発地 *テキストボックス - 2021.11.15 コメント記入
        # + travel_data[2]  # Chandi
        + pointOfDeparture  # Nakabayashi
        + Keys.TAB
        # 到着地 *テキストボックス- 2021.11.15 コメント記入 
        # + travel_data[3]  # Chandi
        + pointOfDestination    # Nakabayashi
        + Keys.TAB
        + Keys.TAB      # Nakabayashi→【利用区分】エントリー画面展開ボタンへ移動
        + Keys.ENTER    # 改行キーを叩く
    )
    # TODO 利用区分もCSVから読み取り、選択する必要がある。 ← Chandi
    # Element を拾って、ポップアップ画面で利用区分コード入力。検索および選択が必要。別のロボでその検索、選択お処理しているので参考にしてください。 ← Chandi
    
    # 利用区分エントリー処理
    entry_usageClassification(usageClassificationCd, price)     # --- CHANGE 2021.12.26

    # iframeから元のフレームへ戻り、エントリーを続ける
    driver.switch_to.default_content()
    time.sleep(2)
    # 料金をエントリー ※注意: この時点で税抜金額と消費税が自動算定されるが、後のプロセスでエントリーし直す！
    price_entry = driver.find_element_by_xpath(
        "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[1]/input"
    )
    time.sleep(2)   # 必置 --- 2021.11.24
    price_entry.send_keys(int(price))      # ExceptionをThrowするのでCast。Python3はlong型がなく、int型しかないらしい！！
    price_entry.send_keys(
        Keys.TAB
    )
    # 特急料金（未FIX）をエントリー
    # limited_express_charge_entry = driver.find_element_by_xpath(
    #    "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[2]/input"
    # )
    # time.sleep(2)
    # if limitedExpressCharge is None \
    #    or limitedExpressCharge.isnumeric() == False:   # NoneとはPythonにおけるNullの事
    #        limitedExpressCharge = 0    # 2021.11.24
    # limited_express_charge_entry.send_keys(int(limitedExpressCharge))
    # time.sleep(2)

    # 片道/往復メニュー（※変数内は「片道」か「往復」という漢字二文字）
    # detail_type_select_field = driver.find_element_by_xpath(
    #    "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[3]/select"
    # )
    # if oneWayOrRoundtrip is not None \
    #    and oneWayOrRoundtrip != "":    # 2021.11.25（※スペースやNullである可能性は低い）
    #        detail_type_select_field.send_keys(str(oneWayOrRoundtrip))
    # time.sleep(2)

    # distance_entry = driver.find_element_by_xpath(
    #    "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[4]/div[4]/input"
    # )
    # time.sleep(1)
    # distance_entry.clear()      # 自動で値が入る為、clear()で削除
    # time.sleep(2)
    # if distance is None or distance.isnumeric != true:  # 2021.11.25
    #     distance_entry.send_keys(0)                     # 距離の取得元列は未詳（2021.11.25現在）
    # else:
    #     distance_entry.send_keys(int(distance))         # 数値型でCast
    # time.sleep(2)

    # 料金 *テキストボックスへ料金をエントリー後、Tabキーを打鍵→税込金額と消費税額は自動反映（Chandi仕様） - 2021.11.15 コメント記入
    # price_entry.send_keys(
    #    distance                # 距離
    #    + Keys.TAB
    #    + taxIncludedAmount     # 税込金額
    #    + Keys.TAB
    #    + consumptionTax        # 消費税
    #    + Keys.TAB
    #    + detailedRoute         # 経路詳細
    #    + Keys.TAB
    #    + Keys.TAB              # 負担部門ボタンへカーソルが移動
    #    + keys.ENTER            # 改行キーを打鍵
    # )     # 前任者仕様はエントリーする値を一切顧慮せず使用に堪えない為、全廃 --- 2021.11.25

    # details_route = driver.find_element_by_xpath(
    #    "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[6]/div/textarea"
    # )
    ### ■原版は(1)特急料金、(2)片道/往復 *、(3)距離（出発地～到着地）のエントリーを無視。
    ### 自動反映の(4)税込金額、(5)消費税額も然り - 2021.11.15 コメント記入
    ### 経路詳細テキストボックスへエントリー - 2021.11.15 コメント記入
    # details_route.send_keys("Test route")
    ### ■負担部門を処置せずスキップ、無視 - 2021.11.15 コメント記入
    taxIncludedAmount_entry = driver.find_element_by_xpath(
        "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[5]/div[1]/input"
    )
    taxIncludedAmount_entry.clear()  # 2021.11.25
    time.sleep(1)
    taxIncludedAmount_entry.send_keys(taxIncludedAmount)        # 税込金額の取得元列は未詳（2021.11.25現在）
    time.sleep(1)   # 税込金額右手の消費税はエントリー項目でなく自動算定
    taxIncludedAmount_entry.send_keys(
        Keys.TAB
    )

    consumptionTax_entry = driver.find_element_by_xpath(
        "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[5]/div[3]/input"
    )
    time.sleep(1)
    consumptionTax_entry.clear()        # 2021.11.25
    time.sleep(1)
    # if oneWayOrRoundtrip == "往復":               # --- UPDATE 2021.12.13 --- Start
    #    consumptionTax_entry.send_keys(consumptionTax * 2)    # 税込運賃が「往復」選択時に倍になる。税金はBCP取得時の値の値でよいか？
    #    # consumptionTax_entry.send_keys(consumptionTax)
    #else:
    #    consumptionTax_entry.send_keys(consumptionTax)
    # time.sleep(2)                               # --- UPDATE 2021.12.13 --- End
    consumptionTax_entry.send_keys(consumptionTax)  # --- MOVE 2021.12.28 --- 高速代、ガス代対応の為、場所移設
    time.sleep(1)

    detailedRoute_entry = driver.find_element_by_xpath(             # 備考枠である（電車賃側では詳細経路であった）
        "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[3]/div/div[6]/div/textarea"
    )
    crlf = '\r\n'   # CRLF改行
    # detailedRoute_entry.send_keys(str(detailedRoute))
    if detailedRoute is not None:
        detailedRoute_entry.send_keys(
            str(detailedRoute) + crlf + str(remark)     # --- 2021.12.28 ADD NEW 備考文字列を経路文字列の改行後に追記
        )
    else:
        detailedRoute_entry.send_keys(
            str(remark)                                 # --- 2021.12.28 ADD NEW 備考文字列のみ
        )
    time.sleep(2)
    detailedRoute_entry.send_keys(      # Add --- 2021.11.25
        Keys.TAB
        + Keys.TAB
        + Keys.ENTER    # 負担部門検索iframeウィンドウを開くボタンへ改行キーを打鍵
    )

    # 負担部門（※iframe使用）
    # select_burdenDepartmentCode(burdenDepartmentCode)                         # --- INVALID 2021.12.26
    select_burdenDepartmentCode(dpt.returnDepartment(burdenDepartmentCode))     # --- UPDATE 2021.12.26
    # iframeから元のフレームへ戻り、エントリーを続ける
    driver.switch_to.default_content()

    ### 確定ボタンクリック
    confirm_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div[2]/div/form/div[7]/div[2]/div[2]/div[5]/div/button[1]"
    ) 
    my_sleep_click(confirm_button)
    
    # 証憑PDFファイルをサブミットする（最大10を確保）
    submit_voucher_pdf_file(
        pdf_voucher0,
        pdf_voucher1,
        pdf_voucher2,
        pdf_voucher3,
        pdf_voucher4,
        pdf_voucher5,
        pdf_voucher6,
        pdf_voucher7,
        pdf_voucher8,
        pdf_voucher9
    )
    
    # 申請ボタンクリック --- 2021.11.22 UPDATE
    click_appeal_button()

def entry_usageClassification(wkUsageClassificationCd, wkPrice):
    main_page = driver.current_window_handle
    # 利用区分内「品目コード」へ利用区分コードをエントリー
    # print(driver.current_url) <--- Chandi
    iframe = driver.find_element_by_xpath('//*[@id="zk_iFrameDialog"]/iframe')
    driver.switch_to.frame(iframe)

    # gen = fc.returnFare(wkUsageClassificationCd, wkPrice)     # --- UPDATE 2021.12.26 yield戻り値をgeneratorへ格納
    gen = htc.returnFare(wkUsageClassificationCd)               # --- UPDATE 2021.12.28
    wkXpath = gen.__next__()        # テーブルの段数を格納
    wk2UsageClassificationCd = gen.__next__()   # 交通費区分を格納
    first_input = driver.find_element_by_xpath(
            "/html/body/div[1]/div/form/div/div[1]/input"
    )
    first_input.send_keys(
        wk2UsageClassificationCd.strip()    # --- UPDATE 2021.12.23 strip()はTRIMと同じ役割
    )
    # タブキーを三回（※品目名テキストボックス→軽減税率テキストボックス→検索ボタン）打鍵し、検索ボタンを改行キーで打鍵する
    first_input.send_keys(
        Keys.TAB
        + Keys.TAB
        + Keys.TAB
        + Keys.ENTER    # 検索ボタンへ改行キーを打鍵
    )
    time.sleep(3)   # 待たせないと例外をスローして止まる！ 要注意
    ### first_input.send_keys(
    ###    Keys.TAB      # すべて開くボタン上へカーソルが移動
    ### )
    ### time.sleep(3)
    ### first_input.send_keys(
    ###     Keys.ENTER    # すべて開くボタンへ改行キーを打鍵→テーブルが開く
    ### )
    click_open_all_button = driver.find_element_by_xpath(   # 2021.11.24 すべて開くボタンをクリック
        "/html/body/div[1]/div/div[1]/div[2]/button"
    )
    click_open_all_button.click()
    # 各コードにおいてテーブル上の[+]ボタンを何回クリック後テーブルをクリックするか、分類する（※分類不詳コード残存: 2021.11.19）→更改: 2021.12.23
    # 1. 一回のパターン（つまり二段目になる）
    if wkXpath == 2:                                                             # --- CONVERT 2021.12.26 --- Start
        click_fare_classification_table = driver.find_element_by_xpath(
            "/html/body/div[1]/div/div[1]/div[4]/div/div/label/div/div[1]"      # --- 変わらない様子
        )
    elif wkXpath == 3:                                                          # --- 使う事は絶無だが置いておく！！
        click_fare_classification_table = driver.find_element_by_xpath(
            "/html/body/div[1]/div/div[1]/div[4]/div/div/div/label/div/div[1]"
        )
    click_fare_classification_table.click()                                     # --- CONVERT 2021.12.26 --- End

def select_burdenDepartmentCode(wkBurdenDepartmentCode):
    main_page = driver.current_window_handle
    iframe = driver.find_element_by_xpath('//*[@id="zk_iFrameDialog"]/iframe')
    driver.switch_to.frame(iframe)
    # 負担部門コードをテキストボックスへエントリー
    entry_burdenDepartmentCode = driver.find_element_by_xpath(
        "/html/body/div[1]/div/form/div/div[1]/input"
    )
    # entry_burdenDepartmentCode.send_keys(wkBurdenDepartmentCode)      # --- INVALID 2021.12.26
    entry_burdenDepartmentCode.send_keys(                               # --- UPDATE 2021.12.26 共通部品化 Start
        dpt.returnDepartment(wkBurdenDepartmentCode)                    #
    )                                                                   # --- UPDATE 2021.12.26 共通部品化 End
    
    entry_burdenDepartmentCode.send_keys(
        Keys.TAB
        + Keys.TAB
        + Keys.ENTER        # 検索ボタンへ改行キーを打鍵
    )
    time.sleep(3)   # 時間差を設けないとExceptionをThrowされて仕舞う 2021.11.25
    click_burdenDepartment_table = driver.find_element_by_xpath(
        "/html/body/div[1]/div/div[1]/div[3]/label/div/div[1]"
    )
    click_burdenDepartment_table.click()

    driver.implicitly_wait(2)   # 引数10であったが過大と判断し2へ変更 2021.11.25
    time.sleep(2)               # 引数5であったが過大と判断し2へ変更 2021.11.25
    driver.implicitly_wait(2)   # 引数10であったが過大と判断し3へ変更 2021.11.25
### 別ウィンドウを開く - 2021.11.15 コメント記入
#    iframe = driver.find_element_by_xpath('//*[@id="zk_iFrameDialog"]/iframe')
#    driver.switch_to.frame(iframe)

# TODO ← Chandi
# 部門コードもCSVから検索して入力が必要。今表示の中でDACを選んでいる。 ← Chandi
### Chandi氏の意図は、4行目にある「通勤交通費」をクリックし、
### 続いて「通勤_開発部門（Web開発課除く）・カスタマーサポート課」を決打ちする事に有った
# select_travel_type = driver.find_element_by_xpath(
#    "/html/body/div[1]/div/div[1]/div[4]/div[4]"
# )
# select_travel_type.click()

# select_travel_details = driver.find_element_by_xpath(
#    "/html/body/div[1]/div/div[1]/div[4]/div[4]/div/label[1]/div"
# )

#    my_sleep_click(select_travel_details)   ← Chandi

# close_button = driver.find_element_by_xpath('/html/body/div[1]/div/div[2]/form/button')   ← Chandi
# my_sleep_click(close_button)  ← Chandi

#    driver.switch_to.default_content()      ← Chandi

def submit_voucher_pdf_file(voucher0,
                            voucher1,
                            voucher2,
                            voucher3,
                            voucher4,
                            voucher5,
                            voucher6,
                            voucher7,
                            voucher8,
                            voucher9
    ):
    # Attach File function ← Chandi
    # PDF証憑ファイルを添付。1部とは限らない。PDFとは限らない。最大10部。
    vouchers = VoucherPDFs(voucher0,
                           voucher1,
                           voucher2,
                           voucher3,
                           voucher4,
                           voucher5,
                           voucher6,
                           voucher7,
                           voucher8,
                           voucher9
    )    # Objectをnewする

    attach_file_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div[2]/div/form/div[5]/div/ul/li[3]/a"
    )
    my_sleep_click(attach_file_button)

    # 証憑ファイル複数対応 --- 2021.11.19
    j = 0
    if vouchers.getVpdf(j) != "N/A":     # --- ADD NEW 2021.12.13 --- 1件目が存在する事が前提
        while vouchers.getVpdf(j) is not None and j < 10:

            # TODO ← Chandi
            # Need to designate attachment file file path from CSV record ← Chandi
            # EXPENSE_SETTLE_FILE = "C:\Output\MFZ\Test01.pdf" # ←Chandi
            if j == 0:  # 2回目以降は不要である為
                select_file_button = driver.find_element_by_xpath(
                # 添付ファイルタブに所在する添付ファイル-選択ボタン - 2021.11.15 コメント記入
                    "/html/body/div[1]/div[2]/div/form/div[5]/div/div/div["
                    "3]/div/div/div/div[2]/div[2]/div/span/button"
                )
                my_sleep_click(select_file_button)

            upload_file_button = driver.find_element_by_xpath(
                # 次画面にある「添付ファイルを選択してください」→選択ボタンをクリック - 2021.11.15 コメント記入
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[2]/div[2]"
            )
            upload_file_button.click()
            time.sleep(2)   # 4を2へ減数
            # Windows File Dialogにおける処理 - 2021.11.15 コメント記入
            # pyautogui.write(EXPENSE_SETTLE_FILE)    # <--- 2Byte文字非対応につきChandi版を改変（アカンやろ）
            # ---------------------------------------- 2021.11.25 Change Start
            pyperclip.copy(vouchers.getVpdf(j))         # Clip Boardへコピーし、
            pg.hotkey('ctrl', 'v')               # Pasteする。
            # ---------------------------------------- 2021.11.25 Change End
            pg.press("enter")
            time.sleep(2)   # 4を2へ減数
            # (+)追加ボタンをクリック→「選択されたファイル」テーブルへ登録される - 2021.11.15 コメント記入
            add_green_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[3]/div/a"
            )
            my_sleep_click(add_green_button)
            time.sleep(2)
            # 右隣セルの値へ移動
            j += 1

    vouchers = None
    ### 確定ボタンをクリック - 2021.11.15 コメント記入
    if j != 0:  # --- ADD NEW 2021.12.24 --- 証憑が一つ以上あった場合のみ確定ボタンをクリック
        confirm_button = driver.find_element_by_xpath(
            "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"
        )
        my_sleep_click(confirm_button)

def click_appeal_button():      # UPDATE 2021.11.22
    # Final submission ← Chandi
    # 申請ボタンをクリック→次画面へ遷移 - 2021.11.15 コメント記入
    final_register_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div[2]/div/form/div[11]/button[1]"
    )

    click_submit_button(driver, logger, "Validation", final_register_button)

    comment_field = driver.find_element_by_xpath(
        "/html/body/div[1]/div/div/form/textarea"
    )
    comment_field.send_keys(MIGRATION_COMMENT)

    # --- ADD NEW 2021.12.23 通知チェックボックスのレ点を解除
    notificationCheckBox = driver.find_element_by_xpath(
        "/html/body/div[1]/div/div/form/div[3]/label/input[1]"
    )
    my_sleep_click(notificationCheckBox)

    # 「レ 実行」ボタンクリック - 2021.11.15 コメント記入
    final_submit_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div/div/form/div[4]/button[1]"
    )

    # click_main(driver, logger, 'Submission', final_submit_button)   # ※ 2021.11.25停止中！！
    final_submit_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div/div/form/div[4]/button[1]"
    )
    my_sleep_click(final_submit_button)  # <--- UPDATE 2021.12.23     # 左記を正とする

def sign_out_procedure():    # 2021.12.15
    # ①申請完了後の「閉じる」ボタンをクリック
    close_button = driver.find_element_by_xpath(
        # "/html/body/div[1]/div/div/form/div/button[4]"    # --- DELETE 2021.12.23
        "/html/body/div[1]/div/div/form/div/button[3]"      # --- UPDATE 2021.12.23
    )
    my_sleep_click(close_button)

    driver.implicitly_wait(3)                           # --- ADD NEW 2021.12.22
    driver.switch_to.window(driver.window_handles[0])   # --- ADD NEW 2021.12.22 要素ゼロのウィンドウへ制御を戻す！

    # ②画面右上「サインアウト」ボタン（非常出口アイコン）をクリック
    sign_out_button = driver.find_element_by_xpath(
        "/html/body/div[2]/nav/div/div[2]/ul/li[6]/form/a"
    )
    my_sleep_click(sign_out_button)
    time.sleep(2)
    # ③【ログイン】ボタンをラストにクリック
    sign_in_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    )
    my_sleep_click(sign_in_button)
    time.sleep(2)

def main():
    # 初期処理
    initiate_program()
    # エントリー手続き
    entry_procedure()
    logger.info("Robot completed")

main()

