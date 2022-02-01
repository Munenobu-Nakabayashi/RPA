# 【経費】仮払申請書
import logging
import pyautogui as pg      # --- UPDATE 2021.12.23
pg.FAILSAFE = False         # --- UPDATE 2021.12.23

from chrome_driver_dl import get_latest_driver
from common import *

# --- ADD 2021.12.07 Start
import openpyxl
import sys
import pyperclip
import gc
from datetime import datetime
import concurrent.futures
import re

from selenium import webdriver
driver = webdriver.Chrome(
    executable_path="C:/Users/digiworker_biz_02/PycharmProjects/pythonProject/Chrome/chromedriver.exe"
)
# --- ADD 2021.12.07 End

# Gets or creates a logger
logger = logging.getLogger("04")
# set log level
logger.setLevel(logging.INFO)
# define file handler and set formatter
file_handler = logging.FileHandler("../Logs/04_down_pay_application.log")
formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")
file_handler.setFormatter(formatter)
# add file handler to logger
logger.addHandler(file_handler)

logger.info("Started")

# driver = webdriver.Chrome(get_latest_driver())    # --- INVALID 2021.12.24
driver.get(MFZC_URL)
driver.fullscreen_window()
driver.refresh()

class ExpenseProvisionalPaymentApplication:
    def __init__(self,
                wkUserId,
                wkPassWd,
                wkPurposeOfUse,
                wkTemporaryPaymentType,
                wkTemporaryPaymentAmount,
                wkDesiredDateReceivingPayment,
                wkScheduledSettlementDate,
                wkRemark,
                wkVoucher0,
                wkVoucher1,
                wkVoucher2,
                wkVoucher3,
                wkVoucher4,
                wkVoucher5,
                wkVoucher6,
                wkVoucher7,
                wkVoucher8,
                wkVoucher9
    ):
        self.userId = wkUserId
        self.passWd = wkPassWd
        self.porposeOfUse = wkPurposeOfUse
        self.temporaryPaymentType = wkTemporaryPaymentType
        self.temporaryPaymentAmount = wkTemporaryPaymentAmount
        self.disiredDateReceivingPayment = wkDesiredDateReceivingPayment
        self.scheduledSettlementDate = wkScheduledSettlementDate
        self.remark = wkRemark
        self.voucher0 = wkVoucher0
        self.voucher1 = wkVoucher1
        self.voucher2 = wkVoucher2
        self.voucher3 = wkVoucher3
        self.voucher4 = wkVoucher4
        self.voucher5 = wkVoucher5
        self.voucher6 = wkVoucher6
        self.voucher7 = wkVoucher7
        self.voucher8 = wkVoucher8
        self.voucher9 = wkVoucher9

    def getSignInInfo(self):
        yield 100
        yield self.userID
        yield self.passWd

    def getPurposeOfUse(self):
        return self.porposeOfUse
    
    def getTemporaryPaymentType(self):
        if self.temporaryPaymentAmount == 1:    # 2021.12.08 --- 左記は仮設。納品先要確認。
            return "振込"
        else:
            return "現金"
    
    def getTemporaryPaymentAmount(self):
        return self.temporaryPaymentAmount
    
    def getDisiredDateReceivingPayment(self):
        return datetime.strptime(str(self.disiredDateReceivingPayment).replace('-',''), '%Y%m%d %H:%M:%S').strftime('%Y-%m-%d')
    
    def getScheduledSettlementDate(self):
        return datetime.strptime(str(self.scheduledSettlementDate).replace('-',''), '%Y%m%d %H:%M:%S').strftime('%Y-%m-%d')
    
    def getRemark(self):
        return self.remark
    
    def getVoucherFiles(self):
        yield self.voucher0
        yield self.voucher1
        yield self.voucher2
        yield self.voucher3
        yield self.voucher4
        yield self.voucher5
        yield self.voucher6
        yield self.voucher7
        yield self.voucher8
        yield self.voucher9

def click_element(web_el):
    my_sleep_click(driver.find_element_by_xpath(web_el))


def click_main(web_el):
    try:
        web_el.click()
        logger.info("Success")
    except NoSuchElementException:
        logger.error("Error")
    time.sleep(5)

def reloadBrowser():
    try:
        time.sleep(2)
        driver.refresh()
        time.sleep(2)
    except Exceptio as e:
        logger.info("Chromeブラウザーをリロードできず。ドライバーを確認せよ")
        # driver.quit()
    finally:
        pass

def entry_procedure():

    reloadBrowser()
    # click_element(                                                    # --- ループ内へ移動 --- 2021.12.24
    #    "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    # )
    # 【経費】仮払申請書の伝票タイプはEV
    book = openpyxl.load_workbook("C:\XLSX\XM_SLIP\EV.xlsx")    # Book名がBV
    sheet = book['EV']                                          # 仮払申請書

    with concurrent.futures.ProcessPoolExecutor() as executor:
        i = 2   # 1行目は見出し行。2行目から開始。A列セルをNull判定しEOFを見極める。
        while sheet.cell(row=i, column=1).value is not None:
            ep = ExpenseProvisionalPaymentApplication(
                                sheet.cell(row=i, column=9).value,              # CREATOR
                                '0Nu4M0%4N0',                                   # 共通パスワード（※未定）
                                sheet.cell(row=i, column=81).value,             # 利用目的【SS15】
                                sheet.cell(row=i, column=46).value,             # 仮払種別【PAY_TYPE】（※左記は未定、要確認。※1は振込、以外は現金で仮設）
                                sheet.cell(row=i, column=39).value,             # 仮払金額【TEMP_PRICE】
                                sheet.cell(row=i, column=71).value,             # 仮払受取希望日【SS05】（確証無し。WISH_DATEはゼロが入っている）
                                sheet.cell(row=i, column=71).value,             # 精算予定日【SS05】（確証無し。PAY_GEN_DATEはゼロが入っている）
                                sheet.cell(row=i, column=86).value,             # 備考【SS20】
                                sheet.cell(row=i, column=109).value,            # 証憑ファイル1
                                sheet.cell(row=i, column=110).value,            # 証憑ファイル2
                                sheet.cell(row=i, column=111).value,            # 証憑ファイル3
                                sheet.cell(row=i, column=112).value,            # 証憑ファイル4
                                sheet.cell(row=i, column=113).value,            # 証憑ファイル5
                                sheet.cell(row=i, column=114).value,            # 証憑ファイル6
                                sheet.cell(row=i, column=115).value,            # 証憑ファイル7
                                sheet.cell(row=i, column=116).value,            # 証憑ファイル8
                                sheet.cell(row=i, column=117).value,            # 証憑ファイル9
                                sheet.cell(row=i, column=118).value             # 証憑ファイル10
            )

            # 引数: 左からテナントID、職員番号、パスワード
            gen = ep.getSignInInfo()
            # signin_procedure(gen.__next__(), gen.__next__(), gen.__next__())   <----- サインイン方法確定後、有効化せよ！！
            gen = None

            click_element(                                                  # --- テナントID直下のサインインボタン。
                "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"    # --- 本稼働時（ID、共通PWでサインインする）は、
            )                                                               # --- 無効化せよ！！

            click_element(
                "/html/body/div[2]/div[1]/ul[1]/li[4]/a"
            )

            click_element(
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[5]/label[4]/div/div[1]/input"
            )

            driver.implicitly_wait(5)       # 引数10だったが5へ減数

            click_element(
                "/html/body/div[2]/div[2]/div[1]/div[1]/div[2]/form/div[3]/button"
            )

            driver.implicitly_wait(5)       # 引数10だったが5へ減数
            driver.switch_to.window(driver.window_handles[1])
            # create_button = driver.find_element_by_xpath('//*[text()="起票する"]')

            create_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/form/div/div[3]/button[2]"
            )
            my_sleep_click(create_button)

            # PURPOSE = "利用目的"                                              # --- DELETE 2021.12.07 前任者による定数エントリー
            purpose_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[2]/div/input"
            )
            # purpose_fld.send_keys(PURPOSE)                                    # --- DELETE 2021.12.08
            purpose_fld.send_keys(ep.getPurposeOfUse())                         # --- UPDATE 2021.12.08

            # --- ADD NEW 2021.12.08 前任者未処理項目につき追加 --- Start
            temp_pay_type_field = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div[1]/select"
            )
            temp_pay_type_field.send_keys(ep.getTemporaryPaymentType())
            # --- ADD NEW 2021.12.08 前任者未処理項目につき追加 --- Stop

            # 仮払金額
            # DOWN_PAY = "0"                                                    # --- DELETE 2021.12.08 前任者による定数エントリー
            down_pay_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div[2]/input"
            )
            # down_pay_fld.send_keys(DOWN_PAY)                                  # --- DELETE 2021.12.08
            down_pay_fld.send_keys(ep.getTemporaryPaymentAmount())              # --- UPDATE 2021.12.08

            # 仮払受取希望日
            # DOWN_PAY_EXPECT_DATE = "2021-08-02"                               # --- DELETE 2021.12.08 前任者による定数エントリー
            down_pay_expect_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div[3]/div/input"
            )
            # down_pay_expect_fld.send_keys(DOWN_PAY_EXPECT_DATE)               # --- DELETE 2021.12.08
            down_pay_expect_fld.send_keys(ep.getDisiredDateReceivingPayment())  # --- UPDATE 2021.12.08

            # 精算予定日
            # SETTLE_DATE = "2021-08-02"                                        # --- DELETE 2021.12.08 前任者による定数エントリー
            settle_date_fld = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[6]/div[2]/div/div[3]/div[4]/div/input"
            )
            # settle_date_fld.send_keys(SETTLE_DATE)                            # --- DELETE 2021.12.08
            settle_date_fld.send_keys(ep.getScheduledSettlementDate())          # --- UPDATE 2021.12.08

            # Attach File function                                              # --- INVALIDATION 2021.12.08 --- Start
            # attach_file_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div[2]/div/form/div[5]/div/ul/li[3]/a"
            # ) 
            # my_sleep_click(attach_file_button)
            # DOWN_PAY_ATTACH_FILE = "C:\Output\MFZ\Test01.pdf"

            # select_file_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div[2]/div/form/div[5]/div/div/div["
            #    "3]/div/div/div/div[2]/div[2]/div/span/button"
            # )

            # my_sleep_click(select_file_button)

            # upload_file_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[2]/div[2]"
            # )
            # upload_file_button.click()
            # time.sleep(4)
            # pyautogui.write(DOWN_PAY_ATTACH_FILE)
            # pyautogui.press("enter")
            # time.sleep(4)

            # add_green_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[3]/div/a"
            # )

            # my_sleep_click(add_green_button)

            # upload_file_button.send_keys(TRIP_ATTACH_FILE)

            # confirm_button = driver.find_element_by_xpath(
            #    "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"
            # )

            # my_sleep_click(confirm_button)                                # --- INVALIDATION 2021.12.08 --- END

            gen = ep.getVoucherFiles()                                      # Generatorへ格納し、next()で取得する
            submit_voucher_files(gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__(),
                                 gen.__next__()
            )

            gen = None

            final_register_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div[2]/div/form/div[11]/button[1]"
            )

            click_submit_button(driver, logger, "Validation", final_register_button)

            # --- ADD NEW 2021.12.23 通知チェックボックスのレ点を解除
            notificationCheckBox = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/form/div[3]/label/input[1]"
            )
            my_sleep_click(notificationCheckBox)

            comment_field = driver.find_element_by_xpath(       # --- CAUTION --- 実運用時は解除せよ！！ --- END
                "/html/body/div[1]/div/div/form/textarea"
            )
            comment_field.send_keys(MIGRATION_COMMENT)          # --- CAUTION --- 実運用時は解除せよ！！ --- END

            final_submit_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/form/div[4]/button[1]"
            )

            # TODO uncomment below after testing
            # click_main(driver, logger, 'Submission', final_register_button)   #
            my_sleep_click(final_submit_button)  # <--- UPDATE 2021.12.22

            # サインアウト処理
            sign_out_procedure()  # サインアウトを止めている

            ep = None

            # ガーベジコレクター
            if i % 100 == 0:
                gc.collect()
            # EPシートの次行へ移動
            i += 1

def submit_voucher_files(vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9): 
    # 【添付ファイル】タブをクリック
    attach_file_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div[2]/div/form/div[5]/div/ul/li[3]/a"
    )
    my_sleep_click(attach_file_button)
    # 配列へ格納 --- ※証憑はPDFのみとは限らない
    va = [vf0, vf1, vf2, vf3, vf4, vf5, vf6, vf7, vf8, vf9]

    j = 0   # ゼロオリジン
    if va[j] != "N/A":
        while va[j] is not None and j < 10:
            if j == 0:  # 2回目以降は不要である為
                select_file_button = driver.find_element_by_xpath(
                # 添付ファイルタブに所在する添付ファイル-選択ボタン - 2021.11.15 コメント記入
                    "/html/body/div[1]/div[2]/div/form/div[5]/div/div/div["
                    "3]/div/div/div/div[2]/div[2]/div/span/button"
                )
                my_sleep_click(select_file_button)

            upload_file_button = driver.find_element_by_xpath(
                # 次画面にある「添付ファイルを選択してください」→選択ボタンをクリック
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[2]/div[2]"
            )
            upload_file_button.click()
            time.sleep(2)
            pyperclip.copy(va[j])                      # Clip Boardへコピーし、
            pg.hotkey('ctrl', 'v')              # Pasteする。
            pg.press("enter")
            time.sleep(2)
            add_green_button = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div[2]/form/div[2]/div[1]/div[3]/div/a"
            )
            my_sleep_click(add_green_button)
            time.sleep(2)
            # 右隣セルの値へ移動
            j += 1

    # --- MOVE 2021.12.09 --- 証憑ファイル最終確定ボタンクリックである為、此処へ配置する事！
    if j != 0:  # --- ADD NEW 2021.12.24 --- 証憑が一つ以上あった場合のみ確定ボタンをクリック
        confirm_button = driver.find_element_by_xpath(
            "/html/body/div[1]/div/div[2]/form/div[4]/button[1]"
        )
        my_sleep_click(confirm_button)

def signin_procedure(tenantId, empId, passWd):
    # テナントID = 100
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[1]/input"
    )
    first_input.send_keys(tenantId)
    # 職員コード
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[2]/input"
    )
    first_input.send_keys(empId)
    # パスワード（※2021.11.17時点、方針未決）
    first_input = driver.find_element_by_xpath(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[3]/input"
    )
    first_input.send_keys(passWd)
    # ログインボタンをクリック
    click_element(
        "/html/body/div[1]/div[1]/form/div[1]/div[2]/div[5]/button"
    )
    time.sleep(3)

def sign_out_procedure():    # 2021.12.15
    # ①申請完了後の「閉じる」ボタンをクリック
    close_button = driver.find_element_by_xpath(
        # "/html/body/div[1]/div/div/form/div/button[4]"    # --- INVALID 2021.12.24
        "/html/body/div[1]/div/div/form/div/button[2]"      # --- VALID 2021.12.24
    )
    my_sleep_click(close_button)

    driver.implicitly_wait(3)                           # --- ADD NEW 2021.12.22
    driver.switch_to.window(driver.window_handles[0])   # --- ADD NEW 2021.12.22 要素ゼロのウィンドウへ制御を戻す！

    # ②画面右上「サインアウト」ボタン（非常出口アイコン）をクリック
    sign_out_button = driver.find_element_by_xpath(
        "/html/body/div[2]/nav/div/div[2]/ul/li[6]/form/a"
    )
    my_sleep_click(sign_out_button)
    time.sleep(2)
    # ③【ログイン】ボタンをラストにクリック
    sign_in_button = driver.find_element_by_xpath(
        "/html/body/div[1]/div/form/div[1]/div[2]/div[2]/button"
    )
    my_sleep_click(sign_in_button)
    time.sleep(2)

def main():
    entry_procedure()
    logger.info("Robot completed")

main()
